"""Archive builder — generates XLSX + ZIP for a closed accident and stores in object storage.

Called as a background task after an accident is closed.  The resulting archive is
attached to the Accident record (archive_object_key) and an AccidentArchive row is
created with metadata.
"""

from __future__ import annotations

import json
import logging
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import select

from ..database import SessionLocal
from ..models import Accident, AccidentArchive, AccidentVideoUpload
from .accident_numbering import format_accident_number
from .accident_situation_table import build_situation_rows
from .admin_updates import notify_admin_data_changed
from .object_storage import _local_root, _make_boto3_client, _use_remote, upload_stream
from .time_utils import now_sgt

_logger = logging.getLogger(__name__)

# Column order in the XLSX (A=1, B=2, …)
# A  Horário
# B  Atividade/Local
# C  Nome
# D  Chave
# E  Projetos
# F  Local
# G  Ciência
# H  Zona de
# I  Situação
# J  Contato
# K  Registros

COLUMN_ORDER = [
    "Horário",
    "Atividade/Local",
    "Nome",
    "Chave",
    "Projetos",
    "Local",
    "Ciência",
    "Zona de",
    "Situação",
    "Contato",
    "Registros",
]

_COL_REGISTROS = len(COLUMN_ORDER)  # 1-based index of "Registros" column


def _slugify(value: str) -> str:
    """Convert an arbitrary string to a safe filename segment."""
    return re.sub(r"[^A-Za-z0-9_-]+", "_", value)[:60]


@dataclass(frozen=True)
class _AccidentFacts:
    """The Accident columns the archive needs, detached from the ORM session.

    Holding an ORM instance would force the DB session to stay open across the
    network I/O below just to keep it loadable.
    """

    id: int
    accident_number: int
    project_name_snapshot: str
    location_name_snapshot: str
    opened_at: datetime | None
    description: str


def _build_xlsx(
    accident: _AccidentFacts,
    snapshot_rows,
    video_files_by_user_chave: dict[str, list[str]],
) -> BytesIO:
    """Build the 'Situação de Pessoal' spreadsheet and return it as a BytesIO."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Situacao de Pessoal"

    # Metadata header rows
    bold = Font(bold=True)
    header_rows = [
        (f"Acidente N.º: {format_accident_number(accident.accident_number)}", ""),
        (f"Projeto: {accident.project_name_snapshot}", ""),
        (f"Local: {accident.location_name_snapshot}", ""),
        (f"Data abertura: {accident.opened_at.strftime('%d/%m/%Y %H:%M') if accident.opened_at else ''}", ""),
        (f"Descrição: {accident.description or '(sem descrição)'}", ""),
        ("", ""),  # blank separator
    ]
    for label, _ in header_rows:
        ws.append([label])
        ws.cell(row=ws.max_row, column=1).font = bold

    # Column headers
    ws.append(COLUMN_ORDER)
    for col_idx in range(1, len(COLUMN_ORDER) + 1):
        ws.cell(row=ws.max_row, column=col_idx).font = bold

    for row in snapshot_rows:
        user_chave = row.chave
        videos = video_files_by_user_chave.get(user_chave, [])
        registros_text = "\n".join(f"Registros/{user_chave}/{filename}" for filename in videos)

        ciencia = "Ciente" if row.awareness_status == "acknowledged" else "Aguardando"

        ws.append([
            row.event_time.isoformat(),
            row.activity_local or "",
            row.name,
            row.chave,
            ", ".join(row.projects),
            row.local or "",
            ciencia,
            row.zone,
            row.status,
            row.phone or "",
            registros_text,
        ])
        cell = ws.cell(row=ws.max_row, column=_COL_REGISTROS)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if videos:
            cell.hyperlink = f"Registros/{user_chave}/{videos[0]}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _read_video_bytes(object_key: str, client=None) -> bytes:
    """Fetch raw video bytes from storage (local or remote).

    Accepts a pre-built boto3 client so the caller reuses one across every video
    instead of constructing (and re-resolving credentials for) one per file.
    Returns b"" when the object is missing so one lost video cannot sink the whole
    archive — the spreadsheet still lists it.
    """
    if _use_remote():
        from ..core.config import settings
        from .object_storage import _make_boto3_client

        client = client or _make_boto3_client()
        try:
            result = client.get_object(Bucket=settings.do_spaces_bucket, Key=object_key)
            return result["Body"].read()
        except Exception:
            _logger.warning("Accident archive: could not read video %s", object_key, exc_info=True)
            return b""

    target = _local_root() / object_key
    return target.read_bytes() if target.exists() else b""


def build_and_attach_archive_for_accident(accident_id: int) -> None:
    """Build XLSX + ZIP archive for *accident_id*, upload to storage, persist metadata.

    Runs as a background task after an accident is closed, so nothing is waiting on
    it and an escaping exception would be swallowed by the task runner with the
    accident silently left without an archive. Log it here instead.
    """
    try:
        _build_and_attach_archive(accident_id)
    except Exception:
        _logger.exception(
            "Failed to build accident archive for accident_id=%s — the accident is "
            "closed but has no downloadable archive",
            accident_id,
        )


def _build_and_attach_archive(accident_id: int) -> None:
    # Phase 1 — read everything the build needs, then let the DB session go.
    # The build downloads every video and uploads two objects; holding the session
    # open across that kept a pooled connection idle-in-transaction for as long as
    # the network took, which is how this project has already exhausted its pool.
    with SessionLocal() as db:
        accident_row = db.get(Accident, accident_id)
        if accident_row is None:
            return
        facts = _AccidentFacts(
            id=accident_row.id,
            accident_number=accident_row.accident_number,
            project_name_snapshot=accident_row.project_name_snapshot,
            location_name_snapshot=accident_row.location_name_snapshot,
            opened_at=accident_row.opened_at,
            description=accident_row.description or "",
        )
        snapshot_rows = build_situation_rows(db, accident=accident_row)
        chave_by_user_id: dict[int, str] = {row.user_id: row.chave for row in snapshot_rows}

        videos = (
            db.execute(
                select(AccidentVideoUpload).where(
                    AccidentVideoUpload.accident_id == facts.id
                )
            )
            .scalars()
            .all()
        )
        # Detach to plain tuples so nothing below re-reads the ORM.
        video_facts = [
            (v.user_id, v.object_key, v.content_type, v.idempotency_key)
            for v in videos
        ]

    # Phase 2 — no DB session held from here until the very end.
    videos_by_user: dict[int, list[tuple[int, str, str, str]]] = {}
    for video in video_facts:
        videos_by_user.setdefault(video[0], []).append(video)

    video_files_by_user_chave: dict[str, list[str]] = {}
    video_payloads: dict[str, bytes] = {}  # key = full zip path
    storage_client = _make_boto3_client() if _use_remote() else None

    for user_id, user_videos in videos_by_user.items():
        user_chave = chave_by_user_id.get(user_id) or str(user_id)
        for idx, (_, object_key, content_type, idempotency_key) in enumerate(user_videos, start=1):
            ext = content_type.split("/")[-1]
            if ext == "quicktime":
                ext = "mov"
            filename = f"{idx:02d}_{_slugify(idempotency_key)}.{ext}"
            zip_path = f"Registros/{user_chave}/{filename}"
            video_files_by_user_chave.setdefault(user_chave, []).append(filename)
            video_payloads[zip_path] = _read_video_bytes(object_key, storage_client)

    xlsx_buffer = _build_xlsx(facts, snapshot_rows, video_files_by_user_chave)

    # Build ZIP
    zip_buffer = BytesIO()
    acc_label = format_accident_number(facts.accident_number)
    xlsx_name = f"{acc_label}.xlsx"
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(xlsx_name, xlsx_buffer.getvalue())
        for zip_path, payload in video_payloads.items():
            zf.writestr(zip_path, payload)
    zip_buffer.seek(0)

    xlsx_key = f"accidents/{acc_label}/archive/{xlsx_name}"
    zip_key = f"accidents/{acc_label}/archive/{acc_label}.zip"

    upload_stream(
        object_key=xlsx_key,
        stream=BytesIO(xlsx_buffer.getvalue()),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    upload_stream(
        object_key=zip_key,
        stream=zip_buffer,
        content_type="application/zip",
    )

    size_bytes = zip_buffer.seek(0, 2) or 0
    zip_buffer.seek(0)

    # Phase 3 — short write session.
    with SessionLocal() as db:
        accident_row = db.get(Accident, accident_id)
        if accident_row is None:
            return
        existing = db.execute(
            select(AccidentArchive).where(AccidentArchive.accident_id == accident_id)
        ).scalars().first()
        if existing is None:
            db.add(AccidentArchive(
                accident_id=accident_id,
                snapshot_json=json.dumps(
                    [row.model_dump() for row in snapshot_rows], default=str
                ),
                xlsx_object_key=xlsx_key,
                zip_object_key=zip_key,
                size_bytes=size_bytes,
                generated_at=now_sgt(),
            ))
        else:
            # Re-running the builder (e.g. after fixing a failure) must not trip
            # uq_accident_archives_accident_id.
            existing.snapshot_json = json.dumps(
                [row.model_dump() for row in snapshot_rows], default=str
            )
            existing.xlsx_object_key = xlsx_key
            existing.zip_object_key = zip_key
            existing.size_bytes = size_bytes
            existing.generated_at = now_sgt()
        accident_row.archive_object_key = zip_key
        db.commit()

    notify_admin_data_changed(
        "accident_closed",
        metadata={"accident_id": accident_id, "archive_ready": True},
    )
