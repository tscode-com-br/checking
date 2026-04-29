from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..core.config import settings
from ..models import TransportAssignment, TransportRequest, User, Vehicle
from ..schemas import TransportOperationalProposal, TransportOperationalSnapshot
from .time_utils import now_sgt


_PAIRED_ROUTE_KIND = {
    "home_to_work": "work_to_home",
    "work_to_home": "home_to_work",
}


def _build_transport_export_file_name(timestamp: datetime) -> str:
    return f"Transport List - {timestamp:%Y%m%d - %H%M%S}.xlsx"


def _build_transport_operational_plan_file_name(timestamp: datetime) -> str:
    return f"Transport Operational Plan - {timestamp:%Y%m%d - %H%M%S}.xlsx"


def _resolve_transport_export_path(file_name: str) -> Path:
    export_dir = Path(settings.transport_exports_dir)
    export_dir.mkdir(parents=True, exist_ok=True)

    candidate = export_dir / file_name
    if not candidate.exists():
        return candidate

    counter = 2
    while True:
        deduplicated = export_dir / f"{candidate.stem} ({counter}){candidate.suffix}"
        if not deduplicated.exists():
            return deduplicated
        counter += 1


def build_transport_list_export(
    db: Session,
    *,
    service_date: date,
    selected_route_kind: str,
) -> tuple[str, bytes]:
    return build_transport_operational_plan_export(
        db,
        service_date=service_date,
        selected_route_kind=selected_route_kind,
        proposal=None,
    )


def _collect_transport_list_rows(
    db: Session,
    *,
    service_date: date,
    selected_route_kind: str,
) -> list[list[str | None]]:
    from .transport_vehicle_operations import find_transport_vehicle_schedule

    route_priority = {
        selected_route_kind: 0,
        _PAIRED_ROUTE_KIND.get(selected_route_kind, "work_to_home"): 1,
    }

    exported_rows: list[tuple[tuple[str, str, int], list[str | None]]] = []
    assignments = db.execute(
        select(TransportAssignment, TransportRequest, User, Vehicle)
        .join(TransportRequest, TransportRequest.id == TransportAssignment.request_id)
        .join(User, User.id == TransportRequest.user_id)
        .join(Vehicle, Vehicle.id == TransportAssignment.vehicle_id)
        .where(
            TransportAssignment.service_date == service_date,
            TransportAssignment.status == "confirmed",
            TransportAssignment.vehicle_id.is_not(None),
        )
    ).all()

    assignments_by_request: dict[int, tuple[int, TransportAssignment, User, Vehicle]] = {}
    for assignment, transport_request, user, vehicle in assignments:
        candidate_priority = route_priority.get(assignment.route_kind, 2)
        current = assignments_by_request.get(transport_request.id)
        if current is None or candidate_priority < current[0]:
            assignments_by_request[transport_request.id] = (candidate_priority, assignment, user, vehicle)

    for _priority, assignment, user, vehicle in assignments_by_request.values():
        schedule = find_transport_vehicle_schedule(
            db,
            vehicle=vehicle,
            service_date=service_date,
            route_kind=assignment.route_kind,
        )
        exported_rows.append(
            (
                (
                    user.nome.lower(),
                    user.chave,
                    assignment.id,
                ),
                [
                    user.nome,
                    user.chave,
                    user.projeto,
                    user.end_rua,
                    service_date.isoformat(),
                    schedule.departure_time if schedule is not None else None,
                ],
            )
        )

    return [row for _, row in sorted(exported_rows, key=lambda item: item[0])]


def _snapshot_request_rows(snapshot: TransportOperationalSnapshot):
    return snapshot.regular_requests + snapshot.weekend_requests + snapshot.extra_requests


def _snapshot_vehicle_registry_rows(snapshot: TransportOperationalSnapshot):
    return (
        snapshot.regular_vehicle_registry
        + snapshot.weekend_vehicle_registry
        + snapshot.extra_vehicle_registry
    )


def _append_sheet_rows(worksheet, rows: list[list[object]]) -> None:
    for row in rows:
        worksheet.append(row)


def _build_executive_summary_rows(
    *,
    snapshot: TransportOperationalSnapshot,
    generated_at: datetime,
    proposal: TransportOperationalProposal | None,
) -> list[list[object]]:
    request_status_counts = Counter(request.assignment_status for request in _snapshot_request_rows(snapshot))
    summary_rows: list[list[object]] = [
        ["Campo/Field", "Valor/Value"],
        ["Modo/Mode", "proposal_review" if proposal is not None else "current_state"],
        ["Gerado em/Generated At", generated_at.isoformat()],
        ["Data/Service Date", snapshot.service_date.isoformat()],
        ["Rota/Route", snapshot.route_kind],
        ["Saida Work to Home/Work to Home Departure", snapshot.work_to_home_departure_time],
        ["Total Requests", len(_snapshot_request_rows(snapshot))],
        ["Pending Requests", request_status_counts.get("pending", 0)],
        ["Confirmed Requests", request_status_counts.get("confirmed", 0)],
        ["Rejected Requests", request_status_counts.get("rejected", 0)],
        ["Cancelled Requests", request_status_counts.get("cancelled", 0)],
        ["Total Vehicles", len(_snapshot_vehicle_registry_rows(snapshot))],
    ]

    if proposal is None:
        return summary_rows

    summary_rows.extend(
        [
            ["Proposal Key", proposal.proposal_key],
            ["Proposal Status", proposal.proposal_status],
            ["Proposal Origin", proposal.origin],
            ["Proposal Created At", proposal.created_at.isoformat()],
            ["Proposal Expires At", proposal.expires_at.isoformat() if proposal.expires_at else None],
            ["Total Decisions", proposal.summary.total_decisions],
            ["Confirmed Decisions", proposal.summary.confirmed_decisions],
            ["Rejected Decisions", proposal.summary.rejected_decisions],
            ["Pending Decisions", proposal.summary.pending_decisions],
            ["Validation Issues", len(proposal.validation_issues)],
            ["Audit Entries", len(proposal.audit_trail)],
        ]
    )
    return summary_rows


def _build_vehicle_load_rows(
    *,
    snapshot: TransportOperationalSnapshot,
    proposal: TransportOperationalProposal | None,
) -> list[list[object]]:
    projected_confirmations = Counter(
        decision.vehicle_id
        for decision in (proposal.decisions if proposal is not None else [])
        if decision.suggested_status == "confirmed" and decision.vehicle_id is not None
    )
    rows: list[list[object]] = [[
        "Placa/Plate",
        "Tipo/Type",
        "Capacidade/Capacity",
        "Alocados Atuais/Current Assigned",
        "Alocados Projetados/Projected Assigned",
        "Saldo Projetado/Projected Remaining",
        "Data/Date",
        "Rota/Route",
        "Partida/Departure",
    ]]

    for registry in sorted(
        _snapshot_vehicle_registry_rows(snapshot),
        key=lambda row: (row.placa or "", row.route_kind or "", row.vehicle_id),
    ):
        projected_assigned = registry.assigned_count + projected_confirmations.get(registry.vehicle_id, 0)
        projected_remaining = None if registry.lugares is None else registry.lugares - projected_assigned
        rows.append(
            [
                registry.placa,
                registry.tipo,
                registry.lugares,
                registry.assigned_count,
                projected_assigned,
                projected_remaining,
                registry.service_date.isoformat() if registry.service_date else None,
                registry.route_kind,
                registry.departure_time,
            ]
        )

    return rows


def _build_snapshot_request_rows(snapshot: TransportOperationalSnapshot) -> list[list[object]]:
    rows: list[list[object]] = [[
        "Request ID",
        "Tipo/Kind",
        "Status",
        "Horario/Time",
        "Data/Date",
        "Chave/Key",
        "Nome/Name",
        "Projeto/Project",
        "Workplace",
        "Endereco/Address",
        "Veiculo/Vehicle",
        "Resposta/Response",
    ]]

    for request in sorted(
        _snapshot_request_rows(snapshot),
        key=lambda row: (row.request_kind, row.nome.lower(), row.chave, row.id),
    ):
        rows.append(
            [
                request.id,
                request.request_kind,
                request.assignment_status,
                request.requested_time,
                request.service_date.isoformat(),
                request.chave,
                request.nome,
                request.projeto,
                request.workplace,
                request.end_rua,
                request.assigned_vehicle.placa if request.assigned_vehicle is not None else None,
                request.response_message,
            ]
        )

    return rows


def _build_proposed_decision_rows(
    *,
    snapshot: TransportOperationalSnapshot,
    proposal: TransportOperationalProposal,
) -> list[list[object]]:
    request_index = {request.id: request for request in _snapshot_request_rows(snapshot)}
    vehicle_index = {registry.vehicle_id: registry for registry in _snapshot_vehicle_registry_rows(snapshot)}
    rows: list[list[object]] = [[
        "Request ID",
        "Tipo/Kind",
        "Chave/Key",
        "Nome/Name",
        "Status Sugerido/Suggested Status",
        "Veiculo Sugerido/Suggested Vehicle",
        "Mensagem/Response",
        "Justificativa/Rationale",
    ]]

    for decision in proposal.decisions:
        request = request_index.get(decision.request_id)
        vehicle = vehicle_index.get(decision.vehicle_id) if decision.vehicle_id is not None else None
        rows.append(
            [
                decision.request_id,
                decision.request_kind,
                request.chave if request is not None else None,
                request.nome if request is not None else None,
                decision.suggested_status,
                vehicle.placa if vehicle is not None else decision.vehicle_id,
                decision.response_message,
                decision.rationale,
            ]
        )

    return rows


def _build_exception_rows(
    *,
    snapshot: TransportOperationalSnapshot,
    proposal: TransportOperationalProposal | None,
) -> list[list[object]]:
    rows: list[list[object]] = [["Tipo/Type", "Referencia/Reference", "Status ou Codigo/Status or Code", "Mensagem/Message"]]

    for request in sorted(
        _snapshot_request_rows(snapshot),
        key=lambda row: (row.assignment_status, row.nome.lower(), row.id),
    ):
        if request.assignment_status == "confirmed":
            continue
        rows.append(
            [
                "snapshot_request",
                f"request:{request.id}",
                request.assignment_status,
                f"{request.nome} ({request.chave}) remains {request.assignment_status}.",
            ]
        )

    if proposal is not None:
        for issue in proposal.validation_issues:
            reference_parts = []
            if issue.request_id is not None:
                reference_parts.append(f"request:{issue.request_id}")
            if issue.vehicle_id is not None:
                reference_parts.append(f"vehicle:{issue.vehicle_id}")
            rows.append(
                [
                    "validation_issue",
                    ", ".join(reference_parts) or proposal.proposal_key,
                    issue.code,
                    issue.message,
                ]
            )

        for decision in proposal.decisions:
            if decision.suggested_status == "confirmed":
                continue
            rows.append(
                [
                    "proposal_decision",
                    f"request:{decision.request_id}",
                    decision.suggested_status,
                    decision.response_message or decision.rationale,
                ]
            )

    return rows


def _build_audit_rows(proposal: TransportOperationalProposal) -> list[list[object]]:
    rows: list[list[object]] = [[
        "Acao/Action",
        "Resultado/Outcome",
        "Ator/Actor Key",
        "Nome/Actor Name",
        "Instante/Occurred At",
        "Mensagem/Message",
    ]]
    for entry in proposal.audit_trail:
        rows.append(
            [
                entry.action,
                entry.outcome,
                entry.actor.chave,
                entry.actor.nome_completo,
                entry.occurred_at.isoformat(),
                entry.message,
            ]
        )
    return rows


def build_transport_operational_plan_export(
    db: Session,
    *,
    service_date: date,
    selected_route_kind: str,
    proposal: TransportOperationalProposal | None,
) -> tuple[str, bytes]:
    from openpyxl import Workbook

    from .transport import now_sgt as transport_now_sgt
    from .transport_proposals import build_transport_operational_snapshot

    timestamp = transport_now_sgt()
    file_name = (
        _build_transport_operational_plan_file_name(timestamp)
        if proposal is not None
        else _build_transport_export_file_name(timestamp)
    )
    snapshot = proposal.snapshot if proposal is not None else build_transport_operational_snapshot(
        db,
        service_date=service_date,
        route_kind=selected_route_kind,
        captured_at=timestamp,
    )
    export_rows = _collect_transport_list_rows(
        db,
        service_date=service_date,
        selected_route_kind=selected_route_kind,
    )

    workbook = Workbook()
    transport_list_sheet = workbook.active
    transport_list_sheet.title = "Transport List"
    transport_list_sheet.append([
        "Nome/Name",
        "Chave/Key",
        "Projeto/Project",
        "Endereço/Address",
        "Data/Date",
        "Partida/Departure",
    ])
    for row in export_rows:
        transport_list_sheet.append(row)

    executive_summary_sheet = workbook.create_sheet("Executive Summary")
    _append_sheet_rows(
        executive_summary_sheet,
        _build_executive_summary_rows(snapshot=snapshot, generated_at=timestamp, proposal=proposal),
    )

    vehicle_load_sheet = workbook.create_sheet("Vehicle Load")
    _append_sheet_rows(vehicle_load_sheet, _build_vehicle_load_rows(snapshot=snapshot, proposal=proposal))

    snapshot_requests_sheet = workbook.create_sheet("Snapshot Requests")
    _append_sheet_rows(snapshot_requests_sheet, _build_snapshot_request_rows(snapshot))

    if proposal is not None:
        proposed_decisions_sheet = workbook.create_sheet("Proposed Decisions")
        _append_sheet_rows(
            proposed_decisions_sheet,
            _build_proposed_decision_rows(snapshot=snapshot, proposal=proposal),
        )

    exceptions_sheet = workbook.create_sheet("Exceptions")
    _append_sheet_rows(exceptions_sheet, _build_exception_rows(snapshot=snapshot, proposal=proposal))

    if proposal is not None:
        audit_trail_sheet = workbook.create_sheet("Audit Trail")
        _append_sheet_rows(audit_trail_sheet, _build_audit_rows(proposal))

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    content = output.getvalue()
    output.close()

    export_path = _resolve_transport_export_path(file_name)
    export_path.write_bytes(content)
    return export_path.name, content
