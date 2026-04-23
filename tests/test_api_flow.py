import asyncio
import io
import os
import json
import shutil
import socket
import threading
import time
import uuid
from contextlib import closing, contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from sqlalchemy import select, text
import uvicorn

# Override settings before app import.
os.environ["APP_ENV"] = "development"
os.environ["DATABASE_URL"] = "sqlite+pysqlite:///./test_checking.db"
os.environ["FORMS_URL"] = "https://example.com/form"
os.environ["DEVICE_SHARED_KEY"] = "device-test-key"
os.environ["MOBILE_APP_SHARED_KEY"] = "mobile-test-key"
os.environ["PROVIDER_SHARED_KEY"] = "PETROBRASP80P82P83"
os.environ["ADMIN_SESSION_SECRET"] = "test-admin-session-secret"
os.environ["BOOTSTRAP_ADMIN_KEY"] = "HR70"
os.environ["BOOTSTRAP_ADMIN_NAME"] = "Tamer Salmem"
os.environ["BOOTSTRAP_ADMIN_PASSWORD"] = "eAcacdLe2"
os.environ["FORMS_QUEUE_ENABLED"] = "false"
os.environ["TRANSPORT_EXPORTS_DIR"] = "./test_transport_exports"

test_db = Path("test_checking.db")
if test_db.exists():
    test_db.unlink()

test_transport_exports_dir = Path("test_transport_exports")
if test_transport_exports_dir.exists():
    shutil.rmtree(test_transport_exports_dir)

from fastapi.testclient import TestClient

from sistema.app.main import app
from sistema.app.core.config import settings
from sistema.app.database import Base, SessionLocal, engine
from sistema.app.models import (
    AdminAccessRequest,
    CheckEvent,
    CheckingHistory,
    FormsSubmission,
    Project,
    TransportAssignment,
    TransportRequest,
    TransportVehicleSchedule,
    TransportVehicleScheduleException,
    User,
    UserSyncEvent,
    Vehicle,
    Workplace,
)
from sistema.app.routers import admin as admin_router
from sistema.app.services.admin_updates import AdminUpdatesBroker, admin_updates_broker, notify_transport_data_changed
from sistema.app.services.forms_worker import FormsWorker
from sistema.app.services.forms_queue import process_forms_submission_queue_once
from sistema.app.services import forms_worker as forms_worker_module
from sistema.app.services import location_settings as location_settings_module
from sistema.app.services import transport as transport_service_module
from sistema.app.services import user_activity as user_activity_module
from sistema.app.services.passwords import hash_password, verify_password
from sistema.app.services.project_catalog import seed_default_projects
from sistema.app.services.time_utils import now_sgt
from sistema.app.services.user_sync import find_user_by_chave, find_user_by_rfid, normalize_event_time
from sistema.app.routers import web_check as web_check_router


ADMIN_LOGIN_CHAVE = "HR70"
ADMIN_LOGIN_SENHA = "eAcacdLe2"
MOBILE_HEADERS = {"x-mobile-shared-key": "mobile-test-key"}
PROVIDER_HEADERS = {"x-provider-shared-key": "PETROBRASP80P82P83"}

Base.metadata.create_all(bind=engine)
seed_default_projects()


def login_admin(client: TestClient, *, chave: str = ADMIN_LOGIN_CHAVE, senha: str = ADMIN_LOGIN_SENHA):
    return client.post("/api/admin/auth/login", json={"chave": chave, "senha": senha})


def ensure_admin_session(client: TestClient) -> None:
    session_response = client.get("/api/admin/auth/session")
    assert session_response.status_code == 200
    if not session_response.json().get("authenticated"):
        login_response = login_admin(client)
        assert login_response.status_code == 200, login_response.text

    transport_session_response = client.get("/api/transport/auth/session")
    assert transport_session_response.status_code == 200
    if transport_session_response.json().get("authenticated"):
        return

    transport_login_response = client.post(
        "/api/transport/auth/verify",
        json={"chave": ADMIN_LOGIN_CHAVE, "senha": ADMIN_LOGIN_SENHA},
    )
    assert transport_login_response.status_code == 200, transport_login_response.text
    assert transport_login_response.json()["authenticated"] is True


def ensure_web_user_exists(*, chave: str, projeto: str = "P80", nome: str = "Oriundo da Web") -> None:
    with SessionLocal() as db:
        existing = find_user_by_chave(db, chave)
        if existing is not None:
            return

        db.add(
            User(
                rfid=None,
                nome=nome,
                chave=chave,
                projeto=projeto,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                cargo=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
        )
        db.commit()


def ensure_project_exists(project_name: str) -> None:
    normalized_name = str(project_name).strip().upper()
    assert normalized_name

    with SessionLocal() as db:
        existing = db.execute(select(Project).where(Project.name == normalized_name)).scalar_one_or_none()
        if existing is not None:
            return

        db.add(Project(name=normalized_name))
        db.commit()


def register_web_password(
    client: TestClient,
    *,
    chave: str,
    senha: str = "abc123",
    projeto: str = "P80",
    ensure_user_exists: bool = True,
):
    if ensure_user_exists:
        ensure_web_user_exists(chave=chave, projeto=projeto)

    return client.post(
        "/api/web/auth/register-password",
        json={"chave": chave, "senha": senha, "projeto": projeto},
    )


def login_web_password(client: TestClient, *, chave: str, senha: str):
    return client.post(
        "/api/web/auth/login",
        json={"chave": chave, "senha": senha},
    )


def make_test_key(prefix: str) -> str:
    normalized_prefix = str(prefix or "T").strip().upper()[:1]
    assert normalized_prefix and normalized_prefix.isalnum()
    return f"{normalized_prefix}{uuid.uuid4().hex[:3].upper()}"


def reserve_tcp_port() -> int:
    with closing(socket.socket(socket.AF_INET, socket.SOCK_STREAM)) as sock:
        sock.bind(("127.0.0.1", 0))
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        return int(sock.getsockname()[1])


@contextmanager
def live_app_server():
    port = reserve_tcp_port()
    config = uvicorn.Config(app, host="127.0.0.1", port=port, log_level="warning")
    server = uvicorn.Server(config)
    server.install_signal_handlers = lambda: None
    thread = threading.Thread(target=server.run, daemon=True)
    thread.start()

    deadline = time.monotonic() + 10
    while time.monotonic() < deadline:
        if getattr(server, "started", False):
            break
        try:
            with closing(socket.create_connection(("127.0.0.1", port), timeout=0.2)):
                break
        except OSError:
            time.sleep(0.05)
    else:
        server.should_exit = True
        thread.join(timeout=5)
        raise RuntimeError("Nao foi possivel iniciar o servidor HTTP para o teste E2E do admin.")

    try:
        yield f"http://127.0.0.1:{port}"
    finally:
        server.should_exit = True
        thread.join(timeout=10)


def get_user_by_rfid(db, rfid: str) -> User:
    user = find_user_by_rfid(db, rfid)
    assert user is not None
    return user


def get_user_by_chave(db, chave: str) -> User:
    user = find_user_by_chave(db, chave)
    assert user is not None
    return user


def add_transport_schedule(
    db,
    *,
    vehicle: Vehicle,
    service_scope: str,
    route_kind: str,
    recurrence_kind: str,
    service_date: date | None = None,
    weekday: int | None = None,
    departure_time: str | None = None,
) -> TransportVehicleSchedule:
    timestamp = now_sgt()
    schedule = TransportVehicleSchedule(
        vehicle_id=vehicle.id,
        service_scope=service_scope,
        route_kind=route_kind,
        recurrence_kind=recurrence_kind,
        service_date=service_date,
        weekday=weekday,
        departure_time=departure_time,
        is_active=True,
        created_at=timestamp,
        updated_at=timestamp,
    )
    db.add(schedule)
    db.flush()
    return schedule


def set_user_checkin_state(*, chave: str, event_time: datetime, local: str = "Web") -> None:
    with SessionLocal() as db:
        user = get_user_by_chave(db, chave)
        user.checkin = True
        user.time = event_time
        user.local = local
        db.commit()


def test_health():
    with TestClient(app) as client:
        res = client.get("/api/health")
        assert res.status_code == 200
        assert res.json()["status"] == "ok"


def test_vehicle_schema_and_user_transport_fields_persist_expected_values():
    with SessionLocal() as db:
        db.add(Workplace(workplace="Innovation Hub", address="1 Harbour Front", zip="098632", country="Singapore"))
        vehicle = Vehicle(placa="SGX1234A", tipo="van", color="White", lugares=18, tolerance=12, service_scope="regular")
        db.add(vehicle)
        db.flush()

        user = User(
            rfid=None,
            nome="Usuario Transporte",
            chave="TR01",
            projeto="P80",
            workplace="Innovation Hub",
            placa="SGX1234A",
            end_rua="123 Harbour Road",
            zip="0012345678",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.commit()
        db.refresh(user)

        persisted_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "SGX1234A")).scalar_one()
        persisted_user = db.execute(select(User).where(User.chave == "TR01")).scalar_one()

        assert persisted_vehicle.tipo == "van"
        assert persisted_vehicle.color == "White"
        assert persisted_vehicle.lugares == 18
        assert persisted_vehicle.tolerance == 12
        assert persisted_vehicle.service_scope == "regular"
        assert persisted_user.workplace == "Innovation Hub"
        assert persisted_user.placa == "SGX1234A"
        assert persisted_user.end_rua == "123 Harbour Road"
        assert persisted_user.zip == "0012345678"


def test_mobile_sync_records_checkinghistory_entry():
    event_time = now_sgt().replace(microsecond=0)

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "HC01",
                "projeto": "P82",
                "action": "checkin",
                "event_time": event_time.isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert response.status_code == 200

    with SessionLocal() as db:
        row = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "HC01")
            .order_by(CheckingHistory.id.desc())
            .limit(1)
        ).scalar_one()

    assert row.atividade == "check-in"
    assert row.projeto == "P82"
    assert row.informe == "normal"
    assert row.time.replace(tzinfo=ZoneInfo(settings.tz_name)) == event_time


def test_admin_users_support_extended_profile_fields_and_key_updates_history():
    with SessionLocal() as db:
        db.add(Workplace(workplace="Refinery West", address="2 Harbour Road", zip="112233", country="Singapore"))
        db.add(Vehicle(placa="TRP1234A", tipo="van", color="Blue", lugares=18, tolerance=10, service_scope="regular"))
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post(
            "/api/admin/users",
            json={
                "rfid": "USR9001",
                "nome": "Usuario Completo",
                "chave": "UF01",
                "projeto": "P82",
                "workplace": "Refinery West",
                "placa": "TRP1234A",
                "end_rua": "123 Harbour Road",
                "zip": "0012345678",
                "cargo": "Cargo Inicial",
                "email": "USER@EXAMPLE.COM",
            },
        )
        assert created.status_code == 200

        sync_res = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "UF01",
                "projeto": "P82",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert sync_res.status_code == 200

        users = client.get("/api/admin/users")
        assert users.status_code == 200
        user_row = next(row for row in users.json() if row["chave"] == "UF01")
        assert user_row["workplace"] == "Refinery West"
        assert user_row["placa"] == "TRP1234A"
        assert user_row["end_rua"] == "123 Harbour Road"
        assert user_row["zip"] == "0012345678"
        assert user_row["cargo"] == "Cargo Inicial"
        assert user_row["email"] == "user@example.com"

        updated = client.post(
            "/api/admin/users",
            json={
                "user_id": user_row["id"],
                "rfid": "USR9002",
                "nome": "Usuario Ajustado",
                "chave": "UF02",
                "projeto": "P83",
                "workplace": "Refinery West",
                "placa": "TRP1234A",
                "end_rua": "456 Harbour Road",
                "zip": "0000001234",
                "cargo": "Cargo Final",
                "email": "final@example.com",
            },
        )
        assert updated.status_code == 200

        users_after = client.get("/api/admin/users")
        assert users_after.status_code == 200
        updated_row = next(row for row in users_after.json() if row["id"] == user_row["id"])
        assert updated_row["rfid"] == "USR9002"
        assert updated_row["nome"] == "Usuario Ajustado"
        assert updated_row["chave"] == "UF02"
        assert updated_row["projeto"] == "P83"
        assert updated_row["workplace"] == "Refinery West"
        assert updated_row["placa"] == "TRP1234A"
        assert updated_row["end_rua"] == "456 Harbour Road"
        assert updated_row["zip"] == "0000001234"
        assert updated_row["cargo"] == "Cargo Final"
        assert updated_row["email"] == "final@example.com"

    with SessionLocal() as db:
        sync_rows = db.execute(
            select(UserSyncEvent)
            .where(UserSyncEvent.user_id == user_row["id"])
            .order_by(UserSyncEvent.id)
        ).scalars().all()
        history_rows = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "UF02")
            .order_by(CheckingHistory.id)
        ).scalars().all()

    assert sync_rows
    assert all(row.chave == "UF02" for row in sync_rows)
    assert history_rows
    assert all(row.chave == "UF02" for row in history_rows)


def test_admin_projects_catalog_lists_creates_and_blocks_linked_user_deletion():
    with TestClient(app) as client:
        ensure_admin_session(client)

        listed = client.get("/api/admin/projects")
        assert listed.status_code == 200
        listed_names = {row["name"] for row in listed.json()}
        assert {"P80", "P82", "P83"}.issubset(listed_names)

        created = client.post("/api/admin/projects", json={"name": "P90"})
        assert created.status_code == 200, created.text
        project_payload = created.json()
        assert project_payload["name"] == "P90"

        with SessionLocal() as db:
            db.add(
                User(
                    rfid=None,
                    nome="Usuario Vinculado",
                    chave="PJ90",
                    projeto="P90",
                    perfil=0,
                    local=None,
                    checkin=None,
                    time=None,
                    last_active_at=now_sgt(),
                    inactivity_days=0,
                )
            )
            db.commit()

        blocked = client.delete(f"/api/admin/projects/{project_payload['id']}")
        assert blocked.status_code == 409
        assert blocked.json()["detail"] == "Nao e possivel remover um projeto com usuarios vinculados."

        with SessionLocal() as db:
            linked_user = get_user_by_chave(db, "PJ90")
            db.delete(linked_user)
            db.commit()

        removed = client.delete(f"/api/admin/projects/{project_payload['id']}")
        assert removed.status_code == 200
        assert removed.json()["ok"] is True

        listed_after = client.get("/api/admin/projects")
        assert listed_after.status_code == 200
        listed_after_names = {row["name"] for row in listed_after.json()}
        assert "P90" not in listed_after_names


def test_admin_project_delete_reassigns_admin_only_users_to_fallback_project():
    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post("/api/admin/projects", json={"name": "P91"})
        assert created.status_code == 200, created.text
        project_payload = created.json()

        with SessionLocal() as db:
            db.add(
                User(
                    rfid=None,
                    nome="Admin Vinculado",
                    chave="PA91",
                    projeto="P91",
                    perfil=1,
                    local=None,
                    checkin=None,
                    time=None,
                    last_active_at=now_sgt(),
                    inactivity_days=0,
                )
            )
            db.commit()

        removed = client.delete(f"/api/admin/projects/{project_payload['id']}")
        assert removed.status_code == 200, removed.text
        assert removed.json()["ok"] is True

    with SessionLocal() as db:
        admin_user = get_user_by_chave(db, "PA91")
        removed_project = db.execute(select(Project).where(Project.name == "P91")).scalar_one_or_none()

    assert removed_project is None
    assert admin_user.projeto in {"P80", "P82", "P83"}
    assert admin_user.projeto != "P91"


def test_web_projects_endpoint_lists_catalog_and_authenticated_project_update_persists():
    ensure_project_exists("P95")
    ensure_web_user_exists(chave="WP95", projeto="P80", nome="Usuario Projeto Web")

    with TestClient(app) as client:
        listed = client.get("/api/web/projects")
        assert listed.status_code == 200
        listed_names = {row["name"] for row in listed.json()}
        assert {"P80", "P82", "P83", "P95"}.issubset(listed_names)

        register_response = register_web_password(
            client,
            chave="WP95",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert register_response.status_code == 200, register_response.text

        updated = client.put("/api/web/project", json={"chave": "WP95", "projeto": "P95"})
        assert updated.status_code == 200, updated.text
        assert updated.json()["ok"] is True
        assert updated.json()["project"] == "P95"

        history_state = client.get("/api/web/check/state", params={"chave": "WP95"})
        assert history_state.status_code == 200, history_state.text
        assert history_state.json()["projeto"] == "P95"

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WP95")

    assert user.projeto == "P95"


def test_provider_endpoint_requires_valid_shared_key():
    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            json={
                "chave": "PV01",
                "nome": "Usuario Provider",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "08:00:00",
            },
        )

    assert response.status_code == 401
    assert response.json()["detail"] == "Invalid provider shared key"


def test_provider_endpoint_creates_user_and_history_with_normalized_name():
    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV11",
                "nome": "ADRIANO JOSE DA SILVA",
                "projeto": "P82",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "07:26:00",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["created_user"] is True
        assert payload["updated_current_state"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV11")
        history_rows = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "PV11")
            .order_by(CheckingHistory.id)
        ).scalars().all()
        provider_events = db.execute(
            select(UserSyncEvent)
            .where(UserSyncEvent.chave == "PV11", UserSyncEvent.source == "provider")
            .order_by(UserSyncEvent.id)
        ).scalars().all()
        forms_rows = db.execute(
            select(FormsSubmission)
            .where(FormsSubmission.chave == "PV11")
            .order_by(FormsSubmission.id)
        ).scalars().all()

    assert user.nome == "Adriano Jose da Silva"
    assert user.projeto == "P82"
    assert user.checkin is True
    assert user.time.strftime("%d/%m/%Y %H:%M:%S") == "17/04/2026 07:26:00"
    assert len(history_rows) == 1
    assert history_rows[0].atividade == "check-in"
    assert history_rows[0].informe == "normal"
    assert len(provider_events) == 1
    assert forms_rows == []


def test_provider_endpoint_never_enqueues_forms_even_for_multiple_events():
    with TestClient(app) as client:
        first = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV21",
                "nome": "USUARIO FORMS",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "07:30:00",
            },
        )
        second = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV21",
                "nome": "USUARIO FORMS",
                "projeto": "P80",
                "atividade": "check-out",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "18:00:00",
            },
        )
        assert first.status_code == 200
        assert second.status_code == 200

    with SessionLocal() as db:
        forms_rows = db.execute(
            select(FormsSubmission)
            .where(FormsSubmission.chave == "PV21")
            .order_by(FormsSubmission.id)
        ).scalars().all()
        provider_log_rows = db.execute(
            select(CheckEvent)
            .where(CheckEvent.source == "provider", CheckEvent.project == "P80")
            .order_by(CheckEvent.id.desc())
        ).scalars().all()

    assert forms_rows == []
    assert provider_log_rows
    assert all("forms_skipped=true" in (row.details or "") for row in provider_log_rows[:2])


def test_provider_endpoint_updates_project_but_keeps_existing_name():
    with SessionLocal() as db:
        db.add(
            User(
                rfid=None,
                chave="PV12",
                nome="Nome Original",
                projeto="P80",
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
        )
        db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV12",
                "nome": "NOME NAO DEVE TROCAR",
                "projeto": "P83",
                "atividade": "check-out",
                "informe": "retroativo",
                "data": "17/04/2026",
                "hora": "19:40:00",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["created_user"] is False
        assert payload["updated_project"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV12")
        history_row = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "PV12")
            .order_by(CheckingHistory.id.desc())
            .limit(1)
        ).scalar_one()

    assert user.nome == "Nome Original"
    assert user.projeto == "P83"
    assert user.checkin is False
    assert history_row.atividade == "check-out"
    assert history_row.informe == "retroativo"


def test_provider_endpoint_keeps_newer_current_user_state_while_recording_older_history():
    newer_time = datetime(2026, 4, 18, 9, 0, tzinfo=ZoneInfo(settings.tz_name))
    with SessionLocal() as db:
        db.add(
            User(
                rfid=None,
                chave="PV13",
                nome="Usuario Atual",
                projeto="P82",
                local="main",
                checkin=False,
                time=newer_time,
                last_active_at=newer_time,
                inactivity_days=0,
            )
        )
        db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV13",
                "nome": "IGNORADO",
                "projeto": "P82",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "08:00:00",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["updated_current_state"] is False

        duplicate = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV13",
                "nome": "IGNORADO",
                "projeto": "P82",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "08:00:00",
            },
        )
        assert duplicate.status_code == 200
        assert duplicate.json()["duplicate"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV13")
        history_rows = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "PV13")
            .order_by(CheckingHistory.id)
        ).scalars().all()
        provider_events = db.execute(
            select(UserSyncEvent)
            .where(UserSyncEvent.chave == "PV13", UserSyncEvent.source == "provider")
            .order_by(UserSyncEvent.id)
        ).scalars().all()

    assert normalize_event_time(user.time) == newer_time
    assert user.checkin is False
    assert user.local == "main"
    assert len(history_rows) == 2
    assert history_rows[0].atividade == "check-out"
    assert history_rows[1].atividade == "check-in"
    assert len(provider_events) == 1


def test_provider_same_day_events_do_not_override_web_state_and_are_reported_in_forms():
    web_checkin_time = datetime(2026, 4, 20, 7, 10, 0, tzinfo=ZoneInfo(settings.tz_name))
    web_checkout_time = datetime(2026, 4, 20, 21, 7, 14, tzinfo=ZoneInfo(settings.tz_name))

    with TestClient(app) as client:
        registered = register_web_password(client, chave="PV31", senha="web123", projeto="P80")
        assert registered.status_code == 200, registered.text

        web_checkin = client.post(
            "/api/web/check",
            json={
                "chave": "PV31",
                "projeto": "P80",
                "action": "checkin",
                "informe": "normal",
                "local": "Web",
                "event_time": web_checkin_time.isoformat(),
                "client_event_id": f"web-checkin-{uuid.uuid4().hex}",
            },
        )
        assert web_checkin.status_code == 200, web_checkin.text

        provider_checkin = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV31",
                "nome": "USUARIO WEB FORMS",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "20/04/2026",
                "hora": "07:12:33",
            },
        )
        assert provider_checkin.status_code == 200, provider_checkin.text
        assert provider_checkin.json()["updated_current_state"] is False

        ensure_admin_session(client)

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200, checkin_rows.text
        checkin_row = next(row for row in checkin_rows.json() if row["chave"] == "PV31")
        assert normalize_event_time(datetime.fromisoformat(checkin_row["time"])) == web_checkin_time
        assert checkin_row["local"] == "Web"

        web_checkout = client.post(
            "/api/web/check",
            json={
                "chave": "PV31",
                "projeto": "P80",
                "action": "checkout",
                "informe": "normal",
                "local": "Web",
                "event_time": web_checkout_time.isoformat(),
                "client_event_id": f"web-checkout-{uuid.uuid4().hex}",
            },
        )
        assert web_checkout.status_code == 200, web_checkout.text

        provider_checkout = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV31",
                "nome": "USUARIO WEB FORMS",
                "projeto": "P80",
                "atividade": "check-out",
                "informe": "normal",
                "data": "20/04/2026",
                "hora": "21:09:37",
            },
        )
        assert provider_checkout.status_code == 200, provider_checkout.text
        assert provider_checkout.json()["updated_current_state"] is False

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200, checkout_rows.text
        checkout_row = next(row for row in checkout_rows.json() if row["chave"] == "PV31")
        assert normalize_event_time(datetime.fromisoformat(checkout_row["time"])) == web_checkout_time
        assert checkout_row["local"] == "Web"

        history_state = client.get("/api/web/check/state", params={"chave": "PV31"})
        assert history_state.status_code == 200, history_state.text
        assert normalize_event_time(datetime.fromisoformat(history_state.json()["last_checkout_at"])) == web_checkout_time

        forms_rows = client.get("/api/admin/forms")
        assert forms_rows.status_code == 200, forms_rows.text
        provider_rows = [row for row in forms_rows.json() if row["chave"] == "PV31"]
        assert any(row["atividade"] == "check-in" and row["hora"] == "07:12:33" for row in provider_rows)
        assert any(row["atividade"] == "check-out" and row["hora"] == "21:09:37" for row in provider_rows)


def test_provider_current_state_uses_forms_as_local_when_provider_event_wins():
    provider_time = datetime(2026, 4, 21, 18, 5, 0, tzinfo=ZoneInfo(settings.tz_name))

    with TestClient(app) as client:
        ensure_admin_session(client)

        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV32",
                "nome": "USUARIO PROVIDER LOCAL",
                "projeto": "P82",
                "atividade": "check-out",
                "informe": "retroativo",
                "data": "21/04/2026",
                "hora": "18:05:00",
            },
        )
        assert response.status_code == 200, response.text
        assert response.json()["updated_current_state"] is True

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200, checkout_rows.text
        checkout_row = next(row for row in checkout_rows.json() if row["chave"] == "PV32")
        assert normalize_event_time(datetime.fromisoformat(checkout_row["time"])) == provider_time
        assert checkout_row["local"] == "Forms"
        assert checkout_row["assiduidade"] == "Retroativo"

        forms_rows = client.get("/api/admin/forms")
        assert forms_rows.status_code == 200, forms_rows.text
        forms_row = next(row for row in forms_rows.json() if row["chave"] == "PV32")
        assert forms_row["nome"] == "USUARIO PROVIDER LOCAL"
        assert forms_row["atividade"] == "check-out"
        assert forms_row["informe"] == "retroativo"
        assert forms_row["hora"] == "18:05:00"


def test_web_check_ignores_provider_checkout_when_deciding_same_day_submission():
    first_event_time = datetime(2026, 4, 22, 8, 0, 0, tzinfo=ZoneInfo(settings.tz_name))
    second_event_time = datetime(2026, 4, 22, 11, 0, 0, tzinfo=ZoneInfo(settings.tz_name))

    with TestClient(app) as client:
        registered = register_web_password(client, chave="PV33", senha="web123", projeto="P80")
        assert registered.status_code == 200, registered.text

        first = client.post(
            "/api/web/check",
            json={
                "chave": "PV33",
                "projeto": "P80",
                "action": "checkin",
                "informe": "normal",
                "local": "Web",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"web-provider-ignore-1-{uuid.uuid4().hex}",
            },
        )
        assert first.status_code == 200, first.text
        assert first.json()["queued_forms"] is True

        provider_checkout = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV33",
                "nome": "USUARIO WEB IGNORA PROVIDER",
                "projeto": "P80",
                "atividade": "check-out",
                "informe": "normal",
                "data": "22/04/2026",
                "hora": "10:00:00",
            },
        )
        assert provider_checkout.status_code == 200, provider_checkout.text
        assert provider_checkout.json()["updated_current_state"] is True

        second = client.post(
            "/api/web/check",
            json={
                "chave": "PV33",
                "projeto": "P80",
                "action": "checkout",
                "informe": "normal",
                "local": "Web",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"web-provider-ignore-2-{uuid.uuid4().hex}",
            },
        )
        assert second.status_code == 200, second.text
        assert second.json()["queued_forms"] is True

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.chave == "PV33").order_by(FormsSubmission.id)
            ).scalars().all()
            assert len(queued) == 2


def test_admin_can_clear_forms_rows_without_archiving():
    timestamp = now_sgt().replace(microsecond=0)
    provider_event_key = uuid.uuid4().hex
    device_event_key = uuid.uuid4().hex

    with SessionLocal() as db:
        db.add_all(
            [
                CheckEvent(
                    idempotency_key=provider_event_key,
                    source="provider",
                    rfid=None,
                    action="checkin",
                    status="success",
                    message="Entrada Forms registrada",
                    details="chave=PF99; nome=USUARIO LIMPAR FORMS; projeto=P80; atividade=check-in; informe=normal; data=22/04/2026; hora=08:00:00",
                    project="P80",
                    device_id="FORMS-CLEAR-01",
                    local="Forms",
                    request_path="/api/provider/updaterecords",
                    http_status=200,
                    ontime=True,
                    event_time=timestamp,
                    submitted_at=timestamp,
                    retry_count=0,
                ),
                CheckEvent(
                    idempotency_key=device_event_key,
                    source="device",
                    rfid="rfid-clear-forms",
                    action="checkin",
                    status="success",
                    message="Entrada por leitor",
                    details="reader=gate-clear",
                    project="P82",
                    device_id="ESP-CLEAR-01",
                    local="Portaria",
                    request_path="/api/scan",
                    http_status=200,
                    ontime=True,
                    event_time=timestamp,
                    submitted_at=timestamp,
                    retry_count=0,
                ),
            ]
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        forms_before = client.get("/api/admin/forms")
        assert forms_before.status_code == 200, forms_before.text
        assert any(row["chave"] == "PF99" for row in forms_before.json())

        cleared = client.delete("/api/admin/forms")
        assert cleared.status_code == 200, cleared.text
        assert cleared.json()["ok"] is True
        assert "Forms" in cleared.json()["message"]

        forms_after = client.get("/api/admin/forms")
        assert forms_after.status_code == 200, forms_after.text
        assert all(row["chave"] != "PF99" for row in forms_after.json())

    with SessionLocal() as db:
        provider_event = db.execute(
            select(CheckEvent).where(CheckEvent.idempotency_key == provider_event_key)
        ).scalar_one_or_none()
        device_event = db.execute(
            select(CheckEvent).where(CheckEvent.idempotency_key == device_event_key)
        ).scalar_one_or_none()

    assert provider_event is None
    assert device_event is not None


def test_admin_stream_requires_valid_session():
    with TestClient(app) as client:
        forbidden = client.get("/api/admin/stream")
        assert forbidden.status_code == 401


def test_admin_routes_do_not_accept_legacy_header_only():
    with TestClient(app) as client:
        forbidden = client.get("/api/admin/pending", headers={"x-admin-key": "admin-test-key"})
        assert forbidden.status_code == 401


def test_admin_updates_broker_publishes_payload():
    broker = AdminUpdatesBroker()
    subscriber_id, queue = broker.subscribe()

    try:
        broker.publish("pending")
        payload = queue.get_nowait()
        assert '"reason": "pending"' in payload
        assert '"emitted_at":' in payload
    finally:
        broker.unsubscribe(subscriber_id)


def test_pending_registration_flow():
    with TestClient(app) as client:
        ensure_admin_session(client)
        heartbeat = client.post(
            "/api/device/heartbeat",
            json={"device_id": "ESP32-TEST", "shared_key": "device-test-key"},
        )
        assert heartbeat.status_code == 200

        scan_pending = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "ABC12345",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_pending.status_code == 200
        assert scan_pending.json()["outcome"] == "pending_registration"
        assert scan_pending.json()["led"] == "orange_4s"

        pending_list = client.get("/api/admin/pending")
        assert pending_list.status_code == 200
        assert len(pending_list.json()) >= 1

        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ABC12345", "nome": "Usuario Teste", "chave": "UT70", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        assert all(row["rfid"] != "ABC12345" for row in checkin_rows.json())

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        assert all(row["rfid"] != "ABC12345" for row in checkout_rows.json())


def test_unknown_rfid_goes_pending():
    with TestClient(app) as client:
        ensure_admin_session(client)
        scan_unknown = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "ZZZ99999",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_unknown.status_code == 200
        assert scan_unknown.json()["outcome"] == "pending_registration"
        assert scan_unknown.json()["led"] == "orange_4s"

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(event["status"] == "received" and event["request_path"] == "/api/scan" for event in events.json())
        assert any(event["status"] == "pending" and event["source"] == "device" for event in events.json())


def test_explicit_checkin_and_checkout_flow(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARD1000", "nome": "Usuario Fluxo", "chave": "AB12", "projeto": "P83"},
        )
        assert save_user.status_code == 200

        scan_checkin = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARD1000",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkin.status_code == 200
        assert scan_checkin.json()["outcome"] == "submitted"
        assert scan_checkin.json()["led"] == "green_1s"

        processed_first = process_forms_submission_queue_once()
        assert processed_first >= 1

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        assert any(row["rfid"] == "CARD1000" and row["local"] == "main" for row in checkin_rows.json())

        scan_checkin_again = client.post(
            "/api/scan",
            json={
                "local": "un83",
                "rfid": "CARD1000",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkin_again.status_code == 200
        assert scan_checkin_again.json()["outcome"] == "local_updated"
        assert scan_checkin_again.json()["led"] == "green_blink_3x_1s"

        checkin_rows_updated = client.get("/api/admin/checkin")
        assert checkin_rows_updated.status_code == 200
        assert any(row["rfid"] == "CARD1000" and row["local"] == "un83" for row in checkin_rows_updated.json())

        checkout_rows_after_checkin = client.get("/api/admin/checkout")
        assert checkout_rows_after_checkin.status_code == 200
        assert all(row["rfid"] != "CARD1000" for row in checkout_rows_after_checkin.json())

        scan_checkout = client.post(
            "/api/scan",
            json={
                "local": "co83",
                "rfid": "CARD1000",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkout.status_code == 200
        assert scan_checkout.json()["outcome"] == "submitted"
        assert scan_checkout.json()["led"] == "green_1s"

        processed_second = process_forms_submission_queue_once()
        assert processed_second >= 1

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        assert any(row["rfid"] == "CARD1000" and row["local"] == "co83" for row in checkout_rows.json())

        checkin_rows_after = client.get("/api/admin/checkin")
        assert checkin_rows_after.status_code == 200
        assert all(row["rfid"] != "CARD1000" for row in checkin_rows_after.json())

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(
            event["rfid"] == "CARD1000"
            and event["chave"] == "AB12"
            and event["request_path"] == "/api/scan"
            for event in events.json()
        )


def test_checkout_without_checkin_returns_red_2s(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARD2000", "nome": "Usuario Sem Checkin", "chave": "EF56", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_checkout = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARD2000",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkout.status_code == 200
        assert scan_checkout.json()["outcome"] == "failed"
        assert scan_checkout.json()["led"] == "red_2s"
        assert "Check-in not found" in scan_checkout.json()["message"]

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(event["rfid"] == "CARD2000" and event["status"] == "blocked" for event in events.json())


def test_repeated_same_day_checkout_updates_state_without_forms_submission(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARD2001", "nome": "Usuario Checkout Repetido", "chave": "EG57", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_checkin = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARD2001",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkin.status_code == 200
        assert scan_checkin.json()["outcome"] == "submitted"

        first_checkout = client.post(
            "/api/scan",
            json={
                "local": "co80-a",
                "rfid": "CARD2001",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert first_checkout.status_code == 200
        assert first_checkout.json()["outcome"] == "submitted"

        second_checkout = client.post(
            "/api/scan",
            json={
                "local": "co80-b",
                "rfid": "CARD2001",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert second_checkout.status_code == 200
        assert second_checkout.json()["outcome"] == "local_updated"
        assert second_checkout.json()["led"] == "green_blink_3x_1s"

        with SessionLocal() as db:
            user = get_user_by_rfid(db, "CARD2001")
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "CARD2001")).scalars().all()

            assert user.checkin is False
            assert user.local == "co80-b"
            assert len(queued) == 2


def test_device_checkout_ignores_provider_only_history():
    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARDPROV1", "nome": "Usuario Provider Ignorado", "chave": "EG59", "projeto": "P80"},
        )
        assert save_user.status_code == 200, save_user.text

        provider_checkin = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "EG59",
                "nome": "USUARIO DEVICE IGNORA PROVIDER",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "22/04/2026",
                "hora": "08:00:00",
            },
        )
        assert provider_checkin.status_code == 200, provider_checkin.text
        assert provider_checkin.json()["updated_current_state"] is True

        scan_checkout = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARDPROV1",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkout.status_code == 200, scan_checkout.text
        assert scan_checkout.json()["outcome"] == "failed"
        assert scan_checkout.json()["led"] == "red_2s"
        assert "Check-in not found" in scan_checkout.json()["message"]

        with SessionLocal() as db:
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "CARDPROV1")).scalars().all()
            assert queued == []


def test_repeated_checkout_after_singapore_midnight_is_queued_again(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARD2002", "nome": "Usuario Midnight", "chave": "EG58", "projeto": "P83"},
        )
        assert save_user.status_code == 200

        with SessionLocal() as db:
            user = get_user_by_rfid(db, "CARD2002")
            prior_checkout = now_sgt() - timedelta(days=1)
            user.checkin = False
            user.local = "co83-old"
            user.time = prior_checkout
            user.last_active_at = prior_checkout
            db.commit()

        scan_checkout = client.post(
            "/api/scan",
            json={
                "local": "co83-new",
                "rfid": "CARD2002",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkout.status_code == 200
        assert scan_checkout.json()["outcome"] == "submitted"
        assert scan_checkout.json()["led"] == "green_1s"

        with SessionLocal() as db:
            user = get_user_by_rfid(db, "CARD2002")
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "CARD2002")).scalars().all()

            assert user.checkin is False
            assert user.local == "co83-new"
            assert len(queued) == 1


def test_forms_step_timeout_returns_red_blink_pattern(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": False,
            "message": "Step 'digitar_chave' not found within 10 seconds",
            "retry_count": 0,
            "error_code": "forms_step_timeout",
            "failed_step": "digitar_chave",
            "audit_events": [
                {
                    "source": "forms",
                    "action": "forms",
                    "status": "failed",
                    "message": "Forms step timeout",
                    "details": "step=digitar_chave; timeout=10",
                }
            ],
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARDFAIL", "nome": "Usuario Falha", "chave": "CD34", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_failed = client.post(
            "/api/scan",
            json={
                "local": "co80",
                "rfid": "CARDFAIL",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_failed.status_code == 200
        assert scan_failed.json()["outcome"] == "submitted"
        assert scan_failed.json()["led"] == "green_1s"

        processed = process_forms_submission_queue_once()
        assert processed >= 1

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(event["rfid"] == "CARDFAIL" and event["status"] == "failed" for event in events.json())
        assert any(event["source"] == "forms" and event["status"] == "failed" for event in events.json())
        assert all(
            not (event["source"] == "forms" and event["message"] == "Forms step timeout")
            for event in events.json()
        )


def test_valid_scan_updates_user_before_forms_processing(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARDFAST", "nome": "Usuario Rapido", "chave": "QP12", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_response = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARDFAST",
                "action": "checkin",
                "device_id": "ESP32-FAST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_response.status_code == 200
        assert scan_response.json()["outcome"] == "submitted"
        assert scan_response.json()["led"] == "green_1s"

        with SessionLocal() as db:
            user = get_user_by_rfid(db, "CARDFAST")
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "CARDFAST")).scalar_one()
            assert user.checkin is True
            assert user.local == "main"
            assert queued.status == "pending"


def test_forms_queue_processing_persists_failure_state(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": False,
            "message": "mocked queue failure",
            "retry_count": 2,
            "error_code": "forms_runtime_error",
            "failed_step": "botao_enviar",
            "audit_events": [
                {
                    "source": "forms",
                    "action": "forms",
                    "status": "failed",
                    "message": "Forms runtime error",
                    "details": "mocked queue failure",
                }
            ],
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "QUEUEFAIL", "nome": "Fila Falha", "chave": "LM34", "projeto": "P83"},
        )
        assert save_user.status_code == 200

        scan_response = client.post(
            "/api/scan",
            json={
                "local": "line-2",
                "rfid": "QUEUEFAIL",
                "action": "checkin",
                "device_id": "ESP32-QUEUE",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_response.status_code == 200
        assert scan_response.json()["outcome"] == "submitted"

        processed = process_forms_submission_queue_once()
        assert processed >= 1

        with SessionLocal() as db:
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "QUEUEFAIL")).scalar_one()
            assert queued.status == "failed"
            assert queued.retry_count == 2
            assert queued.last_error == "mocked queue failure"


def test_forms_success_generates_single_final_forms_event(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": True,
            "message": "Form submitted successfully",
            "retry_count": 0,
            "audit_events": [
                {
                    "source": "forms",
                    "action": "forms",
                    "status": "opened",
                    "message": "Microsoft Forms opened",
                    "details": None,
                },
                {
                    "source": "forms",
                    "action": "forms",
                    "status": "completed",
                    "message": "Microsoft Forms completed",
                    "details": "steps=ok; success_xpath_visible=true",
                },
            ],
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "FORMSOK", "nome": "Forms Final", "chave": "ZX12", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_response = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "FORMSOK",
                "action": "checkin",
                "device_id": "ESP32-FORMS",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_response.status_code == 200

        processed = process_forms_submission_queue_once()
        assert processed >= 1

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        forms_events = [event for event in events.json() if event["rfid"] == "FORMSOK" and event["source"] == "forms"]
        assert len(forms_events) == 1
        assert forms_events[0]["status"] == "success"
        assert "forms_details=steps=ok; success_xpath_visible=true" in (forms_events[0]["details"] or "")


def test_forms_worker_requires_success_xpath_after_submit(tmp_path, monkeypatch):
    xpath_dir = tmp_path / "xpath"
    xpath_dir.mkdir(parents=True)
    xpaths = {
        "digitar_chave.txt": "//digitar_chave",
        "confirmar_chave.txt": "//confirmar_chave",
        "botao_normal.txt": "//botao_normal",
        "botao_retroativo.txt": "//botao_retroativo",
        "botao_checkin.txt": "//botao_checkin",
        "botao_checkout.txt": "//botao_checkout",
        "botao_enviar.txt": "//botao_enviar",
        "sucesso.txt": "//sucesso",
        "botao_projeto_P80.txt": "//botao_projeto_P80",
        "botao_projeto_P82.txt": "//botao_projeto_P82",
        "botao_projeto_P83.txt": "//botao_projeto_P83",
    }
    for name, content in xpaths.items():
        (xpath_dir / name).write_text(content, encoding="utf-8")

    send_selector = "xpath=//botao_enviar"
    success_selector = "xpath=//sucesso"

    class FakeLocator:
        def __init__(self, page, selector: str):
            self.page = page
            self.selector = selector

        def fill(self, value: str) -> None:
            self.page.filled[self.selector] = value

        def click(self) -> None:
            self.page.clicked.append(self.selector)
            self.page.checked.add(self.selector)
            if self.selector == send_selector:
                self.page.success_visible = True
                self.page.url = f"{self.page.url}#submitted"

        def input_value(self) -> str:
            return self.page.filled.get(self.selector, "")

        def is_checked(self) -> bool:
            return self.selector in self.page.checked

        def inner_text(self) -> str:
            if self.selector == success_selector:
                return "Sua resposta foi enviada."
            return ""

    class FakePage:
        def __init__(self):
            self.url = ""
            self.success_visible = False
            self.filled = {}
            self.clicked = []
            self.checked = set()
            self.visible_selectors = {
                "xpath=//digitar_chave",
                "xpath=//confirmar_chave",
                "xpath=//botao_normal",
                "xpath=//botao_checkin",
                "xpath=//botao_checkout",
                send_selector,
                "xpath=//botao_projeto_P80",
                "xpath=//botao_projeto_P83",
            }

        def goto(self, url: str, timeout: int) -> None:
            self.url = url

        def wait_for_selector(self, selector: str, state: str = "visible", timeout: int = 0):
            if selector == success_selector and self.success_visible:
                return True
            if selector in self.visible_selectors:
                return True
            raise forms_worker_module.PlaywrightTimeoutError("timeout")

        def locator(self, selector: str) -> FakeLocator:
            return FakeLocator(self, selector)

    class FakeBrowser:
        def __init__(self, page: FakePage):
            self.page = page

        def new_page(self) -> FakePage:
            return self.page

        def close(self) -> None:
            return None

    class FakePlaywright:
        def __init__(self, page: FakePage):
            self.chromium = SimpleNamespace(launch=lambda headless=True: FakeBrowser(page))

    class FakePlaywrightContext:
        def __init__(self, page: FakePage):
            self.page = page

        def __enter__(self):
            return FakePlaywright(self.page)

        def __exit__(self, exc_type, exc, tb):
            return False

    fake_page = FakePage()
    monkeypatch.setattr(forms_worker_module, "sync_playwright", lambda: FakePlaywrightContext(fake_page))

    worker = FormsWorker(assets_dir=tmp_path)
    result = worker.submit_with_retries(action="checkin", chave="HR70", projeto="P80")

    assert result["success"] is True
    completed_event = next(event for event in result["audit_events"] if event["status"] == "completed")
    assert "steps=digitar_chave:filled+verified,confirmar_chave:filled+verified,botao_normal:clicked+verified,botao_checkin:clicked+verified,botao_projeto_P80:clicked+verified,botao_enviar:clicked,sucesso:visible" in completed_event["details"]
    assert "success_xpath_visible=true" in completed_event["details"]
    assert "submit_to_success_ms=" in completed_event["details"]
    assert "success_text=Sua resposta foi enviada." in completed_event["details"]


def test_forms_worker_rejects_success_xpath_visible_before_submit(tmp_path, monkeypatch):
    xpath_dir = tmp_path / "xpath"
    xpath_dir.mkdir(parents=True)
    for name in [
        "digitar_chave.txt",
        "confirmar_chave.txt",
        "botao_normal.txt",
        "botao_retroativo.txt",
        "botao_checkin.txt",
        "botao_checkout.txt",
        "botao_enviar.txt",
        "sucesso.txt",
        "botao_projeto_P80.txt",
            "botao_projeto_P82.txt",
        "botao_projeto_P83.txt",
    ]:
        (xpath_dir / name).write_text(f"//{name}", encoding="utf-8")

    success_selector = "xpath=//sucesso.txt"

    class FakeLocator:
        def __init__(self, selector: str):
            self.selector = selector

        def fill(self, value: str) -> None:
            return None

        def click(self) -> None:
            return None

        def input_value(self) -> str:
            return "HR70"

        def is_checked(self) -> bool:
            return True

        def inner_text(self) -> str:
            return ""

    class FakePage:
        url = "https://example.com/form"

        def goto(self, url: str, timeout: int) -> None:
            self.url = url

        def wait_for_selector(self, selector: str, state: str = "visible", timeout: int = 0):
            if selector == success_selector:
                return True
            return True

        def locator(self, selector: str) -> FakeLocator:
            return FakeLocator(selector)

    class FakeBrowser:
        def new_page(self) -> FakePage:
            return FakePage()

        def close(self) -> None:
            return None

    class FakePlaywright:
        chromium = SimpleNamespace(launch=lambda headless=True: FakeBrowser())

    class FakePlaywrightContext:
        def __enter__(self):
            return FakePlaywright()

        def __exit__(self, exc_type, exc, tb):
            return False

    monkeypatch.setattr(forms_worker_module, "sync_playwright", lambda: FakePlaywrightContext())

    worker = FormsWorker(assets_dir=tmp_path)
    result = worker.submit_with_retries(action="checkin", chave="HR70", projeto="P80")

    assert result["success"] is False
    assert result["error_code"] == "forms_validation_error"
    assert "XPath de sucesso ja estava visivel antes do envio" in result["message"]


def test_forms_worker_fails_when_first_field_is_not_confirmed(tmp_path, monkeypatch):
    xpath_dir = tmp_path / "xpath"
    xpath_dir.mkdir(parents=True)
    xpaths = {
        "digitar_chave.txt": "//digitar_chave",
        "confirmar_chave.txt": "//confirmar_chave",
        "botao_normal.txt": "//botao_normal",
        "botao_retroativo.txt": "//botao_retroativo",
        "botao_checkin.txt": "//botao_checkin",
        "botao_checkout.txt": "//botao_checkout",
        "botao_enviar.txt": "//botao_enviar",
        "sucesso.txt": "//sucesso",
        "botao_projeto_P80.txt": "//botao_projeto_P80",
        "botao_projeto_P82.txt": "//botao_projeto_P82",
        "botao_projeto_P83.txt": "//botao_projeto_P83",
    }
    for name, content in xpaths.items():
        (xpath_dir / name).write_text(content, encoding="utf-8")

    class FakeLocator:
        def __init__(self, selector: str):
            self.selector = selector

        def fill(self, value: str) -> None:
            return None

        def click(self) -> None:
            return None

        def input_value(self) -> str:
            return ""

        def is_checked(self) -> bool:
            return False

        def inner_text(self) -> str:
            return ""

    class FakePage:
        url = "https://example.com/form"

        def goto(self, url: str, timeout: int) -> None:
            self.url = url

        def wait_for_selector(self, selector: str, state: str = "visible", timeout: int = 0):
            return True

        def locator(self, selector: str) -> FakeLocator:
            return FakeLocator(selector)

    class FakeBrowser:
        def new_page(self) -> FakePage:
            return FakePage()

        def close(self) -> None:
            return None

    class FakePlaywright:
        chromium = SimpleNamespace(launch=lambda headless=True: FakeBrowser())

    class FakePlaywrightContext:
        def __enter__(self):
            return FakePlaywright()

        def __exit__(self, exc_type, exc, tb):
            return False

    monkeypatch.setattr(forms_worker_module, "sync_playwright", lambda: FakePlaywrightContext())

    worker = FormsWorker(assets_dir=tmp_path)
    result = worker.submit_with_retries(action="checkin", chave="HR70", projeto="P80")

    assert result["success"] is False
    assert result["error_code"] == "forms_step_validation_failed"
    assert result["failed_step"] == "digitar_chave"
    assert "expected_value=HR70" in result["audit_events"][0]["details"]


def test_heartbeat_success_is_not_logged_in_events():
    with TestClient(app) as client:
        ensure_admin_session(client)
        heartbeat = client.post(
            "/api/device/heartbeat",
            json={"device_id": "ESP32-TEST", "shared_key": "device-test-key"},
        )
        assert heartbeat.status_code == 200

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert all(
            not (
                event["action"] == "heartbeat"
                and event["status"] == "success"
                and event["device_id"] == "ESP32-TEST"
            )
            for event in events.json()
        )


def test_heartbeat_failure_is_logged_in_events():
    with TestClient(app) as client:
        ensure_admin_session(client)
        heartbeat = client.post(
            "/api/device/heartbeat",
            json={"device_id": "ESP32-BAD", "shared_key": "wrong-key"},
        )
        assert heartbeat.status_code == 200
        assert heartbeat.json()["ok"] is False

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(
            event["action"] == "heartbeat"
            and event["status"] == "failed"
            and event["device_id"] == "ESP32-BAD"
            for event in events.json()
        )


def test_remove_pending_registration():
    with TestClient(app) as client:
        ensure_admin_session(client)
        scan_unknown = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "PENDING01",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_unknown.status_code == 200

        pending_list = client.get("/api/admin/pending")
        assert pending_list.status_code == 200
        pending_id = next(row["id"] for row in pending_list.json() if row["rfid"] == "PENDING01")

        remove_res = client.delete(f"/api/admin/pending/{pending_id}")
        assert remove_res.status_code == 200

        pending_list_after = client.get("/api/admin/pending")
        assert pending_list_after.status_code == 200
        assert all(row["id"] != pending_id for row in pending_list_after.json())


def test_list_and_remove_registered_user():
    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "USERDEL1", "nome": "Usuario Cadastro", "chave": "GH78", "projeto": "P83"},
        )
        assert save_user.status_code == 200

        checkout_before = client.get("/api/admin/checkout")
        assert checkout_before.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in checkout_before.json())

        checkin_before = client.get("/api/admin/checkin")
        assert checkin_before.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in checkin_before.json())

        events_before = client.get("/api/admin/events")
        assert events_before.status_code == 200
        assert any(event["rfid"] == "USERDEL1" for event in events_before.json())

        users_before = client.get("/api/admin/users")
        assert users_before.status_code == 200
        assert any(row["rfid"] == "USERDEL1" and row["nome"] == "Usuario Cadastro" for row in users_before.json())

        user_id = next(row["id"] for row in users_before.json() if row["rfid"] == "USERDEL1")

        remove_user = client.delete(f"/api/admin/users/{user_id}")
        assert remove_user.status_code == 200

        users_after = client.get("/api/admin/users")
        assert users_after.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in users_after.json())

        checkout_after = client.get("/api/admin/checkout")
        assert checkout_after.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in checkout_after.json())

        checkin_after = client.get("/api/admin/checkin")
        assert checkin_after.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in checkin_after.json())

        events_after = client.get("/api/admin/events")
        assert events_after.status_code == 200
        assert any(event["rfid"] == "USERDEL1" for event in events_after.json())


def test_overdue_checkins_stay_visible_and_populate_missing_checkout(monkeypatch):
    fixed_now = datetime(2024, 4, 8, 12, 0, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(admin_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(user_activity_module, "now_sgt", lambda: fixed_now)

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_a = client.post(
            "/api/admin/users",
            json={"rfid": "INA001", "nome": "Zelda Ativa", "chave": "AA11", "projeto": "P80"},
        )
        save_b = client.post(
            "/api/admin/users",
            json={"rfid": "INA002", "nome": "Ana Inativa", "chave": "BB22", "projeto": "P83"},
        )
        save_c = client.post(
            "/api/admin/users",
            json={"rfid": "INA003", "nome": "Bruno Inativo", "chave": "CC33", "projeto": "P80"},
        )
        save_d = client.post(
            "/api/admin/users",
            json={"rfid": "INA004", "nome": "Carlos Checkout", "chave": "DD55", "projeto": "P82"},
        )
        save_e = client.post(
            "/api/admin/users",
            json={"rfid": "INA005", "nome": "Eva Sem Checkout", "chave": "EE66", "projeto": "P83"},
        )
        assert save_a.status_code == 200
        assert save_b.status_code == 200
        assert save_c.status_code == 200
        assert save_d.status_code == 200
        assert save_e.status_code == 200

        with SessionLocal() as db:
            user_active = get_user_by_rfid(db, "INA001")
            user_active.checkin = True
            user_active.local = "main"
            user_active.time = fixed_now - timedelta(hours=3)
            user_active.last_active_at = fixed_now - timedelta(hours=3)
            user_active.inactivity_days = 0

            user_two = get_user_by_rfid(db, "INA002")
            user_two.checkin = True
            user_two.local = "co83"
            user_two.time = datetime(2024, 4, 3, 11, 0, tzinfo=ZoneInfo(settings.tz_name))
            user_two.last_active_at = user_two.time
            user_two.inactivity_days = 0

            user_three = get_user_by_rfid(db, "INA003")
            user_three.checkin = False
            user_three.local = "main"
            user_three.time = datetime(2024, 4, 3, 9, 0, tzinfo=ZoneInfo(settings.tz_name))
            user_three.last_active_at = user_three.time
            user_three.inactivity_days = 0

            user_four = get_user_by_rfid(db, "INA004")
            user_four.checkin = False
            user_four.local = "co80"
            user_four.time = datetime(2024, 4, 4, 9, 0, tzinfo=ZoneInfo(settings.tz_name))
            user_four.last_active_at = user_four.time
            user_four.inactivity_days = 0

            user_five = get_user_by_rfid(db, "INA005")
            user_five.checkin = True
            user_five.local = "un83"
            user_five.time = datetime(2024, 4, 5, 11, 0, tzinfo=ZoneInfo(settings.tz_name))
            user_five.last_active_at = user_five.time
            user_five.inactivity_days = 0
            db.commit()

        inactive_rows = client.get("/api/admin/inactive")
        assert inactive_rows.status_code == 200
        inactive_payload = inactive_rows.json()
        assert all(isinstance(row["id"], int) and row["id"] > 0 for row in inactive_payload)
        assert [row["nome"] for row in inactive_payload] == ["Ana Inativa", "Bruno Inativo"]
        assert [row["inactivity_days"] for row in inactive_payload] == [3, 3]
        assert [row["latest_action"] for row in inactive_payload] == ["checkin", "checkout"]
        assert inactive_payload[0]["latest_time"].startswith("2024-04-03T11:00:00")
        assert inactive_payload[1]["latest_time"].startswith("2024-04-03T09:00:00")

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        checkin_payload = checkin_rows.json()
        assert any(row["rfid"] == "INA001" and row["id"] > 0 for row in checkin_payload)
        assert all(row["rfid"] != "INA002" for row in checkin_payload)
        assert any(row["rfid"] == "INA005" and row["id"] > 0 for row in checkin_payload)

        missing_checkout_rows = client.get("/api/admin/missing-checkout")
        assert missing_checkout_rows.status_code == 200
        missing_checkout_payload = missing_checkout_rows.json()
        assert [row["rfid"] for row in missing_checkout_payload] == ["INA005"]
        assert missing_checkout_payload[0]["time"].startswith("2024-04-05T11:00:00")

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        checkout_payload = checkout_rows.json()
        assert all(row["rfid"] != "INA003" for row in checkout_payload)
        assert any(row["rfid"] == "INA004" and row["id"] > 0 for row in checkout_payload)


def test_weekends_do_not_increase_inactivity_before_threshold(monkeypatch):
    fixed_now = datetime(2024, 4, 7, 12, 0, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(admin_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(user_activity_module, "now_sgt", lambda: fixed_now)

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "INA010", "nome": "Fabio Fim Semana", "chave": "DD44", "projeto": "P82"},
        )
        assert save_user.status_code == 200

        with SessionLocal() as db:
            weekend_user = get_user_by_rfid(db, "INA010")
            weekend_user.checkin = True
            weekend_user.local = "main"
            weekend_user.time = datetime(2024, 4, 4, 9, 0, tzinfo=ZoneInfo(settings.tz_name))
            weekend_user.last_active_at = weekend_user.time
            weekend_user.inactivity_days = 0
            db.commit()

        inactive_rows = client.get("/api/admin/inactive")
        assert inactive_rows.status_code == 200
        assert inactive_rows.json() == []

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        assert any(row["rfid"] == "INA010" for row in checkin_rows.json())

        missing_checkout_rows = client.get("/api/admin/missing-checkout")
        assert missing_checkout_rows.status_code == 200
        assert any(row["rfid"] == "INA010" for row in missing_checkout_rows.json())

        with SessionLocal() as db:
            weekend_user = get_user_by_rfid(db, "INA010")
            assert weekend_user.inactivity_days == 1


def test_users_past_business_inactivity_threshold_stay_inactive_on_weekends(monkeypatch):
    fixed_now = datetime(2024, 4, 7, 12, 0, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(admin_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(user_activity_module, "now_sgt", lambda: fixed_now)

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_checkin_user = client.post(
            "/api/admin/users",
            json={"rfid": "INA011", "nome": "Gil Checkin Weekend", "chave": "GF11", "projeto": "P80"},
        )
        save_checkout_user = client.post(
            "/api/admin/users",
            json={"rfid": "INA012", "nome": "Helena Checkout Weekend", "chave": "HF12", "projeto": "P82"},
        )
        assert save_checkin_user.status_code == 200
        assert save_checkout_user.status_code == 200

        with SessionLocal() as db:
            weekend_checkin_user = get_user_by_rfid(db, "INA011")
            weekend_checkin_user.checkin = True
            weekend_checkin_user.local = "main"
            weekend_checkin_user.time = datetime(2024, 4, 2, 9, 0, tzinfo=ZoneInfo(settings.tz_name))
            weekend_checkin_user.last_active_at = weekend_checkin_user.time
            weekend_checkin_user.inactivity_days = 0

            weekend_checkout_user = get_user_by_rfid(db, "INA012")
            weekend_checkout_user.checkin = False
            weekend_checkout_user.local = "co80"
            weekend_checkout_user.time = datetime(2024, 4, 2, 8, 0, tzinfo=ZoneInfo(settings.tz_name))
            weekend_checkout_user.last_active_at = weekend_checkout_user.time
            weekend_checkout_user.inactivity_days = 0
            db.commit()

        inactive_rows = client.get("/api/admin/inactive")
        assert inactive_rows.status_code == 200
        inactive_payload = inactive_rows.json()
        assert [row["nome"] for row in inactive_payload] == ["Gil Checkin Weekend", "Helena Checkout Weekend"]
        assert [row["inactivity_days"] for row in inactive_payload] == [3, 3]
        assert [row["latest_action"] for row in inactive_payload] == ["checkin", "checkout"]

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        assert all(row["rfid"] != "INA011" for row in checkin_rows.json())

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        assert all(row["rfid"] != "INA012" for row in checkout_rows.json())

        missing_checkout_rows = client.get("/api/admin/missing-checkout")
        assert missing_checkout_rows.status_code == 200
        assert all(row["rfid"] != "INA011" for row in missing_checkout_rows.json())

        with SessionLocal() as db:
            weekend_checkin_user = get_user_by_rfid(db, "INA011")
            weekend_checkout_user = get_user_by_rfid(db, "INA012")
            assert weekend_checkin_user.inactivity_days == 3
            assert weekend_checkout_user.inactivity_days == 3


def test_admin_presence_lists_follow_latest_activity_even_when_current_state_is_missing_or_stale():
    with SessionLocal() as db:
        stale_user = User(
            rfid="LATE001",
            chave="LT01",
            nome="Usuario Estado Antigo",
            projeto="P80",
            local="main",
            checkin=True,
            time=now_sgt() - timedelta(days=5),
            last_active_at=now_sgt() - timedelta(days=5),
            inactivity_days=0,
        )
        check_event_only_user = User(
            rfid="LATE002",
            chave="LT02",
            nome="Usuario Historico RFID",
            projeto="P83",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt() - timedelta(days=3),
            inactivity_days=0,
        )
        no_activity_user = User(
            rfid="LATE003",
            chave="LT03",
            nome="Usuario Sem Atividade",
            projeto="P82",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add_all([stale_user, check_event_only_user, no_activity_user])
        db.flush()

        db.add(
            UserSyncEvent(
                user_id=stale_user.id,
                chave=stale_user.chave,
                rfid=stale_user.rfid,
                source="android",
                action="checkout",
                projeto=stale_user.projeto,
                local="co80",
                event_time=now_sgt() - timedelta(hours=6),
                created_at=now_sgt(),
                source_request_id=f"android-{uuid.uuid4().hex}",
                device_id=None,
            )
        )
        db.add(
            CheckEvent(
                idempotency_key=f"late-checkout-{uuid.uuid4().hex}",
                source="device",
                rfid="LATE002",
                action="checkout",
                status="queued",
                message="checkout antigo",
                details=None,
                project="P83",
                device_id="ESP32-LATE",
                local="main",
                request_path="/api/scan",
                http_status=202,
                event_time=now_sgt() - timedelta(days=2),
                submitted_at=None,
                retry_count=0,
            )
        )
        db.add(
            CheckEvent(
                idempotency_key=f"late-checkin-{uuid.uuid4().hex}",
                source="device",
                rfid="LATE002",
                action="checkin",
                status="queued",
                message="checkin recente",
                details=None,
                project="P83",
                device_id="ESP32-LATE",
                local="co83",
                request_path="/api/scan",
                http_status=202,
                event_time=now_sgt() - timedelta(hours=3),
                submitted_at=None,
                retry_count=0,
            )
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        checkin_payload = checkin_rows.json()

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        checkout_payload = checkout_rows.json()

        assert all(row["chave"] != "LT01" for row in checkin_payload)
        stale_checkout = next(row for row in checkout_payload if row["chave"] == "LT01")
        assert stale_checkout["local"] == "co80"
        assert stale_checkout["checkin"] is False

        fallback_checkin = next(row for row in checkin_payload if row["chave"] == "LT02")
        assert fallback_checkin["local"] == "co83"
        assert fallback_checkin["checkin"] is True
        assert all(row["chave"] != "LT02" for row in checkout_payload)

        assert all(row["chave"] != "LT03" for row in checkin_payload)
        assert all(row["chave"] != "LT03" for row in checkout_payload)


def test_admin_presence_lists_include_assiduidade_labels():
    with TestClient(app) as client:
        ensure_admin_session(client)

        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ASSI01", "nome": "Usuario RFID", "chave": "AQ11", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        rfid_checkin = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "ASSI01",
                "action": "checkin",
                "device_id": "ESP32-ASSI",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert rfid_checkin.status_code == 200

        retroativo_checkout = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AQ22",
                "projeto": "P82",
                "action": "checkout",
                "informe": "Retroativo",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-forms-assiduidade-{uuid.uuid4().hex}",
            },
        )
        assert retroativo_checkout.status_code == 200

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        checkin_match = next(row for row in checkin_rows.json() if row["chave"] == "AQ11")
        assert checkin_match["assiduidade"] == "Normal"

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        checkout_match = next(row for row in checkout_rows.json() if row["chave"] == "AQ22")
        assert checkout_match["assiduidade"] == "Retroativo"


def test_mobile_sync_autocreates_user_and_updates_state():
    with TestClient(app) as client:
        event_time = now_sgt().isoformat()
        response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP11",
                "projeto": "P80",
                "action": "checkin",
                "event_time": event_time,
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["duplicate"] is False
        assert payload["state"]["found"] is True
        assert payload["state"]["last_checkin_at"] is not None

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AP11")
            assert user.nome == "Oriundo do Aplicativo"
            assert user.rfid is None
            assert user.checkin is True


def test_mobile_submit_autocreates_user_with_app_origin_name():
    client_event_id = f"android-submit-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AS11",
                "projeto": "P83",
                "action": "checkout",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["duplicate"] is False
        assert payload["state"]["found"] is True
        assert payload["state"]["projeto"] == "P83"

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AS11")
            assert user.nome == "Oriundo do Aplicativo"
            assert user.rfid is None
            assert user.projeto == "P83"


def test_mobile_forms_submit_autocreates_user_with_app_origin_name():
    client_event_id = f"android-forms-create-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF11",
                "projeto": "P82",
                "action": "checkin",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["duplicate"] is False
        assert payload["state"]["found"] is True
        assert payload["state"]["projeto"] == "P82"

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AF11")
            assert user.nome == "Oriundo do Aplicativo"
            assert user.rfid is None
            assert user.projeto == "P82"


def test_mobile_forms_submit_skips_same_day_repeated_action():
    first_event_time = now_sgt()
    second_event_time = first_event_time + timedelta(minutes=5)

    with TestClient(app) as client:
        first = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF12",
                "projeto": "P82",
                "action": "checkin",
                "local": "Area A",
                "informe": "normal",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"android-forms-same-day-1-{uuid.uuid4().hex}",
            },
        )
        second = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF12",
                "projeto": "P82",
                "action": "checkin",
                "local": "Area B",
                "informe": "normal",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"android-forms-same-day-2-{uuid.uuid4().hex}",
            },
        )

        assert first.status_code == 200
        assert first.json()["queued_forms"] is True
        assert second.status_code == 200
        assert second.json()["ok"] is True
        assert second.json()["duplicate"] is False
        assert second.json()["queued_forms"] is False

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AF12")
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.chave == "AF12")).scalars().all()
            sync_events = db.execute(
                select(UserSyncEvent).where(UserSyncEvent.chave == "AF12", UserSyncEvent.action == "checkin")
            ).scalars().all()

            assert user.checkin is True
            assert user.local == "Area B"
            assert len(queued) == 1
            assert len(sync_events) == 2


def test_mobile_submit_ignores_provider_checkout_when_evaluating_same_day_repeat():
    first_event_time = datetime(2026, 4, 22, 8, 0, 0, tzinfo=ZoneInfo(settings.tz_name))
    second_event_time = datetime(2026, 4, 22, 11, 0, 0, tzinfo=ZoneInfo(settings.tz_name))

    with TestClient(app) as client:
        first = client.post(
            "/api/mobile/events/submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AS12",
                "projeto": "P82",
                "action": "checkin",
                "local": "Area A",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"mobile-provider-ignore-1-{uuid.uuid4().hex}",
            },
        )
        assert first.status_code == 200, first.text
        assert first.json()["queued_forms"] is True

        provider_checkout = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "AS12",
                "nome": "USUARIO MOBILE IGNORA PROVIDER",
                "projeto": "P82",
                "atividade": "check-out",
                "informe": "normal",
                "data": "22/04/2026",
                "hora": "10:00:00",
            },
        )
        assert provider_checkout.status_code == 200, provider_checkout.text
        assert provider_checkout.json()["updated_current_state"] is True

        second = client.post(
            "/api/mobile/events/submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AS12",
                "projeto": "P82",
                "action": "checkin",
                "local": "Area B",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"mobile-provider-ignore-2-{uuid.uuid4().hex}",
            },
        )
        assert second.status_code == 200, second.text
        assert second.json()["queued_forms"] is False

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AS12")
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.chave == "AS12").order_by(FormsSubmission.id)
            ).scalars().all()

            assert user.checkin is True
            assert user.local == "Area B"
            assert len(queued) == 1


def test_mobile_forms_submit_requeues_same_action_after_singapore_midnight():
    first_event_time = now_sgt() - timedelta(days=1)
    second_event_time = now_sgt()

    with TestClient(app) as client:
        first = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF13",
                "projeto": "P83",
                "action": "checkin",
                "local": "Area A",
                "informe": "normal",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"android-forms-next-day-1-{uuid.uuid4().hex}",
            },
        )
        second = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF13",
                "projeto": "P83",
                "action": "checkin",
                "local": "Area C",
                "informe": "normal",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"android-forms-next-day-2-{uuid.uuid4().hex}",
            },
        )

        assert first.status_code == 200
        assert second.status_code == 200
        assert second.json()["ok"] is True
        assert second.json()["duplicate"] is False
        assert second.json()["queued_forms"] is True

        with SessionLocal() as db:
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.chave == "AF13")).scalars().all()
            user = get_user_by_chave(db, "AF13")

            assert len(queued) == 2
            assert user.checkin is True
            assert user.local == "Area C"


def test_mobile_state_returns_current_location_for_checked_in_user():
    with TestClient(app) as client:
        submit_response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF14",
                "projeto": "P80",
                "action": "checkin",
                "local": "Base P80",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-forms-state-{uuid.uuid4().hex}",
            },
        )
        assert submit_response.status_code == 200

        state_response = client.get(
            "/api/mobile/state?chave=AF14",
            headers=MOBILE_HEADERS,
        )
        assert state_response.status_code == 200
        payload = state_response.json()
        assert payload["found"] is True
        assert payload["current_action"] == "checkin"
        assert payload["current_local"] == "Base P80"


def test_admin_checkin_list_accepts_mobile_user_without_rfid():
    with TestClient(app) as client:
        created = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP15",
                "projeto": "P80",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert created.status_code == 200

        ensure_admin_session(client)
        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        payload = checkin_rows.json()
        matched = next(row for row in payload if row["chave"] == "AP15")
        assert matched["rfid"] is None
        assert matched["checkin"] is True


def test_mobile_sync_notifies_admin_realtime_subscribers():
    subscriber_id, queue = admin_updates_broker.subscribe()

    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/mobile/events/sync",
                headers=MOBILE_HEADERS,
                json={
                    "chave": "AP16",
                    "projeto": "P83",
                    "action": "checkout",
                    "event_time": now_sgt().isoformat(),
                    "client_event_id": f"android-{uuid.uuid4().hex}",
                },
            )

        assert response.status_code == 200
        payload = queue.get_nowait()
        assert '"reason": "checkout"' in payload
    finally:
        admin_updates_broker.unsubscribe(subscriber_id)


def test_mobile_sync_is_idempotent_for_same_event_id():
    with TestClient(app) as client:
        client_event_id = f"android-{uuid.uuid4().hex}"
        payload = {
            "chave": "AP12",
            "projeto": "P83",
            "action": "checkout",
            "event_time": now_sgt().isoformat(),
            "client_event_id": client_event_id,
        }
        first = client.post("/api/mobile/events/sync", headers=MOBILE_HEADERS, json=payload)
        second = client.post("/api/mobile/events/sync", headers=MOBILE_HEADERS, json=payload)
        assert first.status_code == 200
        assert second.status_code == 200
        assert second.json()["duplicate"] is True


def test_admin_can_edit_mobile_created_user_without_rfid():
    with TestClient(app) as client:
        created = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP13",
                "projeto": "P80",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert created.status_code == 200

        ensure_admin_session(client)
        users = client.get("/api/admin/users")
        assert users.status_code == 200
        created_user = next(row for row in users.json() if row["chave"] == "AP13")
        assert created_user["rfid"] is None

        updated = client.post(
            "/api/admin/users",
            json={
                "user_id": created_user["id"],
                "nome": "Nome Corrigido",
                "chave": "AP13",
                "projeto": "P83",
            },
        )
        assert updated.status_code == 200

        users_after = client.get("/api/admin/users")
        assert users_after.status_code == 200
        updated_user = next(row for row in users_after.json() if row["id"] == created_user["id"])
        assert updated_user["nome"] == "Nome Corrigido"
        assert updated_user["projeto"] == "P83"


def test_admin_can_attach_rfid_to_mobile_created_user_by_unique_chave():
    with TestClient(app) as client:
        created = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP14",
                "projeto": "P82",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert created.status_code == 200

        ensure_admin_session(client)
        attached = client.post(
            "/api/admin/users",
            json={
                "rfid": "APPRFID1",
                "nome": "Nome Ajustado",
                "chave": "AP14",
                "projeto": "P82",
            },
        )
        assert attached.status_code == 200
        assert attached.json()["linked_existing_user"] is True

        users = client.get("/api/admin/users")
        assert users.status_code == 200
        matched = [row for row in users.json() if row["chave"] == "AP14"]
        assert len(matched) == 1
        assert matched[0]["rfid"] == "APPRFID1"
        assert matched[0]["nome"] == "Nome Ajustado"
        assert matched[0]["projeto"] == "P82"


def test_pending_registration_links_rfid_to_existing_mobile_user_by_chave():
    with TestClient(app) as client:
        created = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP17",
                "projeto": "P80",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert created.status_code == 200

        ensure_admin_session(client)
        scan_pending = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "LINKRFID17",
                "action": "checkin",
                "device_id": "ESP32-LINK",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_pending.status_code == 200
        assert scan_pending.json()["outcome"] == "pending_registration"

        attached = client.post(
            "/api/admin/users",
            json={
                "rfid": "LINKRFID17",
                "nome": "Nome Vindo da Pendência",
                "chave": "AP17",
                "projeto": "P83",
            },
        )
        assert attached.status_code == 200
        assert attached.json()["linked_existing_user"] is True

        users = client.get("/api/admin/users")
        assert users.status_code == 200
        matched = [row for row in users.json() if row["chave"] == "AP17"]
        assert len(matched) == 1
        assert matched[0]["rfid"] == "LINKRFID17"
        assert matched[0]["nome"] == "Nome Vindo da Pendência"
        assert matched[0]["projeto"] == "P83"

        pending = client.get("/api/admin/pending")
        assert pending.status_code == 200
        assert all(row["rfid"] != "LINKRFID17" for row in pending.json())


def test_mobile_state_reflects_rfid_scan_history(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "RFIDSYNC1", "nome": "Usuario Sync", "chave": "SY11", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_response = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "RFIDSYNC1",
                "action": "checkin",
                "device_id": "ESP32-SYNC",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_response.status_code == 200

        mobile_state = client.get("/api/mobile/state?chave=SY11", headers=MOBILE_HEADERS)
        assert mobile_state.status_code == 200
        payload = mobile_state.json()
        assert payload["found"] is True
        assert payload["current_action"] == "checkin"
        assert payload["last_checkin_at"] is not None


def test_mobile_forms_submit_accepts_retroativo_and_persists_ontime_false(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    client_event_id = f"android-forms-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "RT11",
                "projeto": "P82",
                "action": "checkin",
                "informe": "Retroativo",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["duplicate"] is False
        assert payload["queued_forms"] is True
        assert payload["state"]["current_action"] == "checkin"

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.request_id == client_event_id)
            ).scalar_one()
            sync_event = db.execute(
                select(UserSyncEvent).where(
                    UserSyncEvent.source == "android_forms",
                    UserSyncEvent.source_request_id == client_event_id,
                )
            ).scalar_one()
            assert queued.ontime is False
            assert sync_event.ontime is False

        processed = process_forms_submission_queue_once()
        assert processed >= 1

        ensure_admin_session(client)
        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(
            event["request_path"] == "/api/mobile/events/forms-submit"
            and event["ontime"] is False
            and event["status"] == "queued"
            and event["chave"] == "RT11"
            for event in events.json()
        )
        assert any(
            event["source"] == "forms"
            and event["ontime"] is False
            and event["status"] == "success"
            and event["chave"] == "RT11"
            for event in events.json()
        )


def test_mobile_forms_submit_is_idempotent_for_same_event_id():
    with TestClient(app) as client:
        client_event_id = f"android-forms-{uuid.uuid4().hex}"
        payload = {
            "chave": "RT12",
            "projeto": "P80",
            "action": "checkout",
            "informe": "normal",
            "event_time": now_sgt().isoformat(),
            "client_event_id": client_event_id,
        }
        first = client.post("/api/mobile/events/forms-submit", headers=MOBILE_HEADERS, json=payload)
        second = client.post("/api/mobile/events/forms-submit", headers=MOBILE_HEADERS, json=payload)
        assert first.status_code == 200
        assert second.status_code == 200
        assert second.json()["duplicate"] is True


def test_mobile_check_page_is_served_on_user_path():
    with TestClient(app) as client:
        response = client.get("/user")
        assert response.status_code == 200
        assert "Registrar" in response.text
        assert "Senha" in response.text
        assert "Local" in response.text
        assert "Atualizar local" in response.text
        assert "Chave" in response.text
        assert "Último Check-In" in response.text
        assert "Último Check-Out" in response.text
        assert "/api/web/check" in response.text
        assert "/api/web/auth/status" in response.text
        assert "/api/web/auth/register-password" in response.text
        assert "/api/web/auth/register-user" in response.text
        assert "/api/web/auth/login" in response.text
        assert "/api/web/auth/change-password" in response.text
        assert "/api/web/auth/logout" in response.text
        assert "/api/web/check/state" in response.text
        assert "/api/web/check/location" in response.text
        assert "Cadastro de Usuário" in response.text


def test_transport_page_is_served_on_transport_path():
    with TestClient(app) as client:
        response = client.get("/transport")
        assert response.status_code == 200
        assert "User List" in response.text
        assert 'data-toggle-request-section="extra"' in response.text
        assert 'data-toggle-request-section="weekend"' in response.text
        assert 'data-toggle-request-section="regular"' in response.text
        assert "Regular Transport List" in response.text
        assert "Weekend Transport List" in response.text
        assert "Extra Transport List" in response.text
        assert "System Support" in response.text
        assert "Tamer Salmem (HR70)" in response.text
        assert "Home to Work" in response.text
        assert "Work to Home" in response.text
        assert 'id="tela01menu"' in response.text
        assert 'id="tela01main_dir_down"' in response.text
        assert 'data-date-link' in response.text
        assert 'data-date-today' not in response.text


def test_transport_dashboard_groups_requests_by_selected_date_and_assignment_status():
    friday = date(2026, 4, 17)
    saturday = date(2026, 4, 18)
    timestamp = now_sgt()

    with SessionLocal() as db:
        regular_vehicle = Vehicle(placa="REG9001", tipo="van", color="Silver", lugares=12, tolerance=8, service_scope="regular")
        weekend_vehicle = Vehicle(placa="WKD9001", tipo="van", color="Black", lugares=10, tolerance=15, service_scope="weekend")
        extra_vehicle = Vehicle(placa="EXT9001", tipo="carro", color="Red", lugares=4, tolerance=6, service_scope="extra")
        db.add_all(
            [
                Workplace(workplace="Transport Hub Alpha", address="1 Harbour Front", zip="111111", country="Singapore"),
                regular_vehicle,
                weekend_vehicle,
                extra_vehicle,
            ]
        )
        db.flush()
        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="home_to_work",
            recurrence_kind="matching_weekday",
            weekday=saturday.weekday(),
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="work_to_home",
            recurrence_kind="matching_weekday",
            weekday=saturday.weekday(),
        )
        add_transport_schedule(
            db,
            vehicle=extra_vehicle,
            service_scope="extra",
            route_kind="home_to_work",
            recurrence_kind="single_date",
            service_date=friday,
        )

        regular_user = User(
            rfid=None,
            nome="Regular Rider",
            chave="TD01",
            projeto="P80",
            workplace="Transport Hub Alpha",
            end_rua="10 Regular Street",
            zip="900001",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        weekend_user = User(
            rfid=None,
            nome="Weekend Rider",
            chave="TD02",
            projeto="P82",
            workplace="Transport Hub Alpha",
            end_rua="20 Weekend Street",
            zip="900002",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        extra_user = User(
            rfid=None,
            nome="Extra Rider",
            chave="TD03",
            projeto="P83",
            workplace="Transport Hub Alpha",
            end_rua="30 Extra Street",
            zip="900003",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add_all([regular_user, weekend_user, extra_user])
        db.flush()

        regular_request = TransportRequest(
            user_id=regular_user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:30",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        weekend_request = TransportRequest(
            user_id=weekend_user.id,
            request_kind="weekend",
            recurrence_kind="weekend",
            requested_time="08:10",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        extra_request = TransportRequest(
            user_id=extra_user.id,
            request_kind="extra",
            recurrence_kind="single_date",
            requested_time="09:15",
            single_date=friday,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add_all([regular_request, weekend_request, extra_request])
        db.flush()
        db.add(
            TransportAssignment(
                request_id=regular_request.id,
                service_date=friday,
                route_kind="home_to_work",
                vehicle_id=regular_vehicle.id,
                status="confirmed",
                response_message=None,
                assigned_by_admin_id=None,
                created_at=timestamp,
                updated_at=timestamp,
                notified_at=None,
            )
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        friday_response = client.get("/api/transport/dashboard", params={"service_date": friday.isoformat()})
        saturday_response = client.get("/api/transport/dashboard", params={"service_date": saturday.isoformat()})
        friday_work_to_home_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "work_to_home"},
        )

    assert friday_response.status_code == 200
    assert saturday_response.status_code == 200
    assert friday_work_to_home_response.status_code == 200

    friday_payload = friday_response.json()
    saturday_payload = saturday_response.json()
    friday_work_to_home_payload = friday_work_to_home_response.json()

    assert [row["chave"] for row in friday_payload["regular_requests"]] == ["TD01"]
    assert friday_payload["regular_requests"][0]["assignment_status"] == "confirmed"
    assert friday_payload["regular_requests"][0]["assigned_vehicle"]["placa"] == "REG9001"
    assert [row["chave"] for row in friday_payload["extra_requests"]] == ["TD03"]
    assert friday_payload["extra_requests"][0]["assignment_status"] == "pending"
    assert [row["chave"] for row in friday_payload["weekend_requests"]] == ["TD02"]
    assert friday_payload["weekend_requests"][0]["assignment_status"] == "pending"
    assert friday_payload["weekend_requests"][0]["service_date"] == saturday.isoformat()
    assert [row["placa"] for row in friday_payload["regular_vehicles"]] == ["REG9001"]
    assert [row["placa"] for row in friday_payload["extra_vehicles"]] == ["EXT9001"]
    assert [row["route_kind"] for row in friday_payload["extra_vehicles"]] == ["home_to_work"]
    assert [row["placa"] for row in friday_payload["regular_vehicle_registry"]] == ["REG9001"]
    assert friday_payload["regular_vehicle_registry"][0]["assigned_count"] == 1
    assert [row["placa"] for row in friday_payload["weekend_vehicle_registry"]] == ["WKD9001"]
    assert friday_payload["weekend_vehicle_registry"][0]["assigned_count"] == 0
    assert [row["placa"] for row in friday_payload["extra_vehicle_registry"]] == ["EXT9001"]
    assert friday_payload["extra_vehicle_registry"][0]["service_date"] == friday.isoformat()
    assert friday_payload["extra_vehicle_registry"][0]["route_kind"] == "home_to_work"
    assert friday_payload["selected_route"] == "home_to_work"

    assert [row["chave"] for row in saturday_payload["regular_requests"]] == ["TD01"]
    assert saturday_payload["regular_requests"][0]["assignment_status"] == "pending"
    assert saturday_payload["regular_requests"][0]["assigned_vehicle"] is None
    assert [row["chave"] for row in saturday_payload["weekend_requests"]] == ["TD02"]
    assert [row["chave"] for row in saturday_payload["extra_requests"]] == ["TD03"]
    assert saturday_payload["extra_requests"][0]["assignment_status"] == "pending"
    assert saturday_payload["extra_requests"][0]["service_date"] == friday.isoformat()
    assert [row["placa"] for row in saturday_payload["weekend_vehicles"]] == ["WKD9001"]

    assert [row["placa"] for row in friday_work_to_home_payload["regular_vehicles"]] == ["REG9001"]
    assert [row["placa"] for row in friday_work_to_home_payload["extra_vehicles"]] == ["EXT9001"]
    assert [row["route_kind"] for row in friday_work_to_home_payload["extra_vehicles"]] == ["home_to_work"]
    assert [row["placa"] for row in friday_work_to_home_payload["extra_vehicle_registry"]] == ["EXT9001"]
    assert friday_work_to_home_payload["regular_requests"][0]["assignment_status"] == "confirmed"
    assert friday_work_to_home_payload["regular_requests"][0]["assigned_vehicle"]["placa"] == "REG9001"


def test_transport_vehicle_registration_creates_route_aware_schedules():
    friday = date(2026, 4, 17)
    sunday = date(2026, 4, 19)

    with TestClient(app) as client:
        ensure_admin_session(client)

        extra_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "extra",
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "departure_time": "17:45",
                "tipo": "carro",
                "placa": "EXT7001",
                "color": "Red",
                "lugares": 4,
                "tolerance": 6,
            },
        )
        weekend_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "weekend",
                "service_date": sunday.isoformat(),
                "every_saturday": True,
                "every_sunday": True,
                "tipo": "van",
                "placa": "WKD7001",
                "color": "Black",
                "lugares": 10,
                "tolerance": 12,
            },
        )
        regular_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": friday.isoformat(),
                "tipo": "minivan",
                "placa": "REG7001",
                "color": "White",
                "lugares": 7,
                "tolerance": 9,
            },
        )

    assert extra_response.status_code == 200
    assert weekend_response.status_code == 200
    assert regular_response.status_code == 200

    with SessionLocal() as db:
        extra_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "EXT7001")).scalar_one()
        weekend_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "WKD7001")).scalar_one()
        regular_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "REG7001")).scalar_one()

        extra_schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == extra_vehicle.id)
            .order_by(TransportVehicleSchedule.id)
        ).scalars().all()
        weekend_schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == weekend_vehicle.id)
            .order_by(TransportVehicleSchedule.route_kind)
        ).scalars().all()
        regular_schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == regular_vehicle.id)
            .order_by(TransportVehicleSchedule.route_kind)
        ).scalars().all()

    assert len(extra_schedules) == 1
    assert extra_schedules[0].route_kind == "work_to_home"
    assert extra_schedules[0].recurrence_kind == "single_date"
    assert extra_schedules[0].service_date == friday

    assert len(weekend_schedules) == 4
    assert {row.route_kind for row in weekend_schedules} == {"home_to_work", "work_to_home"}
    assert all(row.recurrence_kind == "matching_weekday" for row in weekend_schedules)
    assert {row.weekday for row in weekend_schedules} == {5, 6}
    assert all(row.service_date == sunday for row in weekend_schedules)

    assert len(regular_schedules) == 2
    assert {row.route_kind for row in regular_schedules} == {"home_to_work", "work_to_home"}
    assert all(row.recurrence_kind == "weekday" for row in regular_schedules)
    assert all(row.service_date == friday for row in regular_schedules)


def test_transport_vehicle_registration_accepts_long_plate_with_special_characters():
    friday = date(2026, 4, 17)
    long_plate = "SG-1234.ABC-789"

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": friday.isoformat(),
                "tipo": "van",
                "placa": long_plate,
                "color": "Gray",
                "lugares": 11,
                "tolerance": 9,
            },
        )

    assert response.status_code == 200, response.text

    with SessionLocal() as db:
        persisted_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == long_plate)).scalar_one()
        assert persisted_vehicle.placa == long_plate


def test_transport_regular_vehicle_registration_supports_selected_weekdays():
    saturday = date(2026, 4, 18)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": saturday.isoformat(),
                "every_monday": True,
                "every_thursday": True,
                "tipo": "minivan",
                "placa": "REG7021",
                "color": "Silver",
                "lugares": 7,
                "tolerance": 9,
            },
        )

    assert response.status_code == 200

    with SessionLocal() as db:
        regular_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "REG7021")).scalar_one()
        regular_schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == regular_vehicle.id)
            .order_by(TransportVehicleSchedule.route_kind, TransportVehicleSchedule.weekday)
        ).scalars().all()

    assert len(regular_schedules) == 4
    assert {row.route_kind for row in regular_schedules} == {"home_to_work", "work_to_home"}
    assert all(row.recurrence_kind == "matching_weekday" for row in regular_schedules)
    assert {row.weekday for row in regular_schedules} == {0, 3}
    assert all(row.service_date == saturday for row in regular_schedules)


def test_transport_extra_vehicle_registration_requires_departure_time():
    friday = date(2026, 4, 17)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "extra",
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "tipo": "carro",
                "placa": "EXT7011",
                "color": "Gray",
                "lugares": 4,
                "tolerance": 6,
            },
        )

    assert response.status_code == 422
    assert "departure_time is required for extra vehicles" in json.dumps(response.json()["detail"])


def test_transport_weekend_vehicle_registration_requires_persistent_weekday_selection():
    saturday = date(2026, 4, 18)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "weekend",
                "service_date": saturday.isoformat(),
                "tipo": "van",
                "placa": "WKD7011",
                "color": "Gray",
                "lugares": 8,
                "tolerance": 10,
            },
        )

    assert response.status_code == 422
    assert "Weekend vehicles must be persistent" in json.dumps(response.json()["detail"])


def test_transport_weekend_vehicle_registration_can_start_from_weekday_dashboard_date():
    wednesday = date(2026, 4, 22)
    friday = date(2026, 4, 24)
    saturday = date(2026, 4, 25)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "weekend",
                "service_date": wednesday.isoformat(),
                "every_saturday": True,
                "tipo": "van",
                "placa": "WKD7022",
                "color": "Blue",
                "lugares": 10,
                "tolerance": 10,
            },
        )
        wednesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": wednesday.isoformat(), "route_kind": "home_to_work"},
        )
        friday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )
        saturday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert wednesday_dashboard.status_code == 200
    assert friday_dashboard.status_code == 200
    assert saturday_dashboard.status_code == 200
    assert all(row["placa"] != "WKD7022" for row in wednesday_dashboard.json()["weekend_vehicles"])
    assert all(row["placa"] != "WKD7022" for row in friday_dashboard.json()["weekend_vehicles"])
    assert any(row["placa"] == "WKD7022" for row in saturday_dashboard.json()["weekend_vehicles"])


def test_transport_regular_vehicle_registration_can_start_from_weekend_dashboard_date():
    saturday = date(2026, 4, 18)
    monday = date(2026, 4, 20)
    tuesday = date(2026, 4, 21)
    wednesday = date(2026, 4, 22)
    thursday = date(2026, 4, 23)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": saturday.isoformat(),
                "every_monday": True,
                "every_tuesday": True,
                "every_thursday": True,
                "tipo": "carro",
                "placa": "REG7022",
                "color": "White",
                "lugares": 4,
                "tolerance": 7,
            },
        )
        saturday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "home_to_work"},
        )
        monday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "home_to_work"},
        )
        tuesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": tuesday.isoformat(), "route_kind": "home_to_work"},
        )
        wednesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": wednesday.isoformat(), "route_kind": "home_to_work"},
        )
        thursday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": thursday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert saturday_dashboard.status_code == 200
    assert monday_dashboard.status_code == 200
    assert tuesday_dashboard.status_code == 200
    assert wednesday_dashboard.status_code == 200
    assert thursday_dashboard.status_code == 200
    assert all(row["placa"] != "REG7022" for row in saturday_dashboard.json()["regular_vehicles"])
    assert any(row["placa"] == "REG7022" for row in monday_dashboard.json()["regular_vehicles"])
    assert any(row["placa"] == "REG7022" for row in tuesday_dashboard.json()["regular_vehicles"])
    assert all(row["placa"] != "REG7022" for row in wednesday_dashboard.json()["regular_vehicles"])
    assert any(row["placa"] == "REG7022" for row in thursday_dashboard.json()["regular_vehicles"])


def test_transport_vehicle_registration_conflict_messages_are_in_english():
    friday = date(2026, 4, 17)

    with TestClient(app) as client:
        ensure_admin_session(client)

        first_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "extra",
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "departure_time": "17:45",
                "tipo": "carro",
                "placa": "EXT7010",
                "color": "Red",
                "lugares": 4,
                "tolerance": 6,
            },
        )
        duplicate_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": friday.isoformat(),
                "tipo": "carro",
                "placa": "EXT7010",
                "color": "Red",
                "lugares": 4,
                "tolerance": 6,
            },
        )

    assert first_response.status_code == 200
    assert duplicate_response.status_code == 409
    assert duplicate_response.json()["detail"] == (
        "A vehicle with this plate already exists in another list: "
        "Extra list (Work to Home on 2026-04-17)."
    )


def test_transport_vehicle_registration_reuses_plate_after_past_single_date_schedule():
    friday = date(2026, 4, 17)
    monday = date(2026, 4, 20)

    with TestClient(app) as client:
        ensure_admin_session(client)

        first_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "extra",
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "departure_time": "17:45",
                "tipo": "carro",
                "placa": "AAA0000A",
                "color": "Gray",
                "lugares": 4,
                "tolerance": 6,
            },
        )
        reused_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": monday.isoformat(),
                "tipo": "minivan",
                "placa": "AAA0000A",
                "color": "Gray",
                "lugares": 7,
                "tolerance": 10,
            },
        )

    assert first_response.status_code == 200
    assert reused_response.status_code == 200

    with SessionLocal() as db:
        vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "AAA0000A")).scalar_one()
        schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == vehicle.id)
            .order_by(TransportVehicleSchedule.id)
        ).scalars().all()

    assert vehicle.service_scope == "regular"
    assert vehicle.tipo == "minivan"
    assert len(schedules) == 3
    assert [row.is_active for row in schedules] == [False, True, True]
    assert schedules[0].recurrence_kind == "single_date"
    assert schedules[0].service_date == friday
    assert {row.route_kind for row in schedules[1:] if row.is_active} == {"home_to_work", "work_to_home"}


def test_transport_vehicle_delete_purges_vehicle_and_returns_requests_to_pending():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Delete Hub", address="7 Delete Road", zip="707070", country="Singapore"))
        vehicle = Vehicle(placa="DEL1700", tipo="carro", color="Red", lugares=4, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()
        first_schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        second_schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.flush()
        db.add(
            TransportVehicleScheduleException(
                vehicle_schedule_id=second_schedule.id,
                service_date=friday,
                created_at=timestamp,
            )
        )

        user = User(
            rfid=None,
            nome="Delete Route Rider",
            chave="TD17",
            projeto="P80",
            workplace="Delete Hub",
            placa="DEL1700",
            end_rua="17 Delete Street",
            zip="170170",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="08:15",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        user_id = user.id
        vehicle_id = vehicle.id
        schedule_id = first_schedule.id
        schedule_ids = [first_schedule.id, second_schedule.id]

    with TestClient(app) as client:
        ensure_admin_session(client)
        assigned = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        assert assigned.status_code == 200

        removed = client.delete(
            f"/api/transport/vehicles/{schedule_id}",
            params={"service_date": friday.isoformat()},
        )
        assert removed.status_code == 200

        refreshed_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )
        paired_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "work_to_home"},
        )

    assert refreshed_dashboard.status_code == 200
    assert paired_dashboard.status_code == 200
    request_row = next(row for row in refreshed_dashboard.json()["regular_requests"] if row["id"] == request_id)
    paired_request_row = next(row for row in paired_dashboard.json()["regular_requests"] if row["id"] == request_id)
    assert request_row["assignment_status"] == "pending"
    assert paired_request_row["assignment_status"] == "pending"
    assert request_row["assigned_vehicle"] is None
    assert paired_request_row["assigned_vehicle"] is None
    assert all(row["placa"] != "DEL1700" for row in refreshed_dashboard.json()["regular_vehicles"])
    assert all(row["placa"] != "DEL1700" for row in paired_dashboard.json()["regular_vehicles"])

    with SessionLocal() as db:
        user_row = db.get(User, user_id)
        vehicle_row = db.get(Vehicle, vehicle_id)
        schedules = db.execute(
            select(TransportVehicleSchedule).where(TransportVehicleSchedule.vehicle_id == vehicle_id)
        ).scalars().all()
        schedule_exceptions = db.execute(
            select(TransportVehicleScheduleException).where(
                TransportVehicleScheduleException.vehicle_schedule_id.in_(schedule_ids)
            )
        ).scalars().all()
        assignments = db.execute(
            select(TransportAssignment).where(TransportAssignment.request_id == request_id)
        ).scalars().all()

    assert user_row is not None
    assert user_row.placa is None
    assert vehicle_row is None
    assert schedules == []
    assert schedule_exceptions == []
    assert assignments == []


def test_transport_vehicle_delete_clears_user_plate_references_for_all_vehicle_lists():
    scenarios = [
        {
            "scope": "regular",
            "plate": "CLR1701",
            "user_key": "VR17",
            "selected_date": date(2026, 4, 17),
            "dashboard_key": "regular_vehicles",
            "schedule_specs": [
                {"route_kind": "home_to_work", "recurrence_kind": "weekday"},
                {"route_kind": "work_to_home", "recurrence_kind": "weekday"},
            ],
        },
        {
            "scope": "weekend",
            "plate": "CLW1901",
            "user_key": "VW19",
            "selected_date": date(2026, 4, 19),
            "dashboard_key": "weekend_vehicles",
            "schedule_specs": [
                {"route_kind": "home_to_work", "recurrence_kind": "matching_weekday", "weekday": 6},
                {"route_kind": "work_to_home", "recurrence_kind": "matching_weekday", "weekday": 6},
            ],
        },
        {
            "scope": "extra",
            "plate": "CLX2101",
            "user_key": "VX21",
            "selected_date": date(2026, 4, 21),
            "dashboard_key": "extra_vehicles",
            "schedule_specs": [
                {"route_kind": "home_to_work", "recurrence_kind": "single_date", "service_date": date(2026, 4, 21)},
            ],
        },
    ]

    created_rows = []
    with SessionLocal() as db:
        for index, scenario in enumerate(scenarios, start=1):
            workplace = Workplace(
                workplace=f"Delete Scope Hub {index}",
                address=f"{index} Scope Road",
                zip=f"70{index:04d}",
                country="Singapore",
            )
            db.add(workplace)
            db.flush()

            vehicle = Vehicle(
                placa=scenario["plate"],
                tipo="carro",
                color="Blue",
                lugares=4,
                tolerance=6,
                service_scope=scenario["scope"],
            )
            db.add(vehicle)
            db.flush()

            first_schedule_id = None
            for schedule_spec in scenario["schedule_specs"]:
                schedule = add_transport_schedule(
                    db,
                    vehicle=vehicle,
                    service_scope=scenario["scope"],
                    route_kind=schedule_spec["route_kind"],
                    recurrence_kind=schedule_spec["recurrence_kind"],
                    service_date=schedule_spec.get("service_date"),
                    weekday=schedule_spec.get("weekday"),
                )
                if first_schedule_id is None:
                    first_schedule_id = schedule.id

            user = User(
                rfid=None,
                nome=f"Delete Scope Rider {index}",
                chave=scenario["user_key"],
                projeto="P80",
                workplace=workplace.workplace,
                placa=scenario["plate"],
                end_rua=f"{index} Delete Avenue",
                zip=f"80{index:04d}",
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(user)
            db.flush()

            created_rows.append(
                {
                    "dashboard_key": scenario["dashboard_key"],
                    "plate": scenario["plate"],
                    "schedule_id": first_schedule_id,
                    "selected_date": scenario["selected_date"],
                    "user_id": user.id,
                    "vehicle_id": vehicle.id,
                }
            )

        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        for created_row in created_rows:
            removed = client.delete(
                f"/api/transport/vehicles/{created_row['schedule_id']}",
                params={"service_date": created_row["selected_date"].isoformat()},
            )
            assert removed.status_code == 200

            dashboard = client.get(
                "/api/transport/dashboard",
                params={
                    "service_date": created_row["selected_date"].isoformat(),
                    "route_kind": "home_to_work",
                },
            )
            assert dashboard.status_code == 200
            assert all(
                row["placa"] != created_row["plate"]
                for row in dashboard.json()[created_row["dashboard_key"]]
            )

    with SessionLocal() as db:
        for created_row in created_rows:
            user_row = db.get(User, created_row["user_id"])
            vehicle_row = db.get(Vehicle, created_row["vehicle_id"])
            schedule_rows = db.execute(
                select(TransportVehicleSchedule).where(
                    TransportVehicleSchedule.vehicle_id == created_row["vehicle_id"]
                )
            ).scalars().all()

            assert user_row is not None
            assert user_row.placa is None
            assert vehicle_row is None
            assert schedule_rows == []


def test_transport_vehicle_delete_removes_legacy_notifications_before_assignments():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Legacy Notification Hub",
            address="9 Legacy Avenue",
            zip="919191",
            country="Singapore",
        )
        vehicle = Vehicle(placa="LEG1701", tipo="onibus", color="White", lugares=40, tolerance=12, service_scope="regular")
        db.add_all([workplace, vehicle])
        db.flush()

        schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Legacy Notification Rider",
            chave="LN17",
            projeto="P80",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="17 Legacy Lane",
            zip="171717",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="08:30",
            single_date=None,
            created_via="admin",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.flush()

        assignment = TransportAssignment(
            request_id=request_row.id,
            service_date=friday,
            route_kind="home_to_work",
            vehicle_id=vehicle.id,
            status="confirmed",
            response_message=None,
            assigned_by_admin_id=None,
            created_at=timestamp,
            updated_at=timestamp,
            notified_at=None,
        )
        db.add(assignment)
        db.commit()

        user_id = user.id
        vehicle_id = vehicle.id
        request_id = request_row.id
        assignment_id = assignment.id
        schedule_id = schedule.id

    with engine.begin() as conn:
        conn.execute(text("DROP TABLE IF EXISTS transport_notifications"))
        conn.execute(
            text(
                """
                CREATE TABLE transport_notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    user_id INTEGER NOT NULL REFERENCES users(id),
                    chat_id VARCHAR(120),
                    request_id INTEGER REFERENCES transport_requests(id),
                    assignment_id INTEGER REFERENCES transport_assignments(id),
                    message VARCHAR(500) NOT NULL,
                    status VARCHAR(16) NOT NULL DEFAULT 'pending',
                    created_at DATETIME NOT NULL,
                    sent_at DATETIME
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO transport_notifications (
                    user_id,
                    chat_id,
                    request_id,
                    assignment_id,
                    message,
                    status,
                    created_at,
                    sent_at
                ) VALUES (
                    :user_id,
                    NULL,
                    :request_id,
                    :assignment_id,
                    :message,
                    'pending',
                    :created_at,
                    NULL
                )
                """
            ),
            {
                "user_id": user_id,
                "request_id": request_id,
                "assignment_id": assignment_id,
                "message": "Legacy notification row",
                "created_at": timestamp.isoformat(),
            },
        )

    try:
        with TestClient(app) as client:
            ensure_admin_session(client)
            removed = client.delete(
                f"/api/transport/vehicles/{schedule_id}",
                params={"service_date": friday.isoformat()},
            )
            assert removed.status_code == 200, removed.text

        with engine.begin() as conn:
            remaining_notifications = conn.execute(
                text("SELECT COUNT(*) FROM transport_notifications WHERE assignment_id = :assignment_id"),
                {"assignment_id": assignment_id},
            ).scalar_one()

        with SessionLocal() as db:
            user_row = db.get(User, user_id)
            vehicle_row = db.get(Vehicle, vehicle_id)
            assignment_row = db.get(TransportAssignment, assignment_id)

        assert remaining_notifications == 0
        assert user_row is not None
        assert user_row.placa is None
        assert vehicle_row is None
        assert assignment_row is None
    finally:
        with engine.begin() as conn:
            conn.execute(text("DROP TABLE IF EXISTS transport_notifications"))


def test_transport_vehicle_delete_purges_generic_legacy_foreign_key_dependencies():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Generic Legacy Hub",
            address="8 Generic Avenue",
            zip="818181",
            country="Singapore",
        )
        vehicle = Vehicle(placa="GLG1701", tipo="onibus", color="Blue", lugares=36, tolerance=10, service_scope="regular")
        db.add_all([workplace, vehicle])
        db.flush()

        schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Generic Legacy Rider",
            chave="GL17",
            projeto="P80",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="17 Generic Road",
            zip="171818",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="08:50",
            single_date=None,
            created_via="admin",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.flush()

        assignment = TransportAssignment(
            request_id=request_row.id,
            service_date=friday,
            route_kind="home_to_work",
            vehicle_id=vehicle.id,
            status="confirmed",
            response_message=None,
            assigned_by_admin_id=None,
            created_at=timestamp,
            updated_at=timestamp,
            notified_at=None,
        )
        db.add(assignment)
        db.commit()

        user_id = user.id
        vehicle_id = vehicle.id
        schedule_id = schedule.id
        assignment_id = assignment.id

    with engine.begin() as conn:
        conn.execute(text("DROP TABLE IF EXISTS legacy_vehicle_links"))
        conn.execute(text("DROP TABLE IF EXISTS legacy_schedule_links"))
        conn.execute(text("DROP TABLE IF EXISTS legacy_assignment_links"))
        conn.execute(
            text(
                """
                CREATE TABLE legacy_vehicle_links (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    vehicle_id INTEGER NOT NULL REFERENCES vehicles(id),
                    note VARCHAR(120) NOT NULL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE legacy_schedule_links (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    vehicle_schedule_id INTEGER NOT NULL REFERENCES transport_vehicle_schedules(id),
                    note VARCHAR(120) NOT NULL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE legacy_assignment_links (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    assignment_id INTEGER NOT NULL REFERENCES transport_assignments(id),
                    note VARCHAR(120) NOT NULL
                )
                """
            )
        )
        conn.execute(
            text("INSERT INTO legacy_vehicle_links (vehicle_id, note) VALUES (:vehicle_id, 'legacy vehicle ref')"),
            {"vehicle_id": vehicle_id},
        )
        conn.execute(
            text(
                "INSERT INTO legacy_schedule_links (vehicle_schedule_id, note) VALUES (:schedule_id, 'legacy schedule ref')"
            ),
            {"schedule_id": schedule_id},
        )
        conn.execute(
            text(
                "INSERT INTO legacy_assignment_links (assignment_id, note) VALUES (:assignment_id, 'legacy assignment ref')"
            ),
            {"assignment_id": assignment_id},
        )

    try:
        with TestClient(app) as client:
            ensure_admin_session(client)
            removed = client.delete(
                f"/api/transport/vehicles/{schedule_id}",
                params={"service_date": friday.isoformat()},
            )
            assert removed.status_code == 200, removed.text

        with engine.begin() as conn:
            remaining_vehicle_links = conn.execute(text("SELECT COUNT(*) FROM legacy_vehicle_links")).scalar_one()
            remaining_schedule_links = conn.execute(text("SELECT COUNT(*) FROM legacy_schedule_links")).scalar_one()
            remaining_assignment_links = conn.execute(text("SELECT COUNT(*) FROM legacy_assignment_links")).scalar_one()

        with SessionLocal() as db:
            user_row = db.get(User, user_id)
            vehicle_row = db.get(Vehicle, vehicle_id)
            assignment_row = db.get(TransportAssignment, assignment_id)

        assert remaining_vehicle_links == 0
        assert remaining_schedule_links == 0
        assert remaining_assignment_links == 0
        assert user_row is not None
        assert user_row.placa is None
        assert vehicle_row is None
        assert assignment_row is None
    finally:
        with engine.begin() as conn:
            conn.execute(text("DROP TABLE IF EXISTS legacy_vehicle_links"))
            conn.execute(text("DROP TABLE IF EXISTS legacy_schedule_links"))
            conn.execute(text("DROP TABLE IF EXISTS legacy_assignment_links"))


def test_transport_regular_assignment_persists_across_weekdays_and_routes():
    friday = date(2026, 4, 17)
    monday = date(2026, 4, 20)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Mirror Hub", address="2 Mirror Road", zip="222222", country="Singapore"))
        vehicle = Vehicle(placa="REG8001", tipo="van", color="Blue", lugares=11, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Mirror Rider",
            chave="MR01",
            projeto="P80",
            workplace="Mirror Hub",
            end_rua="22 Mirror Street",
            zip="220022",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:00",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        monday_home_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "home_to_work"},
        )
        monday_work_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "work_to_home"},
        )

    assert response.status_code == 200
    assert monday_home_dashboard.status_code == 200
    assert monday_work_dashboard.status_code == 200

    with SessionLocal() as db:
        assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == request_id)
            .order_by(TransportAssignment.route_kind)
        ).scalars().all()

    assert [row.route_kind for row in assignments] == ["home_to_work", "work_to_home"]
    assert all(row.status == "confirmed" for row in assignments)
    assert all(row.vehicle_id == vehicle_id for row in assignments)
    monday_home_request = next(
        row for row in monday_home_dashboard.json()["regular_requests"] if row["id"] == request_id
    )
    monday_work_request = next(
        row for row in monday_work_dashboard.json()["regular_requests"] if row["id"] == request_id
    )
    assert monday_home_request["assignment_status"] == "confirmed"
    assert monday_work_request["assignment_status"] == "confirmed"
    assert monday_home_request["assigned_vehicle"]["placa"] == "REG8001"
    assert monday_work_request["assigned_vehicle"]["placa"] == "REG8001"


def test_transport_regular_vehicle_registry_ignores_non_selected_weekdays_for_assigned_count():
    monday = date(2026, 4, 20)
    tuesday = date(2026, 4, 21)
    wednesday = date(2026, 4, 22)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Selected Weekday Hub", address="3 Weekday Road", zip="333333", country="Singapore"))
        vehicle = Vehicle(placa="REG8101", tipo="van", color="Gray", lugares=10, tolerance=7, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Weekday Filter Rider",
            chave="WF01",
            projeto="P80",
            workplace="Selected Weekday Hub",
            end_rua="33 Filter Street",
            zip="330033",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:10",
            selected_weekdays_json=json.dumps([0, 2, 4], ensure_ascii=True, separators=(",", ":")),
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": monday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        tuesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": tuesday.isoformat(), "route_kind": "home_to_work"},
        )
        wednesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": wednesday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert tuesday_dashboard.status_code == 200
    assert wednesday_dashboard.status_code == 200

    tuesday_payload = tuesday_dashboard.json()
    wednesday_payload = wednesday_dashboard.json()
    tuesday_request = next(row for row in tuesday_payload["regular_requests"] if row["id"] == request_id)
    wednesday_request = next(row for row in wednesday_payload["regular_requests"] if row["id"] == request_id)
    tuesday_vehicle_registry_row = next(
        row for row in tuesday_payload["regular_vehicle_registry"] if row["placa"] == "REG8101"
    )
    wednesday_vehicle_registry_row = next(
        row for row in wednesday_payload["regular_vehicle_registry"] if row["placa"] == "REG8101"
    )

    assert tuesday_request["service_date"] == wednesday.isoformat()
    assert tuesday_request["assignment_status"] == "confirmed"
    assert tuesday_request["assigned_vehicle"]["placa"] == "REG8101"
    assert tuesday_vehicle_registry_row["assigned_count"] == 0
    assert wednesday_vehicle_registry_row["assigned_count"] == 1
    assert wednesday_request["service_date"] == wednesday.isoformat()
    assert wednesday_request["assignment_status"] == "confirmed"
    assert wednesday_request["assigned_vehicle"]["placa"] == "REG8101"


def test_transport_weekend_vehicle_registry_ignores_non_selected_days_for_assigned_count():
    saturday = date(2026, 4, 18)
    sunday = date(2026, 4, 19)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Weekend Selected Hub", address="7 Weekend Lane", zip="777777", country="Singapore"))
        vehicle = Vehicle(placa="WKD8101", tipo="van", color="Silver", lugares=10, tolerance=8, service_scope="weekend")
        db.add(vehicle)
        db.flush()
        for weekday in (saturday.weekday(), sunday.weekday()):
            add_transport_schedule(
                db,
                vehicle=vehicle,
                service_scope="weekend",
                route_kind="home_to_work",
                recurrence_kind="matching_weekday",
                weekday=weekday,
            )
            add_transport_schedule(
                db,
                vehicle=vehicle,
                service_scope="weekend",
                route_kind="work_to_home",
                recurrence_kind="matching_weekday",
                weekday=weekday,
            )

        user = User(
            rfid=None,
            nome="Weekend Selected Rider",
            chave="WS01",
            projeto="P80",
            workplace="Weekend Selected Hub",
            end_rua="71 Weekend Street",
            zip="770071",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="weekend",
            recurrence_kind="weekend",
            requested_time="09:10",
            selected_weekdays_json=json.dumps([6], ensure_ascii=True, separators=(",", ":")),
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": sunday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        saturday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "home_to_work"},
        )
        sunday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": sunday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert saturday_dashboard.status_code == 200
    assert sunday_dashboard.status_code == 200

    saturday_payload = saturday_dashboard.json()
    sunday_payload = sunday_dashboard.json()
    saturday_request = next(row for row in saturday_payload["weekend_requests"] if row["id"] == request_id)
    sunday_request = next(row for row in sunday_payload["weekend_requests"] if row["id"] == request_id)
    saturday_vehicle_registry_row = next(
        row for row in saturday_payload["weekend_vehicle_registry"] if row["placa"] == "WKD8101"
    )
    sunday_vehicle_registry_row = next(
        row for row in sunday_payload["weekend_vehicle_registry"] if row["placa"] == "WKD8101"
    )

    assert saturday_request["service_date"] == sunday.isoformat()
    assert saturday_request["assignment_status"] == "confirmed"
    assert saturday_request["assigned_vehicle"]["placa"] == "WKD8101"
    assert saturday_vehicle_registry_row["assigned_count"] == 0
    assert sunday_vehicle_registry_row["assigned_count"] == 1
    assert sunday_request["service_date"] == sunday.isoformat()
    assert sunday_request["assignment_status"] == "confirmed"
    assert sunday_request["assigned_vehicle"]["placa"] == "WKD8101"


def test_transport_weekend_assignment_respects_selected_persistent_weekdays():
    saturday = date(2026, 4, 18)
    sunday = date(2026, 4, 19)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Weekend Persist Hub", address="9 Weekend Road", zip="919191", country="Singapore"))
        vehicle = Vehicle(placa="WKD8801", tipo="van", color="Blue", lugares=12, tolerance=11, service_scope="weekend")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="weekend",
            route_kind="home_to_work",
            recurrence_kind="matching_weekday",
            weekday=saturday.weekday(),
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="weekend",
            route_kind="work_to_home",
            recurrence_kind="matching_weekday",
            weekday=saturday.weekday(),
        )

        user = User(
            rfid=None,
            nome="Weekend Persist Rider",
            chave="WP88",
            projeto="P80",
            workplace="Weekend Persist Hub",
            end_rua="88 Weekend Street",
            zip="880088",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="weekend",
            recurrence_kind="weekend",
            requested_time="09:00",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": saturday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        saturday_home_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "home_to_work"},
        )
        saturday_work_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "work_to_home"},
        )
        sunday_home_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": sunday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert saturday_home_dashboard.status_code == 200
    assert saturday_work_dashboard.status_code == 200
    assert sunday_home_dashboard.status_code == 200

    saturday_home_request = next(
        row for row in saturday_home_dashboard.json()["weekend_requests"] if row["id"] == request_id
    )
    saturday_work_request = next(
        row for row in saturday_work_dashboard.json()["weekend_requests"] if row["id"] == request_id
    )
    sunday_home_request = next(
        row for row in sunday_home_dashboard.json()["weekend_requests"] if row["id"] == request_id
    )
    assert saturday_home_request["assignment_status"] == "confirmed"
    assert saturday_work_request["assignment_status"] == "confirmed"
    assert sunday_home_request["assignment_status"] == "pending"
    assert saturday_home_request["assigned_vehicle"]["placa"] == "WKD8801"
    assert saturday_work_request["assigned_vehicle"]["placa"] == "WKD8801"


def test_transport_dashboard_surfaces_extra_assignment_from_any_route():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Extra Route Hub", address="14 Route Way", zip="414141", country="Singapore"))
        vehicle = Vehicle(placa="EXT8101", tipo="carro", color="Black", lugares=4, tolerance=5, service_scope="extra")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="extra",
            route_kind="work_to_home",
            recurrence_kind="single_date",
            service_date=friday,
            departure_time="18:10",
        )

        user = User(
            rfid=None,
            nome="Extra Route Rider",
            chave="ER01",
            projeto="P80",
            workplace="Extra Route Hub",
            end_rua="14 Route Street",
            zip="410014",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="extra",
            recurrence_kind="single_date",
            requested_time="18:10",
            selected_weekdays_json=None,
            single_date=friday,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert dashboard.status_code == 200

    payload = dashboard.json()
    extra_request = next(row for row in payload["extra_requests"] if row["id"] == request_id)

    assert extra_request["assignment_status"] == "confirmed"
    assert extra_request["assigned_vehicle"]["placa"] == "EXT8101"
    assert extra_request["assigned_vehicle"]["route_kind"] == "work_to_home"


def test_admin_page_is_served_on_admin_path():
    with TestClient(app) as client:
        response = client.get("/admin")
        assert response.status_code == 200
        assert "Checking Admin" in response.text
        assert "Acesso Administrativo" in response.text
    assert "Projetos" in response.text
    assert 'id="projectsBody"' in response.text
    assert 'id="addProjectButton"' in response.text


def test_gerencia_page_is_not_served_anymore():
    with TestClient(app) as client:
        response = client.get("/gerencia")
    assert response.status_code == 404


def test_gerencia_trailing_slash_is_not_served_anymore():
    with TestClient(app) as client:
        response = client.get("/gerencia/", follow_redirects=False)
    assert response.status_code == 404


def test_database_events_endpoint_filters_and_paginates_check_events():
    primary_key = f"D{uuid.uuid4().hex[:3].upper()}"
    secondary_key = f"E{uuid.uuid4().hex[:3].upper()}"
    primary_rfid = f"rfid-db-{uuid.uuid4().hex[:8]}"
    secondary_rfid = f"rfid-db-{uuid.uuid4().hex[:8]}"
    timestamp = now_sgt().replace(microsecond=0)
    idempotency_keys = [uuid.uuid4().hex for _ in range(4)]

    try:
        with SessionLocal() as db:
            db.add_all(
                [
                    User(
                        rfid=primary_rfid,
                        nome="Usuario Banco Primario",
                        chave=primary_key,
                        projeto="P80",
                        workplace=None,
                        placa=None,
                        end_rua=None,
                        zip=None,
                        cargo=None,
                        email=None,
                        local=None,
                        checkin=None,
                        time=None,
                        last_active_at=timestamp,
                        inactivity_days=0,
                    ),
                    User(
                        rfid=secondary_rfid,
                        nome="Usuario Banco Secundario",
                        chave=secondary_key,
                        projeto="P82",
                        workplace=None,
                        placa=None,
                        end_rua=None,
                        zip=None,
                        cargo=None,
                        email=None,
                        local=None,
                        checkin=None,
                        time=None,
                        last_active_at=timestamp,
                        inactivity_days=0,
                    ),
                ]
            )
            db.add_all(
                [
                    CheckEvent(
                        idempotency_key=idempotency_keys[0],
                        source="device",
                        rfid=primary_rfid,
                        action="checkin",
                        status="success",
                        message="Entrada liberada",
                        details="reader=gate-1",
                        project="P80",
                        device_id="ESP-DB-01",
                        local="Portaria Norte",
                        request_path="/api/scan",
                        http_status=200,
                        ontime=True,
                        event_time=timestamp - timedelta(minutes=10),
                        submitted_at=timestamp - timedelta(minutes=10),
                        retry_count=0,
                    ),
                    CheckEvent(
                        idempotency_key=idempotency_keys[1],
                        source="mobile",
                        rfid=primary_rfid,
                        action="checkout",
                        status="queued",
                        message="Saida enviada pelo app",
                        details="client_event_id=db-checkout-1",
                        project="P80",
                        device_id="APP-DB-01",
                        local="Portaria Norte",
                        request_path="/api/mobile/events/sync",
                        http_status=202,
                        ontime=False,
                        event_time=timestamp - timedelta(minutes=2),
                        submitted_at=timestamp - timedelta(minutes=2),
                        retry_count=1,
                    ),
                    CheckEvent(
                        idempotency_key=idempotency_keys[2],
                        source="forms",
                        rfid=secondary_rfid,
                        action="checkin",
                        status="success",
                        message="Entrada confirmada",
                        details="provider=meta",
                        project="P82",
                        device_id="FORMS-DB-01",
                        local="Oficina Sul",
                        request_path="/api/provider/forms",
                        http_status=200,
                        ontime=True,
                        event_time=timestamp - timedelta(minutes=1),
                        submitted_at=timestamp - timedelta(minutes=1),
                        retry_count=0,
                    ),
                    CheckEvent(
                        idempotency_key=idempotency_keys[3],
                        source="admin",
                        rfid=None,
                        action="admin_request",
                        status="success",
                        message="Solicitacao administrativa",
                        details="chave=HR70",
                        project=None,
                        device_id=None,
                        local=None,
                        request_path="/api/admin/administrators/request",
                        http_status=200,
                        ontime=None,
                        event_time=timestamp,
                        submitted_at=timestamp,
                        retry_count=0,
                    ),
                ]
            )
            db.commit()

        with TestClient(app) as client:
            ensure_admin_session(client)

            first_page = client.get(
                "/api/admin/database-events",
                params={"chave": primary_key, "page_size": 1},
            )
            assert first_page.status_code == 200, first_page.text
            first_payload = first_page.json()
            assert first_payload["total"] == 2
            assert first_payload["page"] == 1
            assert first_payload["page_size"] == 1
            assert first_payload["total_pages"] == 2
            filter_options = first_payload["filter_options"]
            assert set(filter_options["action"]).issuperset({"checkin", "checkout"})
            assert set(filter_options["chave"]).issuperset({primary_key, secondary_key})
            assert set(filter_options["rfid"]).issuperset({primary_rfid, secondary_rfid})
            assert set(filter_options["project"]).issuperset({"P80", "P82"})
            assert set(filter_options["source"]).issuperset({"device", "forms", "mobile"})
            assert set(filter_options["status"]).issuperset({"queued", "success"})
            assert len(first_payload["items"]) == 1
            assert first_payload["items"][0]["chave"] == primary_key
            assert first_payload["items"][0]["action"] == "checkout"

            second_page = client.get(
                "/api/admin/database-events",
                params={"chave": primary_key, "page_size": 1, "page": 2},
            )
            assert second_page.status_code == 200, second_page.text
            second_payload = second_page.json()
            assert second_payload["page"] == 2
            assert len(second_payload["items"]) == 1
            assert second_payload["items"][0]["action"] == "checkin"

            key_sort_asc = client.get(
                "/api/admin/database-events",
                params={"sort_by": "chave", "sort_direction": "asc", "page_size": 200},
            )
            assert key_sort_asc.status_code == 200, key_sort_asc.text
            key_sort_asc_payload = key_sort_asc.json()
            key_sort_asc_chaves = [
                item["chave"]
                for item in key_sort_asc_payload["items"]
                if item["chave"] in {primary_key, secondary_key}
            ]
            assert key_sort_asc_chaves == [primary_key, primary_key, secondary_key]

            key_sort_desc = client.get(
                "/api/admin/database-events",
                params={"sort_by": "chave", "sort_direction": "desc", "page_size": 200},
            )
            assert key_sort_desc.status_code == 200, key_sort_desc.text
            key_sort_desc_payload = key_sort_desc.json()
            key_sort_desc_chaves = [
                item["chave"]
                for item in key_sort_desc_payload["items"]
                if item["chave"] in {primary_key, secondary_key}
            ]
            assert key_sort_desc_chaves == [secondary_key, primary_key, primary_key]

            message_sort_asc = client.get(
                "/api/admin/database-events",
                params={"chave": primary_key, "sort_by": "message", "sort_direction": "asc", "page_size": 10},
            )
            assert message_sort_asc.status_code == 200, message_sort_asc.text
            message_sort_asc_payload = message_sort_asc.json()
            assert [item["message"] for item in message_sort_asc_payload["items"]] == [
                "Entrada liberada",
                "Saida enviada pelo app",
            ]

            search_response = client.get(
                "/api/admin/database-events",
                params={"search": "oficina", "project": "P82"},
            )
            assert search_response.status_code == 200, search_response.text
            search_payload = search_response.json()
            assert search_payload["total"] == 1
            assert search_payload["items"][0]["chave"] == secondary_key
            assert search_payload["items"][0]["local"] == "Oficina Sul"
            assert search_payload["items"][0]["action"] == "checkin"
    finally:
        with SessionLocal() as db:
            for row in db.execute(select(CheckEvent).where(CheckEvent.idempotency_key.in_(idempotency_keys))).scalars().all():
                db.delete(row)
            for user in db.execute(select(User).where(User.chave.in_([primary_key, secondary_key]))).scalars().all():
                db.delete(user)
            db.commit()


def test_database_events_endpoint_rejects_invalid_date_ranges():
    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.get(
            "/api/admin/database-events",
            params={"from_date": "2025-01-10", "to_date": "2025-01-09"},
        )
        assert response.status_code == 400
        assert response.json()["detail"] == "Intervalo de datas invalido para a consulta de eventos."


def test_database_events_endpoint_rejects_invalid_sort_parameters():
    with TestClient(app) as client:
        ensure_admin_session(client)

        invalid_column = client.get(
            "/api/admin/database-events",
            params={"sort_by": "unexpected_column"},
        )
        assert invalid_column.status_code == 400
        assert invalid_column.json()["detail"] == "Coluna invalida para ordenacao de eventos."

        invalid_direction = client.get(
            "/api/admin/database-events",
            params={"sort_direction": "sideways"},
        )
        assert invalid_direction.status_code == 400
        assert invalid_direction.json()["detail"] == "Direcao invalida para ordenacao de eventos."


def test_web_password_registration_requires_existing_key_and_hashes_password():
    with TestClient(app) as client:
        response = register_web_password(client, chave="WB11", senha="s#123", projeto="P82")
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["authenticated"] is True
        assert payload["has_password"] is True

        status = client.get("/api/web/auth/status", params={"chave": "WB11"})
        assert status.status_code == 200
        assert status.json() == {
            "found": True,
            "chave": "WB11",
            "has_password": True,
            "authenticated": True,
            "message": "Aplicacao liberada.",
        }

        with SessionLocal() as db:
            user = get_user_by_chave(db, "WB11")
            assert user.nome == "Oriundo da Web"
            assert user.rfid is None
            assert user.projeto == "P82"
            assert user.senha is not None
            assert user.senha != "s#123"
            assert user.senha.startswith("pbkdf2_sha256$")
            assert verify_password("s#123", user.senha) is True


def test_web_password_registration_returns_not_found_for_unknown_key():
    with TestClient(app) as client:
        response = register_web_password(
            client,
            chave="ZZ91",
            senha="s#123",
            projeto="P82",
            ensure_user_exists=False,
        )

        assert response.status_code == 404
        assert response.json()["detail"] == "A chave do usuario nao esta cadastrada"


def test_web_user_self_registration_creates_user_requests_transport_access_and_authenticates_web_session():
    with TestClient(app) as client:
        response = client.post(
            "/api/web/auth/register-user",
            json={
                "chave": "WU11",
                "nome": "maria jose da silva",
                "projeto": "P83",
                "email": "maria.jose@petrobras.com.br",
                "senha": "cad123",
                "confirmar_senha": "cad123",
            },
        )

        assert response.status_code == 201
        payload = response.json()
        assert payload["ok"] is True
        assert payload["authenticated"] is True
        assert payload["has_password"] is True
        assert "aguarde aprovacao" in payload["message"].lower()

        status = client.get("/api/web/auth/status", params={"chave": "WU11"})
        assert status.status_code == 200
        assert status.json() == {
            "found": True,
            "chave": "WU11",
            "has_password": True,
            "authenticated": True,
            "message": "Aplicacao liberada.",
        }

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WU11")
        pending = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == "WU11")).scalar_one_or_none()
        assert user is not None
        assert user.nome == "Maria Jose da Silva"
        assert user.projeto == "P83"
        assert user.perfil == 0
        assert user.end_rua is None
        assert user.zip is None
        assert user.email == "maria.jose@petrobras.com.br"
        assert user.senha is not None
        assert user.senha != "cad123"
        assert verify_password("cad123", user.senha) is True
        assert pending is not None
        assert pending.requested_profile == 2

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        administrators = admin_client.get("/api/admin/administrators")
        assert administrators.status_code == 200
        pending_row = next(row for row in administrators.json() if row["row_type"] == "request" and row["chave"] == "WU11")
        assert pending_row["perfil"] == 2
        assert pending_row["status"] == "pending"
        assert "transport" in pending_row["status_label"].lower()

        approve_response = admin_client.post(f"/api/admin/administrators/requests/{pending_row['id']}/approve")
        assert approve_response.status_code == 200
        assert approve_response.json()["ok"] is True

    with SessionLocal() as db:
        approved_user = get_user_by_chave(db, "WU11")
        pending_after = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == "WU11")).scalar_one_or_none()
        assert approved_user is not None
        assert approved_user.perfil == 2
        assert pending_after is None

    with TestClient(app) as transport_client:
        denied_before_login = transport_client.get("/api/transport/auth/session")
        assert denied_before_login.status_code == 200
        assert denied_before_login.json()["authenticated"] is False

        transport_login = transport_client.post(
            "/api/transport/auth/verify",
            json={"chave": "WU11", "senha": "cad123"},
        )
        assert transport_login.status_code == 200
        assert transport_login.json()["authenticated"] is True
        assert transport_login.json()["user"]["perfil"] == 2


def test_web_user_self_registration_accepts_optional_email_blank():
    with TestClient(app) as client:
        response = client.post(
            "/api/web/auth/register-user",
            json={
                "chave": "WU12",
                "nome": "joao sem email",
                "projeto": "P82",
                "senha": "cad321",
                "confirmar_senha": "cad321",
            },
        )

        assert response.status_code == 201

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WU12")
        assert user is not None
        assert user.email is None


def test_web_check_endpoints_require_authenticated_password_session():
    payload = {
        "chave": "WB90",
        "projeto": "P80",
        "action": "checkin",
        "informe": "normal",
        "event_time": now_sgt().isoformat(),
        "client_event_id": f"web-check-auth-{uuid.uuid4().hex}",
    }

    with TestClient(app) as client:
        submit_response = client.post("/api/web/check", json=payload)
        history_response = client.get("/api/web/check/state", params={"chave": "WB90"})
        locations_response = client.get("/api/web/check/locations")
        location_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.255936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert submit_response.status_code == 401
        assert history_response.status_code == 401
        assert locations_response.status_code == 401
        assert location_response.status_code == 401


def test_web_password_change_replaces_previous_password():
    with TestClient(app) as client:
        registered = register_web_password(client, chave="WB15", senha="abc123", projeto="P80")
        assert registered.status_code == 200

        wrong_change = client.post(
            "/api/web/auth/change-password",
            json={
                "chave": "WB15",
                "senha_antiga": "000000",
                "nova_senha": "n0va#1",
            },
        )
        assert wrong_change.status_code == 401

        changed = client.post(
            "/api/web/auth/change-password",
            json={
                "chave": "WB15",
                "senha_antiga": "abc123",
                "nova_senha": "n0va#1",
            },
        )
        assert changed.status_code == 200
        assert changed.json()["authenticated"] is True
        assert changed.json()["has_password"] is True

        old_login = login_web_password(client, chave="WB15", senha="abc123")
        assert old_login.status_code == 401

        new_login = login_web_password(client, chave="WB15", senha="n0va#1")
        assert new_login.status_code == 200
        assert new_login.json()["authenticated"] is True
        assert new_login.json()["has_password"] is True


def test_web_transport_vehicle_request_returns_pending_state_without_same_day_checkin_requirement(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 7, 30, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT11", projeto="P80", nome="Transport Web Rider")
    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT11")
        user.end_rua = "10 Marina Boulevard"
        user.zip = "123456"
        db.commit()

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT11",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT11", "request_kind": "regular"},
        )
        assert created.status_code == 200
        payload = created.json()
        assert payload["ok"] is True
        assert payload["state"]["status"] == "pending"
        assert payload["state"]["request_kind"] == "regular"
        assert payload["state"]["requested_time"] == "07:30"
        assert payload["state"]["confirmation_deadline_time"] == "07:30"
        assert payload["state"]["end_rua"] == "10 Marina Boulevard"
        assert payload["state"]["zip"] == "123456"

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT11")
        request_row = db.execute(
            select(TransportRequest)
            .where(
                TransportRequest.user_id == user.id,
                TransportRequest.request_kind == "regular",
                TransportRequest.status == "active",
            )
            .order_by(TransportRequest.id.desc())
            .limit(1)
        ).scalar_one()

    assert request_row.created_via == "web"
    assert request_row.requested_time == "07:30"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    assert any(row["chave"] == "WT11" for row in dashboard.json()["regular_requests"])


def test_web_transport_stream_requires_authenticated_matching_session():
    ensure_web_user_exists(chave="WT12", projeto="P80", nome="Transport Stream Guard")

    with TestClient(app) as client:
        response = client.get("/api/web/transport/stream", params={"chave": "WT12"})

    assert response.status_code == 401
    assert response.json()["detail"] == "Sessao do usuario invalida ou expirada"


def test_web_transport_stream_emits_connected_and_transport_events():
    ensure_web_user_exists(chave="WT13", projeto="P80", nome="Transport Stream Rider")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT13")
        user.senha = hash_password("abc123")
        db.commit()

    class DummyWebTransportStreamRequest:
        def __init__(self, session):
            self.session = session

        async def is_disconnected(self):
            return False

    request = DummyWebTransportStreamRequest({web_check_router.WEB_USER_SESSION_KEY: "WT13"})

    with SessionLocal() as db:
        response = asyncio.run(
            web_check_router.stream_web_transport_updates(
                request,
                chave="WT13",
                db=db,
            )
        )
        assert response.media_type == "text/event-stream"
        assert response.headers["cache-control"] == "no-cache"

        first_chunk = asyncio.run(asyncio.wait_for(anext(response.body_iterator), timeout=1))
        first_payload = json.loads(first_chunk.removeprefix("data: ").strip())
        assert first_payload["reason"] == "connected"

        notify_transport_data_changed("event")

        second_chunk = asyncio.run(asyncio.wait_for(anext(response.body_iterator), timeout=1))
        second_payload = json.loads(second_chunk.removeprefix("data: ").strip())
        assert second_payload["reason"] == "event"

        asyncio.run(response.body_iterator.aclose())


def test_web_transport_vehicle_request_supports_weekend_and_extra_lists(monkeypatch):
    fixed_now = datetime(2026, 4, 18, 9, 15, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT14", projeto="P83", nome="Weekend Transport Rider")
    set_user_checkin_state(chave="WT14", event_time=fixed_now, local="Weekend Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT14",
            senha="abc123",
            projeto="P83",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        weekend_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT14", "request_kind": "weekend"},
        )
        assert weekend_created.status_code == 200
        assert weekend_created.json()["state"]["status"] == "pending"
        assert weekend_created.json()["state"]["request_kind"] == "weekend"

        extra_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT14", "request_kind": "extra"},
        )
        assert extra_created.status_code == 200
        assert extra_created.json()["state"]["status"] == "pending"
        assert extra_created.json()["state"]["request_kind"] == "extra"

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT14")
        request_kinds = db.execute(
            select(TransportRequest.request_kind)
            .where(
                TransportRequest.user_id == user.id,
                TransportRequest.status == "active",
            )
            .order_by(TransportRequest.id)
        ).scalars().all()

    assert request_kinds == ["weekend", "extra"]

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    assert any(row["chave"] == "WT14" for row in dashboard.json()["weekend_requests"])
    assert any(row["chave"] == "WT14" for row in dashboard.json()["extra_requests"])


def test_web_transport_weekend_and_extra_requests_remain_visible_before_their_target_date(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 9, 15, tzinfo=ZoneInfo(settings.tz_name))
    saturday = fixed_now.date() + timedelta(days=1)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT24", projeto="P83", nome="Upcoming Weekend Rider")
    set_user_checkin_state(chave="WT24", event_time=fixed_now, local="Friday Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT24",
            senha="abc123",
            projeto="P83",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        weekend_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT24", "request_kind": "weekend", "selected_weekdays": [5]},
        )
        assert weekend_created.status_code == 200
        assert weekend_created.json()["state"]["status"] == "pending"

        extra_created = client.post(
            "/api/web/transport/vehicle-request",
            json={
                "chave": "WT24",
                "request_kind": "extra",
                "requested_date": saturday.isoformat(),
                "requested_time": "18:10",
            },
        )
        assert extra_created.status_code == 200
        assert extra_created.json()["state"]["status"] == "pending"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    weekend_row = next(row for row in dashboard.json()["weekend_requests"] if row["chave"] == "WT24")
    extra_row = next(row for row in dashboard.json()["extra_requests"] if row["chave"] == "WT24")

    assert weekend_row["service_date"] == saturday.isoformat()
    assert weekend_row["assignment_status"] == "pending"
    assert extra_row["service_date"] == saturday.isoformat()
    assert extra_row["requested_time"] == "18:10"
    assert extra_row["assignment_status"] == "pending"


def test_web_transport_regular_request_remains_visible_before_the_next_selected_weekday(monkeypatch):
    fixed_now = datetime(2026, 4, 20, 9, 15, tzinfo=ZoneInfo(settings.tz_name))
    tuesday = fixed_now.date() + timedelta(days=1)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT33", projeto="P80", nome="Future Regular Rider")
    set_user_checkin_state(chave="WT33", event_time=fixed_now, local="Monday Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT33",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT33", "request_kind": "regular", "selected_weekdays": [1]},
        )
        assert created.status_code == 200
        assert created.json()["state"]["status"] == "pending"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    regular_rows = [row for row in dashboard.json()["regular_requests"] if row["chave"] == "WT33"]
    assert len(regular_rows) == 1
    assert regular_rows[0]["service_date"] == tuesday.isoformat()


def test_web_transport_state_accumulates_requests_and_cancels_replaced_regular_assignments(monkeypatch):
    clock = {"now": datetime(2026, 4, 20, 7, 30, tzinfo=ZoneInfo(settings.tz_name))}
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: clock["now"])
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: clock["now"])

    with SessionLocal() as db:
        vehicle = Vehicle(placa="HST2501", tipo="van", color="White", lugares=10, tolerance=5, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WT25", projeto="P80", nome="History Transport Rider")
    set_user_checkin_state(chave="WT25", event_time=clock["now"], local="History Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT25",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        first_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT25", "request_kind": "regular", "selected_weekdays": [0, 1]},
        )
        assert first_created.status_code == 200
        first_request_id = first_created.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": first_request_id,
                    "service_date": clock["now"].date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        clock["now"] = clock["now"] + timedelta(minutes=5)

        second_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT25", "request_kind": "regular", "selected_weekdays": [2, 3]},
        )
        assert second_created.status_code == 200

        state_response = client.get("/api/web/transport/state", params={"chave": "WT25"})
        assert state_response.status_code == 200
        state_payload = state_response.json()

    assert [item["request_id"] for item in state_payload["requests"][:2]] == [
        second_created.json()["state"]["requests"][0]["request_id"],
        first_request_id,
    ]
    assert state_payload["requests"][0]["status"] == "pending"
    assert state_payload["requests"][0]["selected_weekdays"] == [2, 3]
    assert state_payload["requests"][1]["status"] == "cancelled"
    assert state_payload["requests"][1]["is_active"] is False

    with SessionLocal() as db:
        first_request = db.get(TransportRequest, first_request_id)
        assert first_request is not None
        assert first_request.status == "cancelled"

        first_assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == first_request_id)
            .order_by(TransportAssignment.route_kind, TransportAssignment.service_date)
        ).scalars().all()

    assert first_assignments
    assert all(row.status == "cancelled" for row in first_assignments)
    assert all(row.vehicle_id is None for row in first_assignments)


def test_transport_dashboard_reject_marks_web_request_as_cancelled(monkeypatch):
    fixed_now = datetime(2026, 4, 21, 8, 10, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT26", projeto="P82", nome="Rejected Transport Rider")
    set_user_checkin_state(chave="WT26", event_time=fixed_now, local="Reject Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT26",
            senha="abc123",
            projeto="P82",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT26", "request_kind": "regular"},
        )
        assert created.status_code == 200
        request_id = created.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            rejected = admin_client.post(
                "/api/transport/requests/reject",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                },
            )
            assert rejected.status_code == 200

            dashboard = admin_client.get(
                "/api/transport/dashboard",
                params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
            )
            assert dashboard.status_code == 200

        state_response = client.get("/api/web/transport/state", params={"chave": "WT26"})
        assert state_response.status_code == 200
        state_payload = state_response.json()

    assert state_payload["status"] == "available"
    assert state_payload["requests"][0]["request_id"] == request_id
    assert state_payload["requests"][0]["status"] == "cancelled"
    assert state_payload["requests"][0]["is_active"] is False
    assert all(row["chave"] != "WT26" for row in dashboard.json()["regular_requests"])

    with SessionLocal() as db:
        request_row = db.get(TransportRequest, request_id)
        assert request_row is not None
        assert request_row.status == "cancelled"

        assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == request_id)
            .order_by(TransportAssignment.route_kind, TransportAssignment.service_date)
        ).scalars().all()

    assert assignments
    assert all(row.status == "rejected" for row in assignments)
    assert all(row.vehicle_id is None for row in assignments)


def test_transport_dashboard_reject_marks_weekend_and_extra_web_requests_as_cancelled(monkeypatch):
    clock = {"now": datetime(2026, 4, 18, 8, 20, tzinfo=ZoneInfo(settings.tz_name))}
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: clock["now"])
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: clock["now"])

    scenarios = [
        {
            "chave": "WW27",
            "nome": "Weekend Reject Rider",
            "projeto": "P82",
            "local": "Weekend Gate",
            "now": datetime(2026, 4, 18, 8, 20, tzinfo=ZoneInfo(settings.tz_name)),
            "create_payload": {"request_kind": "weekend", "selected_weekdays": [5]},
            "dashboard_key": "weekend_requests",
        },
        {
            "chave": "WX27",
            "nome": "Extra Reject Rider",
            "projeto": "P83",
            "local": "Extra Gate",
            "now": datetime(2026, 4, 21, 9, 5, tzinfo=ZoneInfo(settings.tz_name)),
            "create_payload": {
                "request_kind": "extra",
                "requested_date": "2026-04-21",
                "requested_time": "18:10",
            },
            "dashboard_key": "extra_requests",
        },
    ]

    for scenario in scenarios:
        clock["now"] = scenario["now"]
        ensure_web_user_exists(chave=scenario["chave"], projeto=scenario["projeto"], nome=scenario["nome"])
        set_user_checkin_state(chave=scenario["chave"], event_time=clock["now"], local=scenario["local"])

        with TestClient(app) as client:
            registered = register_web_password(
                client,
                chave=scenario["chave"],
                senha="abc123",
                projeto=scenario["projeto"],
                ensure_user_exists=False,
            )
            assert registered.status_code == 200

            created = client.post(
                "/api/web/transport/vehicle-request",
                json={"chave": scenario["chave"], **scenario["create_payload"]},
            )
            assert created.status_code == 200
            request_id = created.json()["state"]["request_id"]

            with TestClient(app) as admin_client:
                ensure_admin_session(admin_client)
                rejected = admin_client.post(
                    "/api/transport/requests/reject",
                    json={
                        "request_id": request_id,
                        "service_date": clock["now"].date().isoformat(),
                        "route_kind": "home_to_work",
                    },
                )
                assert rejected.status_code == 200

                dashboard = admin_client.get(
                    "/api/transport/dashboard",
                    params={"service_date": clock["now"].date().isoformat(), "route_kind": "home_to_work"},
                )
                assert dashboard.status_code == 200

            state_response = client.get("/api/web/transport/state", params={"chave": scenario["chave"]})
            assert state_response.status_code == 200
            state_payload = state_response.json()

        assert state_payload["status"] == "available"
        assert state_payload["requests"][0]["request_id"] == request_id
        assert state_payload["requests"][0]["status"] == "cancelled"
        assert state_payload["requests"][0]["is_active"] is False
        assert all(row["chave"] != scenario["chave"] for row in dashboard.json()[scenario["dashboard_key"]])

        with SessionLocal() as db:
            request_row = db.get(TransportRequest, request_id)

        assert request_row is not None
        assert request_row.status == "cancelled"


def test_transport_dashboard_pending_assignment_returns_request_to_pending_in_dashboard_and_webapp(monkeypatch):
    fixed_now = datetime(2026, 4, 22, 8, 10, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        vehicle = Vehicle(placa="PEN4222", tipo="carro", color="Blue", lugares=4, tolerance=6, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WT42", projeto="P82", nome="Pending Return Rider")
    set_user_checkin_state(chave="WT42", event_time=fixed_now, local="Pending Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT42",
            senha="abc123",
            projeto="P82",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT42", "request_kind": "regular"},
        )
        assert created.status_code == 200
        request_id = created.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

            returned_to_pending = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "pending",
                },
            )
            assert returned_to_pending.status_code == 200

            dashboard_home = admin_client.get(
                "/api/transport/dashboard",
                params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
            )
            dashboard_work = admin_client.get(
                "/api/transport/dashboard",
                params={"service_date": fixed_now.date().isoformat(), "route_kind": "work_to_home"},
            )

        state_response = client.get("/api/web/transport/state", params={"chave": "WT42"})
        assert state_response.status_code == 200
        state_payload = state_response.json()

    assert dashboard_home.status_code == 200
    assert dashboard_work.status_code == 200

    home_row = next(row for row in dashboard_home.json()["regular_requests"] if row["id"] == request_id)
    work_row = next(row for row in dashboard_work.json()["regular_requests"] if row["id"] == request_id)
    assert home_row["assignment_status"] == "pending"
    assert work_row["assignment_status"] == "pending"
    assert home_row["assigned_vehicle"] is None
    assert work_row["assigned_vehicle"] is None

    assert state_payload["status"] == "pending"
    assert state_payload["request_id"] == request_id
    assert state_payload["requests"][0]["request_id"] == request_id
    assert state_payload["requests"][0]["status"] == "pending"
    assert state_payload["requests"][0]["is_active"] is True

    with SessionLocal() as db:
        assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == request_id)
            .order_by(TransportAssignment.route_kind, TransportAssignment.service_date)
        ).scalars().all()

    assert assignments
    assert all(row.status == "pending" for row in assignments)
    assert all(row.vehicle_id is None for row in assignments)


def test_web_transport_regular_request_stays_visible_on_weekend_dashboard(monkeypatch):
    fixed_now = datetime(2026, 4, 19, 9, 15, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT18", projeto="P80", nome="Regular Weekend Rider")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT18")
        user.end_rua = "40 Weekend Avenue"
        user.zip = "654321"
        db.commit()

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT18",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT18", "request_kind": "regular"},
        )
        assert created.status_code == 200
        assert created.json()["state"]["status"] == "pending"
        assert created.json()["state"]["request_kind"] == "regular"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    regular_rows = [row for row in dashboard.json()["regular_requests"] if row["chave"] == "WT18"]
    assert len(regular_rows) == 1
    assert regular_rows[0]["nome"] == "Regular Weekend Rider"


def test_web_transport_address_update_and_acknowledgement_reflect_on_admin_dashboard(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 8, 0, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        db.commit()

    with SessionLocal() as db:
        vehicle = Vehicle(placa="TWA1234", tipo="van", color="Blue", lugares=12, tolerance=10, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WT12", projeto="P82", nome="Aware Transport Rider")
    set_user_checkin_state(chave="WT12", event_time=fixed_now, local="Main Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT12",
            senha="abc123",
            projeto="P82",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        updated_address = client.post(
            "/api/web/transport/address",
            json={
                "chave": "WT12",
                "end_rua": "Block 3, Harbour Street 55",
                "zip": "654321",
            },
        )
        assert updated_address.status_code == 200
        assert updated_address.json()["state"]["end_rua"] == "Block 3, Harbour Street 55"
        assert updated_address.json()["state"]["zip"] == "654321"

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT12", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        confirmed_state = client.get("/api/web/transport/state", params={"chave": "WT12"})
        assert confirmed_state.status_code == 200
        assert confirmed_state.json()["status"] == "confirmed"
        assert confirmed_state.json()["awareness_confirmed"] is False
        assert confirmed_state.json()["route_kind"] == "work_to_home"
        assert confirmed_state.json()["boarding_time"] == "16:45"
        assert confirmed_state.json()["vehicle_plate"] == "TWA1234"

        acknowledged = client.post(
            "/api/web/transport/acknowledge",
            json={"chave": "WT12", "request_id": request_id},
        )
        assert acknowledged.status_code == 200
        assert acknowledged.json()["state"]["awareness_confirmed"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT12")
        assert user.end_rua == "Block 3, Harbour Street 55"
        assert user.zip == "654321"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        home_dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )
        paired_dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "work_to_home"},
        )

    assert home_dashboard.status_code == 200
    assert paired_dashboard.status_code == 200
    home_row = next(row for row in home_dashboard.json()["regular_requests"] if row["chave"] == "WT12")
    paired_row = next(row for row in paired_dashboard.json()["regular_requests"] if row["chave"] == "WT12")
    assert home_row["awareness_status"] == "aware"
    assert paired_row["awareness_status"] == "aware"


def test_transport_export_endpoint_builds_xlsx_download_and_saves_server_copy(monkeypatch):
    fixed_now = datetime(2026, 4, 22, 15, 26, 45, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    export_dir = Path(settings.transport_exports_dir)
    if export_dir.exists():
        shutil.rmtree(export_dir)

    with SessionLocal() as db:
        vehicle = Vehicle(placa="EXP4521", tipo="van", color="White", lugares=9, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        export_home_user = User(
            nome="Export Home Rider",
            chave="EH01",
            projeto="P80",
            end_rua="10 Export Avenue",
            zip="111111",
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        export_work_user = User(
            nome="Export Work Rider",
            chave="EW02",
            projeto="P81",
            end_rua="20 Return Road",
            zip="222222",
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        export_pending_user = User(
            nome="Pending Export Rider",
            chave="EP03",
            projeto="P82",
            end_rua="30 Waiting Street",
            zip="333333",
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        db.add_all([export_home_user, export_work_user, export_pending_user])
        db.flush()

        home_request, _ = transport_service_module.upsert_transport_request(
            db,
            user=export_home_user,
            request_kind="regular",
            requested_time="07:15",
            requested_date=None,
            created_via="web",
        )
        work_request, _ = transport_service_module.upsert_transport_request(
            db,
            user=export_work_user,
            request_kind="regular",
            requested_time="18:10",
            requested_date=None,
            created_via="web",
        )
        pending_request, _ = transport_service_module.upsert_transport_request(
            db,
            user=export_pending_user,
            request_kind="regular",
            requested_time="08:05",
            requested_date=None,
            created_via="web",
        )
        db.commit()

        home_request_id = home_request.id
        work_request_id = work_request.id
        pending_request_id = pending_request.id
        vehicle_id = vehicle.id

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        home_assignment = admin_client.post(
            "/api/transport/assignments",
            json={
                "request_id": home_request_id,
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        assert home_assignment.status_code == 200

        work_assignment = admin_client.post(
            "/api/transport/assignments",
            json={
                "request_id": work_request_id,
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "work_to_home",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        assert work_assignment.status_code == 200

        pending_assignment = admin_client.post(
            "/api/transport/assignments",
            json={
                "request_id": pending_request_id,
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "home_to_work",
                "status": "pending",
                "vehicle_id": None,
            },
        )
        assert pending_assignment.status_code == 200

        exported = admin_client.get(
            "/api/transport/exports/transport-list",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        exported.headers["content-disposition"]
        == 'attachment; filename="Transport List - 20260422 - 152645.xlsx"'
    )

    workbook = load_workbook(io.BytesIO(exported.content))
    worksheet = workbook.active
    assert worksheet.title == "Transport List"
    assert [worksheet["A1"].value, worksheet["B1"].value, worksheet["C1"].value, worksheet["D1"].value, worksheet["E1"].value, worksheet["F1"].value] == [
        "Nome/Name",
        "Chave/Key",
        "Projeto/Project",
        "Endereço/Address",
        "Data/Date",
        "Partida/Departure",
    ]
    assert [worksheet["A2"].value, worksheet["B2"].value, worksheet["C2"].value, worksheet["D2"].value, worksheet["E2"].value, worksheet["F2"].value] == [
        "Export Home Rider",
        "EH01",
        "P80",
        "10 Export Avenue",
        "2026-04-22",
        None,
    ]
    assert [worksheet["A3"].value, worksheet["B3"].value, worksheet["C3"].value, worksheet["D3"].value, worksheet["E3"].value, worksheet["F3"].value] == [
        "Export Work Rider",
        "EW02",
        "P81",
        "20 Return Road",
        "2026-04-22",
        None,
    ]
    assert worksheet.max_row == 3
    workbook.close()

    saved_exports = sorted(Path(settings.transport_exports_dir).glob("Transport List - 20260422 - 152645*.xlsx"))
    assert len(saved_exports) == 1
    assert saved_exports[0].is_file()


def test_transport_settings_endpoint_updates_work_to_home_boarding_time(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 8, 10, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        vehicle = Vehicle(placa="TWS1810", tipo="carro", color="Silver", lugares=4, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WT19", projeto="P80", nome="Settings Boarding Rider")
    set_user_checkin_state(chave="WT19", event_time=fixed_now, local="Settings Gate")

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        current_settings = admin_client.get("/api/transport/settings")
        assert current_settings.status_code == 200
        assert current_settings.json()["work_to_home_time"] == "16:45"
        assert current_settings.json()["last_update_time"] == "16:00"
        assert current_settings.json()["default_car_seats"] == 3
        assert current_settings.json()["default_minivan_seats"] == 6
        assert current_settings.json()["default_van_seats"] == 10
        assert current_settings.json()["default_bus_seats"] == 40
        assert current_settings.json()["default_tolerance_minutes"] == 5

        updated_settings = admin_client.put(
            "/api/transport/settings",
            json={
                "work_to_home_time": "18:10",
                "last_update_time": "16:20",
                "default_car_seats": 4,
                "default_minivan_seats": 7,
                "default_van_seats": 11,
                "default_bus_seats": 44,
                "default_tolerance_minutes": 9,
            },
        )
        assert updated_settings.status_code == 200
        assert updated_settings.json()["work_to_home_time"] == "18:10"
        assert updated_settings.json()["last_update_time"] == "16:20"
        assert updated_settings.json()["default_car_seats"] == 4
        assert updated_settings.json()["default_minivan_seats"] == 7
        assert updated_settings.json()["default_van_seats"] == 11
        assert updated_settings.json()["default_bus_seats"] == 44
        assert updated_settings.json()["default_tolerance_minutes"] == 9

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT19",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT19", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        pending_state = client.get("/api/web/transport/state", params={"chave": "WT19"})
        assert pending_state.status_code == 200
        assert pending_state.json()["status"] == "pending"
        assert pending_state.json()["confirmation_deadline_time"] == "16:20"

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        confirmed_state = client.get("/api/web/transport/state", params={"chave": "WT19"})
        assert confirmed_state.status_code == 200
        assert confirmed_state.json()["status"] == "confirmed"
        assert confirmed_state.json()["route_kind"] == "work_to_home"
        assert confirmed_state.json()["boarding_time"] == "18:10"
        assert confirmed_state.json()["confirmation_deadline_time"] == "16:20"

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        location_settings_module.upsert_transport_last_update_time(db, last_update_time="16:00")
        db.commit()


def test_transport_date_settings_update_work_to_home_departure_for_selected_date_only(monkeypatch):
    friday_now = datetime(2026, 4, 17, 8, 35, tzinfo=ZoneInfo(settings.tz_name))
    saturday = friday_now.date() + timedelta(days=1)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: friday_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: friday_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")

        regular_vehicle = Vehicle(
            placa="DTD1710",
            tipo="carro",
            color="Graphite",
            lugares=4,
            tolerance=8,
            service_scope="regular",
        )
        weekend_vehicle = Vehicle(
            placa="DTD1720",
            tipo="van",
            color="Blue",
            lugares=8,
            tolerance=10,
            service_scope="weekend",
        )
        friday_extra_vehicle = Vehicle(
            placa="DTD1730",
            tipo="minivan",
            color="White",
            lugares=6,
            tolerance=12,
            service_scope="extra",
        )
        saturday_extra_vehicle = Vehicle(
            placa="DTD1740",
            tipo="carro",
            color="Black",
            lugares=4,
            tolerance=9,
            service_scope="extra",
        )
        db.add_all([regular_vehicle, weekend_vehicle, friday_extra_vehicle, saturday_extra_vehicle])
        db.flush()

        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="home_to_work",
            recurrence_kind="matching_weekday",
            weekday=5,
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="work_to_home",
            recurrence_kind="matching_weekday",
            weekday=5,
        )
        add_transport_schedule(
            db,
            vehicle=friday_extra_vehicle,
            service_scope="extra",
            route_kind="work_to_home",
            recurrence_kind="single_date",
            service_date=friday_now.date(),
        )
        add_transport_schedule(
            db,
            vehicle=saturday_extra_vehicle,
            service_scope="extra",
            route_kind="work_to_home",
            recurrence_kind="single_date",
            service_date=saturday,
        )
        db.commit()

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)

        updated_settings = admin_client.put(
            "/api/transport/date-settings",
            json={
                "service_date": friday_now.date().isoformat(),
                "work_to_home_time": "18:10",
            },
        )
        assert updated_settings.status_code == 200
        assert updated_settings.json() == {
            "service_date": friday_now.date().isoformat(),
            "work_to_home_time": "18:10",
        }

        friday_dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": friday_now.date().isoformat(), "route_kind": "work_to_home"},
        )
        saturday_dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "work_to_home"},
        )

    assert friday_dashboard.status_code == 200
    assert saturday_dashboard.status_code == 200

    friday_payload = friday_dashboard.json()
    saturday_payload = saturday_dashboard.json()
    friday_regular_row = next(row for row in friday_payload["regular_vehicles"] if row["placa"] == "DTD1710")
    friday_extra_row = next(row for row in friday_payload["extra_vehicles"] if row["placa"] == "DTD1730")
    saturday_weekend_row = next(row for row in saturday_payload["weekend_vehicles"] if row["placa"] == "DTD1720")
    saturday_extra_row = next(row for row in saturday_payload["extra_vehicles"] if row["placa"] == "DTD1740")

    assert friday_payload["work_to_home_departure_time"] == "18:10"
    assert friday_regular_row["departure_time"] == "18:10"
    assert friday_extra_row["departure_time"] is None

    assert saturday_payload["work_to_home_departure_time"] == "16:45"
    assert saturday_weekend_row["departure_time"] == "16:45"
    assert saturday_extra_row["departure_time"] is None


def test_web_transport_date_override_applies_only_on_selected_day(monkeypatch):
    current_now = {"value": datetime(2026, 4, 17, 8, 45, tzinfo=ZoneInfo(settings.tz_name))}
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: current_now["value"])
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: current_now["value"])

    saturday = current_now["value"].date() + timedelta(days=1)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")

        regular_vehicle = Vehicle(
            placa="DOW1810",
            tipo="carro",
            color="Silver",
            lugares=4,
            tolerance=8,
            service_scope="regular",
        )
        weekend_vehicle = Vehicle(
            placa="DOW1645",
            tipo="van",
            color="White",
            lugares=8,
            tolerance=10,
            service_scope="weekend",
        )
        db.add_all([regular_vehicle, weekend_vehicle])
        db.flush()

        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="home_to_work",
            recurrence_kind="matching_weekday",
            weekday=5,
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="work_to_home",
            recurrence_kind="matching_weekday",
            weekday=5,
        )
        db.commit()
        regular_vehicle_id = regular_vehicle.id
        weekend_vehicle_id = weekend_vehicle.id

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        updated_settings = admin_client.put(
            "/api/transport/date-settings",
            json={
                "service_date": current_now["value"].date().isoformat(),
                "work_to_home_time": "18:10",
            },
        )
        assert updated_settings.status_code == 200

    ensure_web_user_exists(chave="DW18", projeto="P80", nome="Date Override Friday Rider")
    set_user_checkin_state(chave="DW18", event_time=current_now["value"], local="Date Override Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="DW18",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "DW18", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        friday_request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": friday_request_id,
                    "service_date": current_now["value"].date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": regular_vehicle_id,
                },
            )
            assert assigned.status_code == 200

        friday_state = client.get("/api/web/transport/state", params={"chave": "DW18"})
        assert friday_state.status_code == 200
        assert friday_state.json()["status"] == "confirmed"
        assert friday_state.json()["route_kind"] == "work_to_home"
        assert friday_state.json()["boarding_time"] == "18:10"

    current_now["value"] = datetime(2026, 4, 18, 8, 50, tzinfo=ZoneInfo(settings.tz_name))
    ensure_web_user_exists(chave="DW16", projeto="P80", nome="Date Override Saturday Rider")
    set_user_checkin_state(chave="DW16", event_time=current_now["value"], local="Date Override Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="DW16",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "DW16", "request_kind": "weekend"},
        )
        assert requested.status_code == 200
        saturday_request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": saturday_request_id,
                    "service_date": saturday.isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": weekend_vehicle_id,
                },
            )
            assert assigned.status_code == 200

        saturday_state = client.get("/api/web/transport/state", params={"chave": "DW16"})
        assert saturday_state.status_code == 200
        assert saturday_state.json()["status"] == "confirmed"
        assert saturday_state.json()["route_kind"] == "work_to_home"
        assert saturday_state.json()["boarding_time"] == "16:45"

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        db.commit()


def test_web_transport_state_marks_departed_confirmed_request_as_realized_and_exposes_vehicle_color(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 19, 30, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        vehicle = Vehicle(placa="REL1730", tipo="carro", color="Silver", lugares=4, tolerance=7, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WR17", projeto="P80", nome="Realized History Rider")
    set_user_checkin_state(chave="WR17", event_time=fixed_now, local="Realized Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WR17",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WR17", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        state_response = client.get("/api/web/transport/state", params={"chave": "WR17"})
        assert state_response.status_code == 200

    state_payload = state_response.json()
    assert state_payload["status"] == "realized"
    assert state_payload["route_kind"] == "work_to_home"
    assert state_payload["boarding_time"] == "18:10"
    assert state_payload["vehicle_color"] == "Silver"
    assert state_payload["requests"]
    assert state_payload["requests"][0]["request_id"] == request_id
    assert state_payload["requests"][0]["status"] == "realized"
    assert state_payload["requests"][0]["service_date"] == fixed_now.date().isoformat()
    assert state_payload["requests"][0]["vehicle_color"] == "Silver"


def test_web_transport_request_history_uses_next_service_date_and_effective_departure_time(monkeypatch):
    current_now = {"value": datetime(2026, 4, 18, 9, 10, tzinfo=ZoneInfo(settings.tz_name))}
    monday = current_now["value"].date() + timedelta(days=2)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: current_now["value"])
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: current_now["value"])

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        location_settings_module.upsert_transport_work_to_home_time_for_date(
            db,
            service_date=monday,
            work_to_home_time="18:10",
        )
        db.commit()

    ensure_web_user_exists(chave="WH18", projeto="P80", nome="History Departure Rider")
    set_user_checkin_state(chave="WH18", event_time=current_now["value"], local="Weekend Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WH18",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WH18", "request_kind": "regular"},
        )
        assert created.status_code == 200

        state_response = client.get("/api/web/transport/state", params={"chave": "WH18"})
        assert state_response.status_code == 200

    state_payload = state_response.json()
    assert state_payload["requests"]
    assert state_payload["requests"][0]["status"] == "pending"
    assert state_payload["requests"][0]["service_date"] == monday.isoformat()
    assert state_payload["requests"][0]["boarding_time"] == "18:10"


def test_web_transport_cancel_pending_request_marks_history_item_cancelled(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 8, 35, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT31", projeto="P80", nome="Pending Cancel Rider")
    set_user_checkin_state(chave="WT31", event_time=fixed_now, local="Pending Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT31",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT31", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        cancelled = client.post(
            "/api/web/transport/cancel",
            json={"chave": "WT31", "request_id": request_id},
        )
        assert cancelled.status_code == 200

    cancelled_payload = cancelled.json()
    assert cancelled_payload["state"]["status"] == "available"
    assert cancelled_payload["state"]["requests"]
    assert cancelled_payload["state"]["requests"][0]["request_id"] == request_id
    assert cancelled_payload["state"]["requests"][0]["status"] == "cancelled"
    assert cancelled_payload["state"]["requests"][0]["is_active"] is False


def test_web_transport_cancel_future_regular_request_removes_row_from_dashboard(monkeypatch):
    fixed_now = datetime(2026, 4, 20, 8, 35, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT34", projeto="P80", nome="Future Cancel Rider")
    set_user_checkin_state(chave="WT34", event_time=fixed_now, local="Monday Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT34",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT34", "request_kind": "regular", "selected_weekdays": [1]},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        cancelled = client.post(
            "/api/web/transport/cancel",
            json={"chave": "WT34", "request_id": request_id},
        )
        assert cancelled.status_code == 200

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    assert all(row["chave"] != "WT34" for row in dashboard.json()["regular_requests"])


def test_transport_dashboard_reject_extra_request_does_not_store_transport_session_user_in_admin_fk(monkeypatch):
    fixed_now = datetime(2026, 4, 21, 7, 45, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        transport_operator = User(
            rfid=None,
            nome="Transport Operator",
            chave="TP41",
            senha=hash_password("tp1234"),
            perfil=2,
            projeto="P82",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            cargo=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        rider = User(
            rfid=None,
            nome="Extra Reject Rider",
            chave="XR41",
            senha=None,
            perfil=0,
            projeto="P82",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            cargo=None,
            email=None,
            local=None,
            checkin=True,
            time=fixed_now,
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        db.add_all([transport_operator, rider])
        db.flush()

        request_row = TransportRequest(
            user_id=rider.id,
            request_kind="extra",
            recurrence_kind="single_date",
            requested_time="07:45",
            selected_weekdays_json=None,
            single_date=fixed_now.date(),
            created_via="web",
            status="active",
            created_at=fixed_now,
            updated_at=fixed_now,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id

    with TestClient(app) as transport_client:
        login_response = transport_client.post(
            "/api/transport/auth/verify",
            json={"chave": "TP41", "senha": "tp1234"},
        )
        assert login_response.status_code == 200
        assert login_response.json()["authenticated"] is True

        rejected = transport_client.post(
            "/api/transport/requests/reject",
            json={
                "request_id": request_id,
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "home_to_work",
            },
        )

    assert rejected.status_code == 200

    with SessionLocal() as db:
        request_row = db.get(TransportRequest, request_id)
        assignment = db.execute(
            select(TransportAssignment).where(
                TransportAssignment.request_id == request_id,
                TransportAssignment.service_date == fixed_now.date(),
                TransportAssignment.route_kind == "home_to_work",
            )
        ).scalar_one()

    assert request_row is not None
    assert request_row.status == "cancelled"
    assert assignment.status == "rejected"
    assert assignment.assigned_by_admin_id is None


def test_web_transport_cancel_after_confirmation_removes_request_from_vehicle_dashboard(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 8, 20, tzinfo=ZoneInfo(settings.tz_name))
    user_key = f"W{uuid.uuid4().hex[:3].upper()}"
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        vehicle = Vehicle(placa="TWC1234", tipo="carro", color="White", lugares=4, tolerance=7, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave=user_key, projeto="P80", nome="Cancel Transport Rider")
    set_user_checkin_state(chave=user_key, event_time=fixed_now, local="North Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave=user_key,
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": user_key, "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        cancelled = client.post(
            "/api/web/transport/cancel",
            json={"chave": user_key, "request_id": request_id},
        )
        assert cancelled.status_code == 200
        assert cancelled.json()["state"]["status"] == "available"

    with SessionLocal() as db:
        request_row = db.get(TransportRequest, request_id)
        assert request_row is not None
        assert request_row.status == "cancelled"
        assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == request_id)
            .order_by(TransportAssignment.route_kind)
        ).scalars().all()

    assert assignments
    assert all(row.status == "cancelled" for row in assignments)
    assert all(row.vehicle_id is None for row in assignments)

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    assert all(row["chave"] != user_key for row in dashboard.json()["regular_requests"])


def test_admin_can_clear_registered_user_password_and_allow_new_registration():
    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "USRSPASS1", "nome": "Usuario Senha", "chave": "PW11", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        users = client.get("/api/admin/users")
        assert users.status_code == 200
        user_id = next(row["id"] for row in users.json() if row["chave"] == "PW11")

        registered = register_web_password(client, chave="PW11", senha="abc123", projeto="P80", ensure_user_exists=False)
        assert registered.status_code == 200

        reset_response = client.post(f"/api/admin/users/{user_id}/reset-password")
        assert reset_response.status_code == 200
        assert "nova senha" in reset_response.json()["message"]

        with SessionLocal() as db:
            user = get_user_by_chave(db, "PW11")
            assert user.senha is None

        status = client.get("/api/web/auth/status", params={"chave": "PW11"})
        assert status.status_code == 200
        assert status.json() == {
            "found": True,
            "chave": "PW11",
            "has_password": False,
            "authenticated": False,
            "message": "Digite sua chave e crie uma senha.",
        }

        old_login = login_web_password(client, chave="PW11", senha="abc123")
        assert old_login.status_code == 404

        new_registration = register_web_password(client, chave="PW11", senha="nova123", projeto="P80", ensure_user_exists=False)
        assert new_registration.status_code == 200
        assert new_registration.json()["authenticated"] is True
        assert new_registration.json()["has_password"] is True

        with SessionLocal() as db:
            user = get_user_by_chave(db, "PW11")
            assert user.senha is not None
            assert verify_password("nova123", user.senha) is True

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(
            event["action"] == "password"
            and event["request_path"] == f"/api/admin/users/{user_id}/reset-password"
            and event["status"] == "removed"
            for event in events.json()
        )


def test_web_password_status_returns_not_found_for_unknown_key():
    with TestClient(app) as client:
        response = client.get("/api/web/auth/status", params={"chave": "ZZ99"})

        assert response.status_code == 200
        assert response.json() == {
            "found": False,
            "chave": "ZZ99",
            "has_password": False,
            "authenticated": False,
            "message": "Digite sua chave e crie uma senha.",
        }


def test_web_password_login_accepts_partial_attempts_without_validation_error():
    with TestClient(app) as client:
        registered = register_web_password(client, chave="WB19", senha="abc123", projeto="P80")
        assert registered.status_code == 200

        logout = client.post("/api/web/auth/logout")
        assert logout.status_code == 200

        partial_attempt = client.post(
            "/api/web/auth/login",
            json={
                "chave": "WB19",
                "senha": "a",
            },
        )

        assert partial_attempt.status_code == 401
        assert partial_attempt.json()["detail"] == "Chave ou senha invalida"


def test_web_location_match_returns_known_location_when_accuracy_is_good():
    with TestClient(app) as client:
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL80", senha="loc123", projeto="P80")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Match P80",
                "latitude": 1.255936,
                "longitude": 103.611066,
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.255936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is True
        assert payload["resolved_local"] == "Web Match P80"
        assert payload["label"] == "Web Match P80"
        assert payload["status"] == "matched"
        assert payload["accuracy_threshold_meters"] == 25


def test_web_location_match_blocks_low_accuracy_before_matching():
    with TestClient(app) as client:
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL81", senha="loc123", projeto="P80")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Accuracy P80",
                "latitude": 1.300001,
                "longitude": 103.800001,
                "tolerance_meters": 120,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 15},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.300001,
                "longitude": 103.800001,
                "accuracy_meters": 44,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is False
        assert payload["resolved_local"] is None
        assert payload["label"] == "Precisao insuficiente"
        assert payload["status"] == "accuracy_too_low"
        assert payload["accuracy_threshold_meters"] == 15


def test_web_location_match_returns_unregistered_location_without_message_within_two_km():
    with TestClient(app) as client:
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL82", senha="loc123", projeto="P80")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Nearby P80",
                "latitude": 1.255936,
                "longitude": 103.611066,
                "tolerance_meters": 120,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.260936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is False
        assert payload["resolved_local"] is None
        assert payload["label"] == "Localização não Cadastrada"
        assert payload["status"] == "not_in_known_location"
        assert payload["message"] == ""
        assert payload["nearest_workplace_distance_meters"] < 2000


def test_web_location_match_returns_outside_workplace_without_message():
    with TestClient(app) as client:
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL83", senha="loc123", projeto="P80")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Far P80",
                "latitude": 1.255936,
                "longitude": 103.611066,
                "tolerance_meters": 120,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.285936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is False
        assert payload["resolved_local"] is None
        assert payload["label"] == "Fora do Ambiente de Trabalho"
        assert payload["status"] == "outside_workplace"
        assert payload["message"] == ""
        assert payload["nearest_workplace_distance_meters"] > 2000


def test_web_check_updates_user_local_when_location_is_provided():
    client_event_id = f"web-check-local-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        auth_response = register_web_password(client, chave="WB14", senha="local1", projeto="P80")
        assert auth_response.status_code == 200

        response = client.post(
            "/api/web/check",
            json={
                "chave": "WB14",
                "projeto": "P80",
                "action": "checkin",
                "local": "Web Match P80",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )
        history = client.get("/api/web/check/state", params={"chave": "WB14"})

        assert response.status_code == 200
        assert response.json()["ok"] is True
        assert history.status_code == 200
        assert history.json() == {
            "found": True,
            "chave": "WB14",
            "projeto": "P80",
            "current_action": "checkin",
            "current_local": "Web Match P80",
            "has_current_day_checkin": True,
            "last_checkin_at": response.json()["state"]["last_checkin_at"],
            "last_checkout_at": None,
        }

        with SessionLocal() as db:
            user = get_user_by_chave(db, "WB14")
            assert user.local == "Web Match P80"


def test_web_check_reuses_flutter_like_hidden_project_for_checkout():
    first_event_time = now_sgt()
    second_event_time = first_event_time + timedelta(minutes=4)

    with TestClient(app) as client:
        auth_response = register_web_password(client, chave="WB12", senha="check1", projeto="P83")
        assert auth_response.status_code == 200

        first = client.post(
            "/api/web/check",
            json={
                "chave": "WB12",
                "projeto": "P83",
                "action": "checkout",
                "informe": "retroativo",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"web-check-1-{uuid.uuid4().hex}",
            },
        )
        second = client.post(
            "/api/web/check",
            json={
                "chave": "WB12",
                "projeto": "P83",
                "action": "checkout",
                "informe": "retroativo",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"web-check-2-{uuid.uuid4().hex}",
            },
        )

        assert first.status_code == 200
        assert first.json()["queued_forms"] is True
        assert second.status_code == 200
        assert second.json()["queued_forms"] is False

        with SessionLocal() as db:
            user = get_user_by_chave(db, "WB12")
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.chave == "WB12")).scalars().all()
            sync_events = db.execute(
                select(UserSyncEvent).where(
                    UserSyncEvent.source == "web_forms",
                    UserSyncEvent.chave == "WB12",
                )
            ).scalars().all()
            request_events = db.execute(
                select(CheckEvent).where(CheckEvent.request_path == "/api/web/check", CheckEvent.rfid.is_(None))
            ).scalars().all()

            assert user.nome == "Oriundo da Web"
            assert user.projeto == "P83"
            assert user.checkin is False
            assert len(queued) == 1
            assert len(sync_events) == 2
            assert any(event.ontime is False and event.action == "checkout" for event in sync_events)
            assert any(event.action == "checkout" and event.status == "queued" for event in request_events)


def test_web_check_state_returns_latest_public_history():
    checkin_at = now_sgt() - timedelta(hours=1)
    checkout_at = now_sgt()

    with TestClient(app) as client:
        auth_response = register_web_password(client, chave="WB13", senha="state1", projeto="P80")
        assert auth_response.status_code == 200

        first = client.post(
            "/api/web/check",
            json={
                "chave": "WB13",
                "projeto": "P80",
                "action": "checkin",
                "informe": "normal",
                "event_time": checkin_at.isoformat(),
                "client_event_id": f"web-history-1-{uuid.uuid4().hex}",
            },
        )
        second = client.post(
            "/api/web/check",
            json={
                "chave": "WB13",
                "projeto": "P80",
                "action": "checkout",
                "informe": "normal",
                "event_time": checkout_at.isoformat(),
                "client_event_id": f"web-history-2-{uuid.uuid4().hex}",
            },
        )
        history = client.get("/api/web/check/state", params={"chave": "WB13"})

        assert first.status_code == 200
        assert second.status_code == 200
        assert history.status_code == 200

        payload = history.json()
        assert payload == {
            "found": True,
            "chave": "WB13",
            "projeto": "P80",
            "current_action": "checkout",
            "current_local": "Web",
            "has_current_day_checkin": True,
            "last_checkin_at": checkin_at.replace(tzinfo=None).isoformat(),
            "last_checkout_at": checkout_at.replace(tzinfo=None).isoformat(),
        }


def test_mobile_sync_accepts_project_p82():
    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP82",
                "projeto": "P82",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert response.status_code == 200
        assert response.json()["state"]["projeto"] == "P82"


def test_mobile_checkout_preserves_previous_checkin_history_without_existing_sync_events():
    previous_checkin_at = now_sgt() - timedelta(hours=2)
    checkout_at = now_sgt()

    with SessionLocal() as db:
        user = User(
            rfid=None,
            chave="LG11",
            nome="Legado Mobile",
            projeto="P80",
            local=None,
            checkin=True,
            time=previous_checkin_at,
            last_active_at=previous_checkin_at,
            inactivity_days=0,
        )
        db.add(user)
        db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "LG11",
                "projeto": "P80",
                "action": "checkout",
                "event_time": checkout_at.isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )

        assert response.status_code == 200
        state = response.json()["state"]
        assert state["current_action"] == "checkout"
        assert state["last_checkin_at"] is not None
        assert state["last_checkout_at"] is not None


def test_mobile_state_falls_back_to_check_events_history():
    checkin_at = now_sgt() - timedelta(hours=3)
    checkout_at = now_sgt() - timedelta(hours=1)

    with SessionLocal() as db:
        user = User(
            rfid="RFBACK1",
            chave="FB11",
            nome="Fallback Historico",
            projeto="P80",
            local="main",
            checkin=False,
            time=checkout_at,
            last_active_at=checkout_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()
        db.add(
            CheckEvent(
                idempotency_key=f"fallback-checkin-{uuid.uuid4().hex}",
                source="device",
                rfid="RFBACK1",
                action="checkin",
                status="queued",
                message="checkin historico",
                details=None,
                project="P80",
                device_id="ESP32-FALLBACK",
                local="main",
                request_path="/api/scan",
                http_status=202,
                event_time=checkin_at,
                submitted_at=None,
                retry_count=0,
            )
        )
        db.add(
            CheckEvent(
                idempotency_key=f"fallback-checkout-{uuid.uuid4().hex}",
                source="device",
                rfid="RFBACK1",
                action="checkout",
                status="queued",
                message="checkout historico",
                details=None,
                project="P80",
                device_id="ESP32-FALLBACK",
                local="main",
                request_path="/api/scan",
                http_status=202,
                event_time=checkout_at,
                submitted_at=None,
                retry_count=0,
            )
        )
        db.commit()

    with TestClient(app) as client:
        response = client.get("/api/mobile/state?chave=FB11", headers=MOBILE_HEADERS)

        assert response.status_code == 200
        state = response.json()
        assert state["found"] is True
        assert state["current_action"] == "checkout"
        assert state["last_checkin_at"] is not None
        assert state["last_checkout_at"] is not None


def test_archive_events_creates_csv_clears_table_and_lists_downloads(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ARC1000", "nome": "Usuario Arquivo", "chave": "JK90", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        archive_res = client.post("/api/admin/events/archive")
        assert archive_res.status_code == 200

        archive_payload = archive_res.json()
        assert archive_payload["created"] is True
        assert archive_payload["cleared_count"] >= 1
        assert archive_payload["archive"]["file_name"].endswith(".csv")
        assert " a " in archive_payload["archive"]["period"]
        assert archive_payload["archive"]["record_count"] >= 1
        assert archive_payload["archives"]["total"] >= 1
        assert archive_payload["archives"]["total_size_bytes"] >= archive_payload["archive"]["size_bytes"]
        assert archive_payload["archives"]["items"][0]["file_name"] == archive_payload["archive"]["file_name"]

        events_after = client.get("/api/admin/events")
        assert events_after.status_code == 200
        assert events_after.json() == []

        archives_list = client.get("/api/admin/events/archives")
        assert archives_list.status_code == 200
        assert archives_list.json()["items"][0]["file_name"] == archive_payload["archive"]["file_name"]
        assert archives_list.json()["page"] == 1
        assert archives_list.json()["page_size"] == 8

        download_res = client.get(
            f"/api/admin/events/archives/{archive_payload['archive']['file_name']}",
        )
        assert download_res.status_code == 200
        assert "attachment;" in download_res.headers["content-disposition"]
        assert "event_time" in download_res.text
        assert "ARC1000" in download_res.text

        download_all_res = client.get("/api/admin/events/archives/download-all")
        assert download_all_res.status_code == 200
        assert download_all_res.headers["content-type"].startswith("application/zip")
        assert len(download_all_res.content) > 0


def test_delete_archived_event_csv(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ARCDEL", "nome": "Usuario Delete", "chave": "DL90", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        archive_res = client.post("/api/admin/events/archive")
        assert archive_res.status_code == 200
        file_name = archive_res.json()["archive"]["file_name"]

        delete_res = client.delete(f"/api/admin/events/archives/{file_name}")
        assert delete_res.status_code == 200
        assert delete_res.json()["ok"] is True

        list_res = client.get("/api/admin/events/archives")
        assert list_res.status_code == 200
        assert all(item["file_name"] != file_name for item in list_res.json()["items"])

        missing_res = client.get(f"/api/admin/events/archives/{file_name}")
        assert missing_res.status_code == 404


def test_event_archive_operations_are_logged(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ARCAUD1", "nome": "Usuario Auditoria", "chave": "AU90", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        archive_res = client.post("/api/admin/events/archive")
        assert archive_res.status_code == 200
        file_name = archive_res.json()["archive"]["file_name"]

        single_download = client.get(f"/api/admin/events/archives/{file_name}")
        assert single_download.status_code == 200

        download_all = client.get("/api/admin/events/archives/download-all")
        assert download_all.status_code == 200

        delete_res = client.delete(f"/api/admin/events/archives/{file_name}")
        assert delete_res.status_code == 200

        missing_res = client.get(f"/api/admin/events/archives/{file_name}")
        assert missing_res.status_code == 404

        events_res = client.get("/api/admin/events")
        assert events_res.status_code == 200
        assert events_res.json() == []

        with SessionLocal() as db:
            archive_events = db.execute(
                select(CheckEvent)
                .where(CheckEvent.action == "event_archive")
                .order_by(CheckEvent.id)
            ).scalars().all()

        assert any(
            event.status == "created"
            and event.request_path == "/api/admin/events/archive"
            for event in archive_events
        )
        assert any(
            event.status == "downloaded"
            and event.request_path == f"/api/admin/events/archives/{file_name}"
            for event in archive_events
        )
        assert any(
            event.status == "downloaded"
            and event.request_path == "/api/admin/events/archives/download-all"
            for event in archive_events
        )
        assert any(
            event.status == "removed"
            and event.request_path == f"/api/admin/events/archives/{file_name}"
            for event in archive_events
        )
        assert any(
            event.status == "failed"
            and event.http_status == 404
            and event.request_path == f"/api/admin/events/archives/{file_name}"
            for event in archive_events
        )


def test_archive_events_without_current_rows_returns_existing_archives(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")

    with TestClient(app) as client:
        ensure_admin_session(client)
        first = client.post(
            "/api/admin/users",
            json={"rfid": "ARCEMPTY", "nome": "Usuario Empty", "chave": "EM90", "projeto": "P83"},
        )
        assert first.status_code == 200

        archive_first = client.post("/api/admin/events/archive")
        assert archive_first.status_code == 200
        assert archive_first.json()["created"] is True

        archive_second = client.post("/api/admin/events/archive")
        assert archive_second.status_code == 200
        assert archive_second.json()["created"] is False
        assert archive_second.json()["cleared_count"] == 0
        assert archive_second.json()["archives"]["total"] == 1


def test_list_event_archives_supports_backend_filter_and_pagination(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")
    archive_dir = tmp_path / "event_archives"
    archive_dir.mkdir(parents=True, exist_ok=True)

    for index in range(12):
        file_path = archive_dir / f"2026-03-{index + 1:02d} 10-00-00 a 2026-03-{index + 1:02d} 11-00-00.csv"
        file_path.write_text("id,event_time\n1,2026-03-01T10:00:00\n", encoding="utf-8-sig")

    with TestClient(app) as client:
        ensure_admin_session(client)
        paged = client.get("/api/admin/events/archives?page=2&page_size=5")
        assert paged.status_code == 200
        payload = paged.json()
        assert payload["total"] == 12
        assert payload["page"] == 2
        assert payload["page_size"] == 5
        assert payload["total_pages"] == 3
        assert payload["total_size_bytes"] > 0
        assert len(payload["items"]) == 5

        filtered = client.get("/api/admin/events/archives?q=2026-03-01&page=1&page_size=5")
        assert filtered.status_code == 200
        filtered_payload = filtered.json()
        assert filtered_payload["total"] >= 1
        assert filtered_payload["query"] == "2026-03-01"
        assert all("2026-03-01" in item["period"] for item in filtered_payload["items"])


def test_admin_login_session_and_logout_flow():
    with TestClient(app) as client:
        session_before = client.get("/api/admin/auth/session")
        assert session_before.status_code == 200
        assert session_before.json()["authenticated"] is False

        login_response = login_admin(client)
        assert login_response.status_code == 200
        assert login_response.json()["ok"] is True

        session_after = client.get("/api/admin/auth/session")
        assert session_after.status_code == 200
        assert session_after.json()["authenticated"] is True
        assert session_after.json()["admin"]["chave"] == "HR70"

        admins = client.get("/api/admin/administrators")
        assert admins.status_code == 200
        assert any(
            row["chave"] == "HR70" and row["status"] == "active" and row["perfil"] == 9
            for row in admins.json()
        )

        logout_response = client.post("/api/admin/auth/logout")
        assert logout_response.status_code == 200
        assert logout_response.json()["ok"] is True

        session_final = client.get("/api/admin/auth/session")
        assert session_final.status_code == 200
        assert session_final.json()["authenticated"] is False


def test_admin_request_access_and_approval_flow():
    with TestClient(app) as client:
        request_response = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "TS11",
                "nome_completo": "Teste Solicitante",
                "senha": "Senha@123",
            },
        )
        assert request_response.status_code == 200
        assert request_response.json()["ok"] is True

        duplicate_response = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "TS11",
                "nome_completo": "Teste Solicitante",
                "senha": "Senha@123",
            },
        )
        assert duplicate_response.status_code == 409

        login_response = login_admin(client)
        assert login_response.status_code == 200

        list_response = client.get("/api/admin/administrators")
        assert list_response.status_code == 200
        pending_row = next(row for row in list_response.json() if row["chave"] == "TS11")
        assert pending_row["status"] == "pending"

        approve_response = client.post(f"/api/admin/administrators/requests/{pending_row['id']}/approve")
        assert approve_response.status_code == 200
        assert approve_response.json()["ok"] is True

        relogin = client.post("/api/admin/auth/logout")
        assert relogin.status_code == 200

        new_admin_login = login_admin(client, chave="TS11", senha="Senha@123")
        assert new_admin_login.status_code == 200

        with SessionLocal() as db:
            admin = db.execute(select(User).where(User.chave == "TS11")).scalar_one_or_none()
            pending = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == "TS11")).scalar_one_or_none()
            assert admin is not None
            assert admin.perfil == 1
            assert admin.senha is not None
            assert pending is None


def test_admin_self_service_request_registers_unknown_user_and_allows_profile_override_on_approval():
    with TestClient(app) as client:
        status_response = client.get("/api/admin/auth/request-access/status", params={"chave": "NR11"})
        assert status_response.status_code == 200
        assert status_response.json()["found"] is False

        request_response = client.post(
            "/api/admin/auth/request-access/self-service",
            json={
                "chave": "NR11",
                "nome_completo": "Novo Requisitante",
                "projeto": "P80",
                "senha": "cad123",
                "confirmar_senha": "cad123",
            },
        )
        assert request_response.status_code == 200
        assert request_response.json()["ok"] is True

        with SessionLocal() as db:
            user = db.execute(select(User).where(User.chave == "NR11")).scalar_one_or_none()
            pending = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == "NR11")).scalar_one_or_none()
            assert user is not None
            assert user.perfil == 0
            assert user.projeto == "P80"
            assert user.senha is not None
            assert verify_password("cad123", user.senha) is True
            assert pending is not None
            assert pending.requested_profile == 1

        login_response = login_admin(client)
        assert login_response.status_code == 200

        rows_response = client.get("/api/admin/administrators")
        assert rows_response.status_code == 200
        pending_row = next(row for row in rows_response.json() if row["chave"] == "NR11")
        assert pending_row["row_type"] == "request"
        assert pending_row["perfil"] == 1

        approve_response = client.post(
            f"/api/admin/administrators/requests/{pending_row['id']}/approve",
            json={"perfil": 9},
        )
        assert approve_response.status_code == 200
        assert approve_response.json()["ok"] is True

        with SessionLocal() as db:
            user = db.execute(select(User).where(User.chave == "NR11")).scalar_one_or_none()
            pending = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == "NR11")).scalar_one_or_none()
            assert user is not None
            assert user.perfil == 9
            assert pending is None


def test_admin_self_service_request_uses_registered_transport_user_and_keeps_revoke_available():
    ensure_web_user_exists(chave="RK12", nome="Requisitante Conhecido")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "RK12")
        user.perfil = 2
        db.commit()

    with TestClient(app) as client:
        registration = register_web_password(client, chave="RK12", senha="rk1234", ensure_user_exists=False)
        assert registration.status_code == 200

        status_response = client.get("/api/admin/auth/request-access/status", params={"chave": "RK12"})
        assert status_response.status_code == 200
        assert status_response.json()["found"] is True
        assert status_response.json()["has_password"] is True
        assert status_response.json()["is_admin"] is False
        assert status_response.json()["has_pending_request"] is False

        request_response = client.post(
            "/api/admin/auth/request-access/self-service",
            json={"chave": "RK12"},
        )
        assert request_response.status_code == 200
        assert request_response.json()["ok"] is True

        login_response = login_admin(client)
        assert login_response.status_code == 200

        rows_response = client.get("/api/admin/administrators")
        assert rows_response.status_code == 200
        pending_row = next(row for row in rows_response.json() if row["chave"] == "RK12")
        assert pending_row["row_type"] == "request"

        approve_response = client.post(
            f"/api/admin/administrators/requests/{pending_row['id']}/approve",
            json={"perfil": 2},
        )
        assert approve_response.status_code == 200

        admins_response = client.get("/api/admin/administrators")
        assert admins_response.status_code == 200
        approved_row = next(
            row for row in admins_response.json() if row["chave"] == "RK12" and row["row_type"] == "admin"
        )
        assert approved_row["perfil"] == 12
        assert approved_row["can_revoke"] is True
        profile_update_response = client.post(
            f"/api/admin/administrators/{approved_row['id']}/profile",
            json={"perfil": 2},
        )
        assert profile_update_response.status_code == 200
        assert profile_update_response.json()["ok"] is True

        refreshed_admins_response = client.get("/api/admin/administrators")
        assert refreshed_admins_response.status_code == 200
        refreshed_row = next(
            row for row in refreshed_admins_response.json() if row["chave"] == "RK12" and row["row_type"] == "admin"
        )
        assert refreshed_row["perfil"] == 12
        assert refreshed_row["can_revoke"] is True

        with SessionLocal() as db:
            user = db.execute(select(User).where(User.chave == "RK12")).scalar_one_or_none()
            assert user is not None
            assert user.perfil == 12
            assert user.senha is not None
            assert verify_password("rk1234", user.senha) is True


def test_admin_request_access_frontend_e2e_handles_existing_and_unknown_keys():
    existing_key = make_test_key("E")
    unknown_key = make_test_key("N")
    existing_password = "ex1234"
    unknown_password = "cad123"
    existing_name = "Usuario Existente Front"
    unknown_name = "Usuario Novo Front"

    ensure_web_user_exists(chave=existing_key, projeto="P80", nome=existing_name)
    with TestClient(app) as client:
        register_response = register_web_password(
            client,
            chave=existing_key,
            senha=existing_password,
            projeto="P80",
            ensure_user_exists=False,
        )
        assert register_response.status_code == 200, register_response.text

    with live_app_server() as base_url:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            try:
                page.goto(f"{base_url}/admin", wait_until="domcontentloaded")
                page.locator("#requestAdminButton").wait_for(state="visible")

                page.get_by_role("button", name="Solicitar Administração").click()
                page.locator("#requestAdminModal").wait_for(state="visible")
                page.locator("#requestAdminChave").fill(unknown_key)
                page.locator("#requestAdminRegistrationModal").wait_for(state="visible")
                assert page.input_value("#requestAdminRegistrationChave") == unknown_key
                assert "Chave nao cadastrada" in (page.text_content("#requestAdminRegistrationStatus") or "")

                page.locator("#requestAdminRegistrationNome").fill(unknown_name)
                page.locator("#requestAdminRegistrationProjeto").select_option("P80")
                page.locator("#requestAdminRegistrationSenha").fill(unknown_password)
                page.locator("#requestAdminRegistrationConfirm").fill(unknown_password)
                page.locator("#requestAdminRegistrationSaveButton").click()
                page.wait_for_function(
                    "() => document.querySelector('#authStatus').textContent.includes('Solicitacao enviada')"
                )
                page.wait_for_function(
                    "() => document.querySelector('#requestAdminRegistrationModal').classList.contains('hidden')"
                )

                page.get_by_role("button", name="Solicitar Administração").click()
                page.locator("#requestAdminModal").wait_for(state="visible")
                page.locator("#requestAdminChave").fill(existing_key)
                page.wait_for_function(
                    "() => document.querySelector('#requestAdminStatus').textContent.includes('Solicitacao enviada')"
                )
                assert page.locator("#requestAdminRegistrationModal").is_hidden()
                page.wait_for_function(
                    "() => document.querySelector('#requestAdminModal').classList.contains('hidden')"
                )
            finally:
                browser.close()

    with SessionLocal() as db:
        existing_user = db.execute(select(User).where(User.chave == existing_key)).scalar_one_or_none()
        unknown_user = db.execute(select(User).where(User.chave == unknown_key)).scalar_one_or_none()
        existing_request = db.execute(
            select(AdminAccessRequest).where(AdminAccessRequest.chave == existing_key)
        ).scalar_one_or_none()
        unknown_request = db.execute(
            select(AdminAccessRequest).where(AdminAccessRequest.chave == unknown_key)
        ).scalar_one_or_none()

        assert existing_user is not None
        assert existing_user.nome == existing_name
        assert existing_user.perfil == 0
        assert existing_user.senha is not None
        assert verify_password(existing_password, existing_user.senha) is True

        assert unknown_user is not None
        assert unknown_user.nome == unknown_name
        assert unknown_user.projeto == "P80"
        assert unknown_user.perfil == 0
        assert unknown_user.senha is not None
        assert verify_password(unknown_password, unknown_user.senha) is True

        assert existing_request is not None
        assert existing_request.nome_completo == existing_name
        assert existing_request.requested_profile == 1

        assert unknown_request is not None
        assert unknown_request.nome_completo == unknown_name
        assert unknown_request.requested_profile == 1


def test_stale_checkin_rows_leave_checkin_and_move_to_inactive():
    ensure_web_user_exists(chave="CI11", nome="Checkin Antigo", projeto="P80")
    stale_time = now_sgt() - timedelta(days=1, minutes=5)
    set_user_checkin_state(chave="CI11", event_time=stale_time, local="Web")

    with TestClient(app) as client:
        ensure_admin_session(client)

        checkin_response = client.get("/api/admin/checkin")
        assert checkin_response.status_code == 200
        checkin_rows = checkin_response.json()
        assert all(row["chave"] != "CI11" for row in checkin_rows)

        inactive_response = client.get("/api/admin/inactive")
        assert inactive_response.status_code == 200
        inactive_row = next(row for row in inactive_response.json() if row["chave"] == "CI11")
        assert inactive_row["latest_action"] == "checkin"
        assert inactive_row["inactivity_days"] >= 1


def test_admin_login_rejects_transport_only_profile():
    ensure_web_user_exists(chave="TT20", nome="Transport Only")
    with TestClient(app) as client:
        registration = register_web_password(client, chave="TT20", senha="tt2024", ensure_user_exists=False)
        assert registration.status_code == 200

    with SessionLocal() as db:
        transport_only_user = get_user_by_chave(db, "TT20")
        transport_only_user.perfil = 2
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="TT20", senha="tt2024")
        assert login_response.status_code == 403


def test_administrators_endpoint_lists_admin_profiles_only():
    ensure_web_user_exists(chave="UTO9", nome="Admin Perfil")
    ensure_web_user_exists(chave="TP22", nome="Transport Perfil")
    ensure_web_user_exists(chave="ZZ00", nome="Sem Perfil")

    with SessionLocal() as db:
        get_user_by_chave(db, "UTO9").perfil = 1
        get_user_by_chave(db, "TP22").perfil = 2
        get_user_by_chave(db, "ZZ00").perfil = 0
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        rows = client.get("/api/admin/administrators")
        assert rows.status_code == 200
        rows_by_key = {row["chave"]: row for row in rows.json() if row["row_type"] == "admin"}
        assert rows_by_key["UTO9"]["perfil"] == 1
        assert rows_by_key["UTO9"]["can_revoke"] is True
        assert "TP22" not in rows_by_key
        assert "ZZ00" not in rows_by_key


def test_admin_password_reset_and_redefine_flow():
    with TestClient(app) as client:
        login_response = login_admin(client)
        assert login_response.status_code == 200

        request_new_admin = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "AB12",
                "nome_completo": "Admin Auxiliar",
                "senha": "SenhaNova1",
            },
        )
        assert request_new_admin.status_code == 200

        admins_before = client.get("/api/admin/administrators")
        pending_row = next(row for row in admins_before.json() if row["chave"] == "AB12")
        approve_response = client.post(f"/api/admin/administrators/requests/{pending_row['id']}/approve")
        assert approve_response.status_code == 200

        reset_response = client.post("/api/admin/auth/request-password-reset", json={"chave": "AB12"})
        assert reset_response.status_code == 200
        assert "Outro administrador" in reset_response.json()["message"]

        blocked_login = login_admin(client, chave="AB12", senha="SenhaNova1")
        assert blocked_login.status_code == 403

        admins_after_reset = client.get("/api/admin/administrators")
        reset_row = next(row for row in admins_after_reset.json() if row["chave"] == "AB12")
        assert reset_row["status"] == "password_reset_requested"

        set_password_response = client.post(
            f"/api/admin/administrators/{reset_row['id']}/set-password",
            json={"nova_senha": "SenhaFinal2"},
        )
        assert set_password_response.status_code == 200

        client.post("/api/admin/auth/logout")
        relogin_new_password = login_admin(client, chave="AB12", senha="SenhaFinal2")
        assert relogin_new_password.status_code == 200


def test_admin_self_service_password_change_flow():
    with TestClient(app) as client:
        assert login_admin(client).status_code == 200

        request_new_admin = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "SC12",
                "nome_completo": "Senha Própria",
                "senha": "SenhaIni1",
            },
        )
        assert request_new_admin.status_code == 200

        admins_before = client.get("/api/admin/administrators")
        pending_row = next(row for row in admins_before.json() if row["chave"] == "SC12")
        approve_response = client.post(f"/api/admin/administrators/requests/{pending_row['id']}/approve")
        assert approve_response.status_code == 200

        verify_ok = client.post(
            "/api/admin/auth/verify-current-password",
            json={"chave": "SC12", "senha_atual": "SenhaIni1"},
        )
        assert verify_ok.status_code == 200
        assert verify_ok.json()["valid"] is True

        verify_invalid = client.post(
            "/api/admin/auth/verify-current-password",
            json={"chave": "SC12", "senha_atual": "SenhaErr1"},
        )
        assert verify_invalid.status_code == 200
        assert verify_invalid.json()["valid"] is False

        invalid_change = client.post(
            "/api/admin/auth/change-password",
            json={
                "chave": "SC12",
                "senha_atual": "SenhaErr1",
                "nova_senha": "SenhaSC2",
                "confirmar_senha": "SenhaSC2",
            },
        )
        assert invalid_change.status_code == 401

        change_response = client.post(
            "/api/admin/auth/change-password",
            json={
                "chave": "SC12",
                "senha_atual": "SenhaIni1",
                "nova_senha": "SenhaSC2",
                "confirmar_senha": "SenhaSC2",
            },
        )
        assert change_response.status_code == 200
        assert "Senha alterada com sucesso." in change_response.json()["message"]

        client.post("/api/admin/auth/logout")
        blocked_old_login = login_admin(client, chave="SC12", senha="SenhaIni1")
        assert blocked_old_login.status_code == 401

        relogin_new_password = login_admin(client, chave="SC12", senha="SenhaSC2")
        assert relogin_new_password.status_code == 200

        events_response = client.get("/api/admin/events")
        assert events_response.status_code == 200
        events = events_response.json()
        assert any(
            event["action"] == "password"
            and event["status"] == "updated"
            and event["request_path"] == "/api/admin/auth/change-password"
            and "chave=SC12" in (event["details"] or "")
            for event in events
        )


def test_admin_event_audit_covers_new_auth_lifecycle():
    with TestClient(app) as client:
        first_request = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "EV12",
                "nome_completo": "Evento Auditoria",
                "senha": "SenhaEvt1",
            },
        )
        assert first_request.status_code == 200

        duplicate_request = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "EV12",
                "nome_completo": "Evento Auditoria",
                "senha": "SenhaEvt1",
            },
        )
        assert duplicate_request.status_code == 409

        assert login_admin(client).status_code == 200

        administrators = client.get("/api/admin/administrators")
        pending_row = next(row for row in administrators.json() if row["chave"] == "EV12")

        approve_response = client.post(f"/api/admin/administrators/requests/{pending_row['id']}/approve")
        assert approve_response.status_code == 200

        first_reset = client.post("/api/admin/auth/request-password-reset", json={"chave": "EV12"})
        assert first_reset.status_code == 200

        duplicate_reset = client.post("/api/admin/auth/request-password-reset", json={"chave": "EV12"})
        assert duplicate_reset.status_code == 409

        blocked_login = login_admin(client, chave="EV12", senha="SenhaEvt1")
        assert blocked_login.status_code == 403

        administrators_after_reset = client.get("/api/admin/administrators")
        reset_row = next(row for row in administrators_after_reset.json() if row["chave"] == "EV12")

        set_password = client.post(
            f"/api/admin/administrators/{reset_row['id']}/set-password",
            json={"nova_senha": "SenhaEvt2"},
        )
        assert set_password.status_code == 200

        revoke_response = client.post(f"/api/admin/administrators/{reset_row['id']}/revoke")
        assert revoke_response.status_code == 200

        events_response = client.get("/api/admin/events")
        assert events_response.status_code == 200
        events = events_response.json()

        assert any(
            event["action"] == "admin_request"
            and event["status"] == "pending"
            and event["request_path"] == "/api/admin/auth/request-access"
            and "chave=EV12" in (event["details"] or "")
            for event in events
        )
        assert any(
            event["action"] == "admin_request"
            and event["status"] == "failed"
            and event["http_status"] == 409
            and event["request_path"] == "/api/admin/auth/request-access"
            and "chave=EV12" in (event["details"] or "")
            for event in events
        )
        assert any(
            event["action"] == "admin_request"
            and event["status"] == "approved"
            and event["request_path"] == f"/api/admin/administrators/requests/{pending_row['id']}/approve"
            for event in events
        )
        assert any(
            event["action"] == "password"
            and event["status"] == "pending"
            and event["request_path"] == "/api/admin/auth/request-password-reset"
            for event in events
        )
        assert any(
            event["action"] == "password"
            and event["status"] == "failed"
            and event["http_status"] == 409
            and event["request_path"] == "/api/admin/auth/request-password-reset"
            for event in events
        )
        assert any(
            event["action"] == "login"
            and event["status"] == "blocked"
            and event["request_path"] == "/api/admin/auth/login"
            and "chave=EV12" in (event["details"] or "")
            for event in events
        )
        assert any(
            event["action"] == "password"
            and event["status"] == "updated"
            and event["request_path"] == f"/api/admin/administrators/{reset_row['id']}/set-password"
            for event in events
        )
        assert any(
            event["action"] == "admin_access"
            and event["status"] == "removed"
            and event["request_path"] == f"/api/admin/administrators/{reset_row['id']}/revoke"
            for event in events
        )


def test_transport_inline_auth_respects_user_profile():
    ensure_web_user_exists(chave="TRP2", nome="Transport Access")
    ensure_web_user_exists(chave="AD11", nome="Admin Only")
    with TestClient(app) as client:
        registration = register_web_password(client, chave="TRP2", senha="tp1234", ensure_user_exists=False)
        assert registration.status_code == 200
        admin_registration = register_web_password(client, chave="AD11", senha="ad1111", ensure_user_exists=False)
        assert admin_registration.status_code == 200

    with SessionLocal() as db:
        transport_user = get_user_by_chave(db, "TRP2")
        transport_user.perfil = 2
        admin_only_user = get_user_by_chave(db, "AD11")
        admin_only_user.perfil = 1
        db.commit()

    with TestClient(app) as client:
        denied = client.post("/api/transport/auth/verify", json={"chave": "AD11", "senha": "ad1111"})
        assert denied.status_code == 200
        assert denied.json()["authenticated"] is False
        assert "transport access" in denied.json()["message"].lower()

        granted = client.post("/api/transport/auth/verify", json={"chave": "TRP2", "senha": "tp1234"})
        assert granted.status_code == 200
        assert granted.json()["authenticated"] is True
        assert granted.json()["user"]["perfil"] == 2

        dashboard = client.get("/api/transport/dashboard")
        assert dashboard.status_code == 200

        logout = client.post("/api/transport/auth/logout")
        assert logout.status_code == 200

        dashboard_after_logout = client.get("/api/transport/dashboard")
        assert dashboard_after_logout.status_code == 401


def test_admin_locations_crud_and_mobile_catalog_sync():
    with TestClient(app) as client:
        ensure_admin_session(client)
        reset_location_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 30},
        )
        assert reset_location_settings.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Base P80",
                "latitude": 1.255936,
                "longitude": 103.611066,
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200
        assert create_location.json()["ok"] is True

        locations = client.get("/api/admin/locations")
        assert locations.status_code == 200
        assert locations.json()["location_accuracy_threshold_meters"] == 30
        base_p80 = next(row for row in locations.json()["items"] if row["local"] == "Base P80")
        assert base_p80["coordinates"] == [{"latitude": 1.255936, "longitude": 103.611066}]
        assert base_p80["tolerance_meters"] == 150

        update_location_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 45},
        )
        assert update_location_settings.status_code == 200
        assert update_location_settings.json()["ok"] is True
        assert update_location_settings.json()["location_accuracy_threshold_meters"] == 45

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        location_settings_event = next(
            event
            for event in events.json()
            if event["action"] == "location_config" and event["request_path"] == "/api/admin/locations/settings"
        )
        assert location_settings_event["chave"] == ADMIN_LOGIN_CHAVE
        assert location_settings_event["message"] == (
            "O valor do erro máximo para considerar a coordenada do usuário foi ajustado para 45 metros."
        )

        update_location = client.post(
            "/api/admin/locations",
            json={
                "location_id": base_p80["id"],
                "local": "Base P80",
                "coordinates": [
                    {"latitude": 1.255936, "longitude": 103.611066},
                    {"latitude": 1.260001, "longitude": 103.612002},
                ],
                "tolerance_meters": 250,
            },
        )
        assert update_location.status_code == 200
        assert update_location.json()["ok"] is True

        updated_locations = client.get("/api/admin/locations")
        assert updated_locations.status_code == 200
        updated_base_p80 = next(row for row in updated_locations.json()["items"] if row["local"] == "Base P80")
        assert updated_base_p80["coordinates"] == [
            {"latitude": 1.255936, "longitude": 103.611066},
            {"latitude": 1.260001, "longitude": 103.612002},
        ]
        assert updated_base_p80["latitude"] == 1.255936
        assert updated_base_p80["longitude"] == 103.611066

        mobile_catalog = client.get("/api/mobile/locations", headers=MOBILE_HEADERS)
        assert mobile_catalog.status_code == 200
        assert mobile_catalog.json()["location_accuracy_threshold_meters"] == 45
        assert "coordinate_update_frequency_headers" not in mobile_catalog.json()
        assert "coordinate_update_frequency_rows" not in mobile_catalog.json()
        synced_row = next(row for row in mobile_catalog.json()["items"] if row["local"] == "Base P80")
        assert synced_row["tolerance_meters"] == 250
        assert synced_row["coordinates"] == [
            {"latitude": 1.255936, "longitude": 103.611066},
            {"latitude": 1.260001, "longitude": 103.612002},
        ]
        assert synced_row["latitude"] == 1.255936
        assert synced_row["longitude"] == 103.611066

        remove_location = client.delete(f"/api/admin/locations/{base_p80['id']}")
        assert remove_location.status_code == 200
        assert remove_location.json()["ok"] is True


def test_mobile_forms_submit_uses_default_and_custom_local():
    with TestClient(app) as client:
        first_event = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "LX01",
                "projeto": "P83",
                "action": "checkin",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"mobile-{uuid.uuid4().hex}",
            },
        )
        assert first_event.status_code == 200
        assert first_event.json()["ok"] is True

        with SessionLocal() as db:
            user = get_user_by_chave(db, "LX01")
            queued = db.execute(
                select(FormsSubmission)
                .where(FormsSubmission.chave == "LX01")
                .order_by(FormsSubmission.id.desc())
            ).scalars().first()
            sync_event = db.execute(
                select(UserSyncEvent)
                .where(UserSyncEvent.chave == "LX01")
                .order_by(UserSyncEvent.id.desc())
            ).scalars().first()

            assert user.local == "Aplicativo"
            assert queued is not None and queued.local == "Aplicativo"
            assert sync_event is not None and sync_event.local == "Aplicativo"

        second_event = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "LX01",
                "projeto": "P83",
                "action": "checkout",
                "local": "Base P80",
                "informe": "retroativo",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"mobile-{uuid.uuid4().hex}",
            },
        )
        assert second_event.status_code == 200
        assert second_event.json()["ok"] is True

        with SessionLocal() as db:
            user = get_user_by_chave(db, "LX01")
            queued = db.execute(
                select(FormsSubmission)
                .where(FormsSubmission.chave == "LX01")
                .order_by(FormsSubmission.id.desc())
            ).scalars().first()
            sync_event = db.execute(
                select(UserSyncEvent)
                .where(UserSyncEvent.chave == "LX01")
                .order_by(UserSyncEvent.id.desc())
            ).scalars().first()

            assert user.local == "Base P80"
            assert queued is not None and queued.local == "Base P80"
            assert sync_event is not None and sync_event.local == "Base P80"
